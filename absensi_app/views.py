import json
import csv
from datetime import datetime, timedelta, date
from itertools import groupby
from operator import attrgetter
import logging  
import concurrent.futures
import threading

# 2. IMPORT DJANGO
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.cache import cache
from django.db import transaction
from django.db.models import Q, Count, Sum, Avg
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

# 3. SETUP LOGGER 
logger = logging.getLogger(__name__)

# 4. IMPORT LOKAL
from .services import WorkModeService
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ZK Fingerprint Library
try:
    from zk import ZK, const
except ImportError:
    ZK = None
    const = None

from .models import (
    Pegawai, Absensi, FingerprintTemplate,
    MasterDepartemen, MasterJabatan, MasterCabang,
    MasterMesin, MasterModeJamKerja, ModeJamKerjaJadwal,
    PegawaiModeAssignment, ModeJamKerjaPeriode 
)
from .forms import (
    PegawaiForm, AbsensiAdminForm, LaporanFilterForm,
    PegawaiSearchForm, MasterDepartemenForm, MasterJabatanForm,
    MasterCabangForm, MasterMesinForm
)
from .services import WorkModeService

## FUNGSI BANTUAN UMUM
# Kode ini berisi fungsi-fungsi kecil untuk pengecekan user, pengambilan data master, dan pemfilteran.

def is_staff_or_superuser(user):
    """Cek apakah user adalah staf atau superuser."""
    return user.is_staff or user.is_superuser


def get_active_machines():
    """Ambil semua data mesin yang aktif."""
    return MasterMesin.objects.filter(is_active=True).select_related('cabang')


def get_machine_by_id(mesin_id):
    """Ambil data mesin berdasarkan ID."""
    return get_object_or_404(MasterMesin, id=mesin_id, is_active=True)


def get_pegawai_with_fingerprint():
    """Ambil ID pegawai yang memiliki template sidik jari."""
    try:
        return Pegawai.objects.filter(
            fingerprint_templates__isnull=False
        ).values_list('id', flat=True).distinct()
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Error getting pegawai with fingerprint: {str(e)}")
        return []

def _show_form_errors(request, form):
    """Menampilkan pesan error dari form."""
    for field, errors in form.errors.items():
        for error in errors:
            messages.error(request, f"{field}: {error}")


## FUNGSI BANTUAN KHUSUS CABANG
# Kode ini untuk mengelola Cabang Aktif dari sesi user.

def get_active_cabang(request):
    """
    ✅ FIXED: Ambil data cabang aktif dari session atau set cabang pertama jika belum ada.
    
    Key Fixes:
    1. ✅ Proper session object validation
    2. ✅ Force save session after write
    3. ✅ Proper error handling
    4. ✅ Return None jika tidak ada cabang aktif
    """
    try:
        # ========================================
        # 1. VALIDATE REQUEST & USER
        # ========================================
        if not request.user.is_authenticated:
            return None
        
        if not (request.user.is_staff or request.user.is_superuser):
            return None
        
        # ========================================
        # 2. VALIDATE SESSION OBJECT
        # ========================================
        if not hasattr(request, 'session'):
            logger.warning("❌ Request tidak memiliki session")
            return None
        
        # ✅ FIX: Check session is SessionStore, not dict
        if not hasattr(request.session, 'get'):
            logger.warning("❌ Session bukan SessionStore object")
            return None
        
        # ========================================
        # 3. GET CABANG ID DARI SESSION
        # ========================================
        cabang_id = request.session.get('cabang_aktif_id')
        
        if cabang_id:
            try:
                # ✅ Try convert ke int
                cabang_id = int(cabang_id)
                
                # ✅ Get cabang
                cabang_aktif = MasterCabang.objects.get(
                    id=cabang_id,
                    is_active=True
                )
                
                logger.info(f"✅ Cabang aktif dari session: {cabang_aktif.nama}")
                return cabang_aktif
                
            except (MasterCabang.DoesNotExist, ValueError, TypeError) as e:
                # ========================================
                # 4. CLEANUP INVALID CABANG
                # ========================================
                logger.warning(f"⚠️ Cabang ID {cabang_id} tidak valid: {str(e)}")
                
                try:
                    request.session.pop('cabang_aktif_id', None)
                    request.session.pop('cabang_aktif_nama', None)
                    request.session.modified = True
                    request.session.save()  # ✅ FORCE SAVE
                    
                    logger.info("✅ Cleaned invalid cabang session")
                except Exception as cleanup_error:
                    logger.error(f"❌ Error cleaning session: {str(cleanup_error)}")
        
        # ========================================
        # 5. SET CABANG PERTAMA JIKA BELUM ADA
        # ========================================
        first_cabang = MasterCabang.objects.filter(is_active=True).first()
        
        if first_cabang:
            try:
                request.session['cabang_aktif_id'] = first_cabang.id
                request.session['cabang_aktif_nama'] = first_cabang.nama
                request.session.modified = True
                request.session.save()  # ✅ FORCE SAVE
                
                logger.info(f"✅ Set cabang pertama sebagai aktif: {first_cabang.nama}")
                return first_cabang
                
            except Exception as e:
                logger.error(f"❌ Error setting first cabang: {str(e)}")
                return None
        
        # ========================================
        # 6. TIDAK ADA CABANG AKTIF
        # ========================================
        logger.warning("⚠️ Tidak ada cabang aktif")
        return None
    
    except Exception as e:
        logger.error(f"❌ Critical error in get_active_cabang: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return None
    
def filter_by_cabang(queryset, cabang, field='cabang'):
    """
    Memfilter queryset berdasarkan objek cabang yang aktif.
    
     Handle None cabang dan invalid field gracefully
    """
    if not cabang:
        return queryset
    
    try:
        # Build filter kwargs
        filter_kwargs = {field: cabang}
        return queryset.filter(filter_kwargs)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Error in filter_by_cabang: {str(e)}")
        return queryset

@login_required
def switch_cabang(request):
    """
    ✅ FIXED: Endpoint API untuk mengganti cabang aktif pada sesi user.
    
    Key Fixes:
    1. ✅ Proper session object handling
    2. ✅ Force save + commit to database
    3. ✅ Validate cabang exists & is_active
    4. ✅ Proper error response
    5. ✅ Return updated cabang data
    """
    
    # ========================================
    # 1. VALIDATE REQUEST
    # ========================================
    if not request.user.is_staff:
        return JsonResponse({
            "status": "error", 
            "msg": "Akses ditolak"
        }, status=403)
    
    if request.method != 'POST':
        return JsonResponse({
            "status": "error", 
            "msg": "Method tidak diizinkan"
        }, status=405)
    
    try:
        # ========================================
        # 2. GET CABANG ID
        # ========================================
        cabang_id = request.POST.get('cabang_id', '').strip()
        
        if not cabang_id:
            return JsonResponse({
                "status": "error",
                "msg": "Cabang ID wajib diisi"
            }, status=400)
        
        # ========================================
        # 3. VALIDATE CABANG EXISTS
        # ========================================
        try:
            cabang_id_int = int(cabang_id)
        except (ValueError, TypeError):
            return JsonResponse({
                "status": "error",
                "msg": "Cabang ID harus berupa angka"
            }, status=400)
        
        cabang = MasterCabang.objects.filter(
            id=cabang_id_int,
            is_active=True
        ).first()
        
        if not cabang:
            return JsonResponse({
                "status": "error",
                "msg": f"Cabang ID {cabang_id} tidak ditemukan atau tidak aktif"
            }, status=404)
        
        # ========================================
        # 4. VALIDATE SESSION OBJECT
        # ========================================
        if not hasattr(request, 'session'):
            return JsonResponse({
                "status": "error",
                "msg": "Session tidak tersedia"
            }, status=400)
        
        if not hasattr(request.session, 'get'):
            return JsonResponse({
                "status": "error",
                "msg": "Session object tidak valid"
            }, status=400)
        
        # ========================================
        # 5. SAVE CABANG KE SESSION (PERSISTEN 30 HARI)
        # ========================================
        try:
            request.session['cabang_aktif_id'] = cabang.id
            request.session['cabang_aktif_nama'] = cabang.nama
            request.session.modified = True
            
            # ✅ CRITICAL: Force save to database
            request.session.save()
            
            logger.info(
                f"✅ User {request.user.username} switched cabang to {cabang.nama} (ID: {cabang.id})"
            )
            
        except Exception as save_error:
            logger.error(f"❌ Error saving session: {str(save_error)}")
            
            return JsonResponse({
                "status": "error",
                "msg": f"Gagal menyimpan ke session: {str(save_error)}"
            }, status=500)
        
        # ========================================
        # 6. HITUNG STATISTIK CABANG
        # ========================================
        try:
            total_pegawai = Pegawai.objects.filter(
                cabang=cabang,
                is_active=True
            ).count()
            
            total_mesin = MasterMesin.objects.filter(
                cabang=cabang,
                is_active=True
            ).count()
            
        except Exception as stats_error:
            logger.warning(f"⚠️ Error calculating stats: {str(stats_error)}")
            total_pegawai = 0
            total_mesin = 0
        
        # ========================================
        # 7. RETURN SUCCESS RESPONSE
        # ========================================
        return JsonResponse({
            "status": "success",
            "msg": f"✅ Berhasil beralih ke {cabang.nama}",
            "cabang": {
                "id": cabang.id,
                "nama": cabang.nama,
                "kode": cabang.kode,
                "alamat": cabang.alamat or "-",
                "total_pegawai": total_pegawai,
                "total_mesin": total_mesin,
            },
            "session": {
                "cabang_aktif_id": request.session.get('cabang_aktif_id'),
                "cabang_aktif_nama": request.session.get('cabang_aktif_nama'),
            }
        })
    
    except Exception as e:
        logger.error(f"❌ Critical error in switch_cabang: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        
        return JsonResponse({
            "status": "error",
            "msg": f"Server error: {str(e)}"
        }, status=500)    

## FUNGSI INTEGRASI MESIN FINGERPRINT
# Kode ini untuk menghubungkan, mengambil data, dan mendaftarkan user ke mesin absensi ZKTeco.

def connect_to_fingerprint_machine(ip_address, port=4370, timeout=10):
    """Menghubungkan ke mesin sidik jari ZKTeco."""
    if ZK is None:
        raise Exception("Library pyzk belum terinstal.")
    
    try:
        zk = ZK(ip_address, port=port, timeout=timeout)
        conn = zk.connect()
        return conn
    except Exception as e:
        raise Exception(f"Koneksi ke {ip_address} gagal: {str(e)}")


def get_available_uid_from_machine(conn):
    """Mencari UID yang tersedia di mesin."""
    try:
        users = conn.get_users()
        if not users:
            return 1
        
        uids = sorted([int(user.uid) for user in users])
        
        for i in range(len(uids) - 1):
            if uids[i + 1] - uids[i] > 1:
                return uids[i] + 1
        
        return uids[-1] + 1 if uids else 1
    except Exception:
        return 1


def register_Pegawai_to_machine(conn, userid, nama_lengkap, password=None):
    """Mendaftarkan/memperbarui data pegawai ke mesin sidik jari."""
    try:
        conn.disable_device()
        
        userid = str(userid).strip()
        nama_display = nama_lengkap[:24]
        users = conn.get_users()
        
        existing_user = next(
            (u for u in users if str(getattr(u, 'user_id', u.uid)) == str(userid)),
            None
        )
        
        if existing_user:
            uid = existing_user.uid
            # Update user yang sudah ada
            try:
                conn.set_user(
                    uid=uid,
                    name=nama_display,
                    privilege=const.USER_DEFAULT if const else 0,
                    password=str(password) if password else '',
                    user_id=str(userid),
                    card=0,
                    group_id=''
                )
            except AttributeError:
                from zk.user import User
                user = User(
                    uid=uid,
                    name=nama_display,
                    privilege=0,
                    password=str(password) if password else '',
                    user_id=str(userid)
                )
                conn.set_user(user)
        else:
            uid = get_available_uid_from_machine(conn)
            # Buat user baru
            try:
                conn.set_user(
                    uid=uid,
                    name=nama_display,
                    privilege=const.USER_DEFAULT if const else 0,
                    password=str(password) if password else '',
                    user_id=str(userid),
                    card=0,
                    group_id=''
                )
            except AttributeError:
                from zk.user import User
                user = User(
                    uid=uid,
                    name=nama_display,
                    privilege=0,
                    password=str(password) if password else '',
                    user_id=str(userid)
                )
                conn.set_user(user)
        
        conn.enable_device()
        return uid
    except Exception as e:
        try:
            conn.enable_device()
        except:
            pass
        raise Exception(f"Pendaftaran gagal: {str(e)}")

def sync_fingerprint_template_from_machine(conn, userid):
    """Sinkronisasi template sidik jari dari mesin."""
    try:
        templates = conn.get_templates()
        users = conn.get_users()
        
        target_user = next(
            (u for u in users if str(getattr(u, 'user_id', u.uid)) == str(userid)),
            None
        )
        
        if not target_user:
            return []
        
        user_templates = [
            {
                'uid': t.uid,
                'fid': t.fid,
                'size': t.size,
                'valid': t.valid,
                'template': t.template
            }
            for t in templates if t.uid == target_user.uid
        ]
        
        return user_templates
    except Exception:
        return []


# def _sync_fingerprint_for_pegawai(pegawai, mesin):
#     """Helper untuk menyimpan template sidik jari ke database."""
#     fingers_count = 0
#     try:
#         conn = connect_to_fingerprint_machine(mesin.ip_address, mesin.port)
#         fingers = sync_fingerprint_template_from_machine(conn, pegawai.userid)
        
#         for finger in fingers:
#             FingerprintTemplate.objects.create(
#                 pegawai=pegawai,
#                 uid=finger['uid'],
#                 fid=finger['fid'],
#                 size=finger['size'],
#                 valid=finger['valid'],
#                 template=finger['template']
#             )
#             fingers_count += 1
        
#         conn.disconnect()
#     except Exception as e:
#         print(f"Peringatan sinkronisasi sidik jari: {str(e)}")
    
#     return fingers_count


## FUNGSI OTENTIKASI & OTORISASI
# Kode ini untuk mengelola proses login dan logout user admin.

def user_login(request):
    """Menangani proses login user."""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user:
            if user.is_staff or user.is_superuser:
                login(request, user)
                messages.success(
                    request,
                    f"Selamat datang, {user.get_full_name() or username}"
                )
                return redirect('dashboard')
            else:
                messages.error(request, "Akses ditolak. Hanya untuk Admin.")
        else:
            messages.error(request, "Username atau password salah.")
    
    return render(request, 'absensi_app/login.html')


@login_required
def user_logout(request):
    """Menangani proses logout user."""
    logout(request)
    messages.info(request, "Anda telah logout.")
    return redirect('user_login')


## FUNGSI DASHBOARD & STATISTIK
# Kode ini untuk menampilkan data ringkasan dan statistik absensi di halaman utama.

@login_required
def dashboard(request):
    """Menampilkan halaman dashboard dengan filter cabang."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('user_login')
    
    # ========================================
    #  1: Get cabang aktif dengan defensive checks
    # ========================================
    cabang_aktif = None
    try:
        cabang_aktif = get_active_cabang(request)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Error getting active cabang: {str(e)}")
    
    # ========================================
    #  2: Get pegawai dengan fingerprint
    # ========================================
    pegawai_with_fp = []
    try:
        pegawai_with_fp = list(get_pegawai_with_fingerprint())
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Error getting pegawai with fingerprint: {str(e)}")
    
    # ========================================
    #  3: Query pegawai dengan proper error handling
    # ========================================
    total_pegawai = 0
    try:
        pegawai_query = Pegawai.objects.filter(
            is_active=True,
            uid_mesin__isnull=False
        ).exclude(uid_mesin=0)
        
        # Filter by fingerprint (only if we have data)
        if pegawai_with_fp:
            pegawai_query = pegawai_query.filter(id__in=pegawai_with_fp)
        
        # Filter by cabang (only if cabang_aktif exists)
        if cabang_aktif:
            pegawai_query = pegawai_query.filter(cabang=cabang_aktif)
        
        total_pegawai = pegawai_query.count()
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Error counting pegawai: {str(e)}")
        total_pegawai = 0
    
    # ========================================
    #  4: Get today's attendance with error handling
    # ========================================
    today = datetime.now().date()
    hadir_count = 0
    sakit_count = 0
    izin_count = 0
    absen_count = 0
    pegawai_hadir = 0
    
    try:
        today_absensi = Absensi.objects.filter(tanggal=today)
        
        if cabang_aktif:
            today_absensi = today_absensi.filter(pegawai__cabang=cabang_aktif)
        
        hadir_count = today_absensi.filter(status='Hadir').count()
        sakit_count = today_absensi.filter(status='Sakit').count()
        izin_count = today_absensi.filter(status='Izin').count()
        absen_count = today_absensi.filter(status='Absen').count()
        pegawai_hadir = hadir_count + sakit_count + izin_count
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Error getting today attendance: {str(e)}")
    
    # ========================================
    #  5: Get chart data with error handling
    # ========================================
    period = request.GET.get('period', 'bulan')
    chart_data = {'title': 'Statistik Absensi', 'labels': [], 'hadir': [], 'sakit': [], 'izin': [], 'absen': []}
    
    try:
        chart_data = _get_chart_data(period, cabang_aktif)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Error getting chart data: {str(e)}")
    
    # ========================================
    #  6: Build context
    # ========================================
    context = {
        'total_pegawai': total_pegawai,
        'pegawai_hadir': pegawai_hadir,
        'hadir_count': hadir_count,
        'sakit_count': sakit_count,
        'izin_count': izin_count,
        'absen_count': absen_count,
        'chart_title': chart_data.get('title', 'Statistik Absensi'),
        'chart_labels': json.dumps(chart_data.get('labels', [])),
        'chart_data_hadir': json.dumps(chart_data.get('hadir', [])),
        'chart_data_sakit': json.dumps(chart_data.get('sakit', [])),
        'chart_data_izin': json.dumps(chart_data.get('izin', [])),
        'chart_data_absen': json.dumps(chart_data.get('absen', [])),
        'current_period': period,
        'today_date': today.strftime('%Y-%m-%d'),
        'cabang_aktif': cabang_aktif,
    }
    
    return render(request, "absensi_app/dashboard.html", context)

def _get_chart_data(period, cabang=None):
    """Helper untuk membuat data chart statistik absensi."""
    today = timezone.now().date()
    labels = []
    hadir = []
    sakit = []
    izin = []
    absen = []
    
    # ========================================
    #  Wrap query dalam try-except
    # ========================================
    try:
        base_query = Absensi.objects.all()
        
        if cabang:
            base_query = base_query.filter(pegawai__cabang=cabang)
        
        if period == 'tahun':
            title = "Statistik 5 Tahun"
            for i in range(5):
                year = today.year - i
                labels.insert(0, str(year))
                year_data = base_query.filter(tanggal__year=year)
                hadir.insert(0, year_data.filter(status='Hadir').count())
                sakit.insert(0, year_data.filter(status='Sakit').count())
                izin.insert(0, year_data.filter(status='Izin').count())
                absen.insert(0, year_data.filter(status='Absen').count())
        
        elif period == 'hari':
            title = "Statistik 7 Hari"
            dates = [today - timedelta(days=i) for i in range(6, -1, -1)]
            for date in dates:
                labels.append(date.strftime("%d %b"))
                day_data = base_query.filter(tanggal=date)
                hadir.append(day_data.filter(status='Hadir').count())
                sakit.append(day_data.filter(status='Sakit').count())
                izin.append(day_data.filter(status='Izin').count())
                absen.append(day_data.filter(status='Absen').count())
        
        else:  
            title = "Statistik 6 Bulan"
            for i in range(6):
                date = today - timedelta(days=30 * (5 - i))
                labels.append(date.strftime("%b %Y"))
                month_data = base_query.filter(
                    tanggal__year=date.year,
                    tanggal__month=date.month
                )
                hadir.append(month_data.filter(status='Hadir').count())
                sakit.append(month_data.filter(status='Sakit').count())
                izin.append(month_data.filter(status='Izin').count())
                absen.append(month_data.filter(status='Absen').count())
    
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in _get_chart_data: {str(e)}")
        title = "Statistik Absensi"
    
    return {
        'title': title,
        'labels': labels,
        'hadir': hadir,
        'sakit': sakit,
        'izin': izin,
        'absen': absen
    }

@login_required
def statistik_absensi(request):
    """Endpoint AJAX: Mengambil data chart absensi berdasarkan periode."""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Akses ditolak'}, status=403)
    
    period = request.GET.get('period', 'bulan')
    chart_data = _get_chart_data(period)
    
    return JsonResponse({
        "labels": chart_data['labels'],
        "hadir_data": chart_data['hadir'],
        "sakit_data": chart_data['sakit'],
        "izin_data": chart_data['izin'],
        "absen_data": chart_data['absen'],
    })


@login_required
def detail_absensi_by_status(request):
    """Endpoint AJAX: Mengambil detail absensi hari ini berdasarkan status."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    cabang_aktif = get_active_cabang(request)
    status = request.GET.get('status', '')
    today = datetime.now().date()
    
    if status not in ['Hadir', 'Sakit', 'Izin', 'Absen']:
        return JsonResponse(
            {"status": "error", "msg": "Status tidak valid"},
            status=400
        )
    
    try:
        absensi_list = Absensi.objects.filter(
            tanggal=today,
            status=status
        ).select_related(
            'pegawai',
            'pegawai__departemen',
            'pegawai__jabatan'
        )
        
        if cabang_aktif:
            absensi_list = absensi_list.filter(pegawai__cabang=cabang_aktif)
        
        absensi_list = absensi_list.order_by('pegawai__nama_lengkap')
        
        pegawai_data = []
        for absensi in absensi_list:
            # Hitung total jam kerja
            jam_kerja = absensi.calculate_total_jam_kerja()
            total_jam_kerja = jam_kerja.get('formatted', '-') if jam_kerja else '-'
            
            # Cek violation istirahat
            has_violation = False
            if absensi.tap_masuk and absensi.tap_pulang:
                if not absensi.tap_istirahat_keluar or not absensi.tap_istirahat_masuk:
                    has_violation = True
            
            pegawai_data.append({
                'userid': absensi.pegawai.userid,
                'nama': absensi.pegawai.nama_lengkap,
                'departemen': absensi.pegawai.departemen.nama
                    if absensi.pegawai.departemen else '-',
                'jabatan': absensi.pegawai.jabatan.nama
                    if absensi.pegawai.jabatan else '-',
                'tanggal': absensi.tanggal.strftime('%d %b %Y') if absensi.tanggal else '-',
                'tap_masuk': absensi.tap_masuk.strftime('%H:%M')
                    if absensi.tap_masuk else '-',
                'tap_istirahat_keluar': absensi.tap_istirahat_keluar.strftime('%H:%M')
                    if absensi.tap_istirahat_keluar else '-',
                'tap_istirahat_masuk': absensi.tap_istirahat_masuk.strftime('%H:%M')
                    if absensi.tap_istirahat_masuk else '-',
                'tap_pulang': absensi.tap_pulang.strftime('%H:%M')
                    if absensi.tap_pulang else '-',
                'total_jam_kerja': total_jam_kerja,
                'has_violation': has_violation,
                'keterangan': absensi.keterangan or '-',
                'is_late': absensi.is_late,
                'is_early_departure': absensi.is_early_departure
            })
        
        return JsonResponse({
            "status": "success",
            "pegawai": pegawai_data,
            "total": len(pegawai_data),
            "status_filter": status,
            "tanggal": today.strftime('%Y-%m-%d'),
            "cabang": cabang_aktif.nama if cabang_aktif else None
        })
    except Exception as e:
        return JsonResponse({"status": "error", "msg": str(e)}, status=500)

@login_required
def riwayat_absensi_hari_ini(request):
    """Endpoint AJAX: Mengambil riwayat absensi hari ini (log detail)."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    cabang_aktif = get_active_cabang(request)
    today = datetime.now().date()
    
    try:
        absensi_list = Absensi.objects.filter(
            tanggal=today
        ).select_related(
            'pegawai',
            'pegawai__departemen',
            'pegawai__jabatan'
        )
        
        if cabang_aktif:
            absensi_list = absensi_list.filter(pegawai__cabang=cabang_aktif)
        
        absensi_list = absensi_list.order_by('-tap_masuk', 'pegawai__nama_lengkap')
        
        absensi_data = []
        for absensi in absensi_list:
            # Hitung total jam kerja
            jam_kerja = absensi.calculate_total_jam_kerja()
            total_jam_kerja = jam_kerja.get('formatted', '-') if jam_kerja else '-'
            
            # Cek violation istirahat
            has_violation = False
            if absensi.tap_masuk and absensi.tap_pulang:
                if not absensi.tap_istirahat_keluar or not absensi.tap_istirahat_masuk:
                    has_violation = True
            
            absensi_data.append({
                'id': absensi.id,
                'userid': absensi.pegawai.userid,
                'nama': absensi.pegawai.nama_lengkap,
                'departemen': absensi.pegawai.departemen.nama
                    if absensi.pegawai.departemen else '-',
                'jabatan': absensi.pegawai.jabatan.nama
                    if absensi.pegawai.jabatan else '-',
                'tanggal': absensi.tanggal.strftime('%d %b %Y') if absensi.tanggal else '-',
                'tap_masuk': absensi.tap_masuk.strftime('%H:%M')
                    if absensi.tap_masuk else '-',
                'tap_istirahat_keluar': absensi.tap_istirahat_keluar.strftime('%H:%M')
                    if absensi.tap_istirahat_keluar else '-',
                'tap_istirahat_masuk': absensi.tap_istirahat_masuk.strftime('%H:%M')
                    if absensi.tap_istirahat_masuk else '-',
                'tap_pulang': absensi.tap_pulang.strftime('%H:%M')
                    if absensi.tap_pulang else '-',
                'total_jam_kerja': total_jam_kerja,
                'has_violation': has_violation,
                'status': absensi.status,
                'keterangan': absensi.keterangan or '-',
                'is_late': absensi.is_late,
                'is_early_departure': absensi.is_early_departure
            })
        
        return JsonResponse({
            "status": "success",
            "data": absensi_data,
            "total": len(absensi_data),
            "tanggal": today.strftime('%d %B %Y'),
            "tanggal_raw": today.strftime('%Y-%m-%d'),
            "cabang": cabang_aktif.nama if cabang_aktif else None
        })
    except Exception as e:
        return JsonResponse({"status": "error", "msg": str(e)}, status=500)
    
## FUNGSI MANAJEMEN PEGAWAI (CRUD)
# Kode ini untuk mengelola data master pegawai, termasuk daftar, detail, edit, dan hapus.

@login_required
def daftar_Pegawai(request):
    """Menampilkan daftar pegawai dengan filter status dan pencarian."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    # ========================================
    #  1: Get cabang aktif dengan error handling
    # ========================================
    cabang_aktif = None
    try:
        cabang_aktif = get_active_cabang(request)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Error getting active cabang: {str(e)}")
    
    # ========================================
    #  2: Get status filter
    # ========================================
    status_filter = request.GET.get('status_filter', 'active')
    
    # ========================================
    #  3: Get pegawai with fingerprint
    # ========================================
    pegawai_with_fp = []
    try:
        pegawai_with_fp = list(get_pegawai_with_fingerprint())
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Error getting pegawai with fingerprint: {str(e)}")
    
    # ========================================
    #  4: Query pegawai list
    # ========================================
    try:
        pegawai_list = Pegawai.objects.all()
        
        # Filter by cabang (dengan error handling)
        if cabang_aktif:
            try:
                pegawai_list = pegawai_list.filter(cabang=cabang_aktif)
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Error filtering by cabang: {str(e)}")
        
        # Filter berdasarkan status
        if status_filter == 'active':
            # Pegawai AKTIF: Sudah lengkap (UID + Fingerprint)
            pegawai_list = pegawai_list.filter(
                is_active=True,
                uid_mesin__gt=0
            )
            if pegawai_with_fp:
                pegawai_list = pegawai_list.filter(id__in=pegawai_with_fp)
            
        elif status_filter == 'pending':
            # Pegawai PENDING: Belum lengkap (UID kosong ATAU FP belum ada)
            pegawai_list = pegawai_list.filter(
                is_active=True
            ).filter(
                Q(uid_mesin__isnull=True) | 
                Q(uid_mesin=0) | 
                Q(uid_mesin__lte=0) |
                ~Q(id__in=pegawai_with_fp) if pegawai_with_fp else Q()
            )
            
        elif status_filter == 'inactive':
            # Pegawai NONAKTIF
            pegawai_list = pegawai_list.filter(is_active=False)
            
        else:
            # Default: tampilkan yang aktif
            pegawai_list = pegawai_list.filter(
                is_active=True,
                uid_mesin__gt=0
            )
            if pegawai_with_fp:
                pegawai_list = pegawai_list.filter(id__in=pegawai_with_fp)
            status_filter = 'active'
        
        # ✅ TAMBAHKAN mode_jam_kerja ke select_related
        pegawai_list = pegawai_list.select_related(
            'departemen', 
            'jabatan', 
            'cabang', 
            'mesin',
            'mode_jam_kerja'  # ✅ PENTING: Tambahkan ini
        ).prefetch_related(
            'mode_assignments',  # ✅ Prefetch assignments untuk performa
            'mode_assignments__mode'
        ).order_by('userid')
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error querying pegawai: {str(e)}")
        pegawai_list = Pegawai.objects.none()
    
    # ========================================
    #  5: Filter pencarian
    # ========================================
    search_form = PegawaiSearchForm(request.GET)
    if search_form.is_valid():
        query = search_form.cleaned_data.get('search_query')
        if query:
            try:
                pegawai_list = pegawai_list.filter(
                    Q(nama_lengkap__icontains=query) |
                    Q(userid__icontains=query) |
                    Q(jabatan__nama__icontains=query) |
                    Q(departemen__nama__icontains=query)
                )
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Error in search filter: {str(e)}")
    
    # ========================================
    #  6: Process pegawai untuk display
    # ========================================
    pegawai_display_list = []
    for pegawai in pegawai_list:
        # ✅ Ambil mode jam kerja info
        mode_names = []
        if pegawai.mode_jam_kerja:
            mode_names.append(pegawai.mode_jam_kerja.nama)
        
        # ✅ Tambahkan mode dari assignments
        for assignment in pegawai.mode_assignments.filter(is_active=True):
            if assignment.mode.nama not in mode_names:
                mode_names.append(assignment.mode.nama)
        
        pegawai.mode_names_display = ', '.join(mode_names) if mode_names else '-'
        pegawai.has_mode = len(mode_names) > 0
        
        pegawai_display_list.append(pegawai)
    
    # ========================================
    #  7: Hitung total per status
    # ========================================
    total_active = 0
    total_pending = 0
    total_inactive = 0
    
    try:
        base_query = Pegawai.objects.all()
        if cabang_aktif:
            try:
                base_query = base_query.filter(cabang=cabang_aktif)
            except Exception:
                pass
        
        # Total active
        active_query = base_query.filter(
            is_active=True,
            uid_mesin__gt=0
        )
        if pegawai_with_fp:
            active_query = active_query.filter(id__in=pegawai_with_fp)
        total_active = active_query.count()
        
        # Total pending (hitung manual untuk safety)
        all_active = base_query.filter(is_active=True)
        pending_count = 0
        
        for pegawai in all_active:
            if (pegawai.uid_mesin is None or 
                pegawai.uid_mesin <= 0 or 
                pegawai.id not in pegawai_with_fp):
                pending_count += 1
        
        total_pending = pending_count
        
        # Total inactive
        total_inactive = base_query.filter(is_active=False).count()
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Error counting totals: {str(e)}")
    
    # ========================================
    #  8: Build context
    # ========================================
    context = {
            'pegawai_list': pegawai_display_list,
            'search_form': search_form,
            'status_filter': status_filter,
            'total_active': total_active,
            'total_pending': total_pending,
            'total_inactive': total_inactive,
            'departemen_list': MasterDepartemen.objects.filter(is_active=True).order_by('nama'),
            'jabatan_list': MasterJabatan.objects.filter(is_active=True).order_by('nama'),
            'mode_jam_kerja_list': MasterModeJamKerja.objects.filter(is_active=True).order_by('nama'),
            'cabang_aktif': cabang_aktif,
            # ✅ TAMBAHAN: kirim mesin_list ke template untuk dropdown sync
            'mesin_list': MasterMesin.objects.filter(
                is_active=True, cabang=cabang_aktif
            ).order_by('nama') if cabang_aktif else MasterMesin.objects.filter(is_active=True).order_by('nama'),
        }
    return render(request, 'absensi_app/pegawai/daftar_pegawai.html', context)

@login_required
@require_http_methods(["GET"])
def Pegawai_detail(request, pk):
    """
    ✅ FINAL: Display ONLY assigned group schedules (not all groups)
    """
    from absensi_app.models import (
        Pegawai,
        MasterModeJamKerja,
        ModeJamKerjaJadwal,
        PegawaiModeAssignment
    )
    
    pegawai = get_object_or_404(Pegawai, id=pk)
    
    # Get cabang aktif dari session
    cabang_aktif = get_active_cabang(request)
    
    if cabang_aktif and pegawai.cabang != cabang_aktif:
        messages.error(
            request, 
            f"Pegawai '{pegawai.nama_lengkap}' tidak terdaftar di cabang {cabang_aktif.nama}."
        )
        return redirect('daftar_Pegawai')
    
    today = date.today()
    
    # Get jam kerja hari ini
    jam_kerja_info = WorkModeService.get_jam_kerja_for_pegawai(pegawai, today)
    jadwal_hari_ini = jam_kerja_info.get('jadwal')
    mode_aktif_hari_ini = jam_kerja_info.get('mode')
    periode_aktif = jam_kerja_info.get('periode')
    
    # ✅ Clear cache untuk data fresh
    from django.core.cache import cache
    cache_key = f"pegawai_mode_detail_{pegawai.id}"
    cache.delete(cache_key)
    
    # Get ALL existing assignments (FRESH QUERY)
    all_assignments = PegawaiModeAssignment.objects.filter(
        pegawai=pegawai,
        is_active=True
    ).select_related('mode').order_by('mode__nama')
    
    modes_with_schedule = []
    hari_names = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
    
    for assignment in all_assignments:
        mode = assignment.mode
        
        # ========================================
        # ✅ CRITICAL FIX: Filter jadwal by ASSIGNED group only
        # ========================================
        
        # Get jadwal_per_hari from assignment
        jadwal_per_hari_map = assignment.jadwal_per_hari or {}
        
        if not jadwal_per_hari_map:
            print(f"⚠️ WARNING: No jadwal_per_hari for {pegawai.nama_lengkap} - {mode.nama}")
            continue
        
        # ✅ Get ONLY jadwal IDs that are assigned to this pegawai
        assigned_jadwal_ids = set(jadwal_per_hari_map.values())
        
        if not assigned_jadwal_ids:
            print(f"⚠️ WARNING: Empty jadwal_ids for {pegawai.nama_lengkap} - {mode.nama}")
            continue
        
        # ✅ QUERY: Only jadwal that match assigned IDs
        jadwal_list = ModeJamKerjaJadwal.objects.filter(
            mode=mode,
            id__in=assigned_jadwal_ids  # ← CRITICAL: Filter by assignment
        ).order_by('hari')
        
        if not jadwal_list.exists():
            print(f"⚠️ WARNING: No matching jadwal for {pegawai.nama_lengkap} - {mode.nama}")
            continue
        
        # ✅ BUILD: Hanya hari yang ADA jadwal (dari assignment)
        jadwal_per_hari = {}  # {hari_index: jadwal_object}
        hari_yang_ada = set()
        
        for jadwal in jadwal_list:
            hari = jadwal.hari
            
            # Validasi hari (0-6)
            if 0 <= hari <= 6:
                hari_yang_ada.add(hari)
                
                # Ambil jadwal PERTAMA untuk hari ini (jika multiple shift)
                if hari not in jadwal_per_hari:
                    jadwal_per_hari[hari] = jadwal
        
        # ✅ Sort hari yang ada (Senin-Minggu)
        hari_sorted = sorted(hari_yang_ada)
        
        # ✅ Build jadwal_display (ONLY assigned days)
        jadwal_display = []
        for hari in hari_sorted:
            jadwal_display.append({
                'hari_index': hari,
                'hari_nama': hari_names[hari],
                'jadwal': jadwal_per_hari[hari],
                'is_today': hari == today.weekday()
            })
        
        # Cek mode aktif hari ini
        is_mode_today = mode_aktif_hari_ini and mode.id == mode_aktif_hari_ini.id
        
        # ✅ Cari periode aktif untuk mode ini
        periode_mode = mode.periode_list.filter(
            is_active=True,
            tanggal_mulai__lte=today,
            tanggal_selesai__gte=today
        ).first()
        
        # ✅ Get group name from first jadwal
        group_name = jadwal_list.first().group_name if jadwal_list.exists() else 'Unknown'
        
        modes_with_schedule.append({
            'mode': mode,
            'assignment': assignment,
            'group_name': group_name,  # ← TAMBAHAN: Display group name
            'jadwal_display': jadwal_display,
            'total_hari_kerja': len(hari_sorted),
            'periode_aktif': periode_mode,
            'is_mode_today': is_mode_today,
        })
    
    # Sort: mode aktif hari ini di atas
    modes_with_schedule.sort(key=lambda x: (not x['is_mode_today'], x['mode'].nama))
    
    # Statistics
    absensi_bulan_ini = pegawai.absensi.filter(
        tanggal__month=today.month,
        tanggal__year=today.year
    )
    
    context = {
        'pegawai': pegawai,
        'today': today,
        'jadwal_hari_ini': jadwal_hari_ini,
        'mode_aktif_hari_ini': mode_aktif_hari_ini,
        'periode_aktif': periode_aktif,
        
        # ✅ MODES WITH SCHEDULE (ONLY ASSIGNED GROUP)
        'modes_with_schedule': modes_with_schedule,
        'total_mode_assigned': len(modes_with_schedule),
        
        # Statistics
        'jam_kerja_info': jam_kerja_info,
        'total_hadir': absensi_bulan_ini.filter(status='Hadir').count(),
        'total_terlambat': absensi_bulan_ini.filter(is_late=True).count(),
        'total_pulang_cepat': absensi_bulan_ini.filter(is_early_departure=True).count(),
        'total_sakit_izin': absensi_bulan_ini.filter(status__in=['Sakit', 'Izin']).count(),
        
        'cabang_aktif': cabang_aktif,
    }
    
    return render(request, 'absensi_app/pegawai/pegawai_detail.html', context)

@login_required
def Pegawai_edit(request, pk):
    """
    ✅ FIXED: Edit pegawai dengan mode assignment yang benar
    
    KEY FIX:
    1. Support BOTH format: mode_assignments (new) dan schedules_json (fallback)
    2. Build jadwal_per_hari LENGKAP dari frontend data
    3. Clear cache setelah save
    """
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    pegawai_obj = get_object_or_404(Pegawai, pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # ========================================
                # 1. UPDATE DATA PRIBADI & ORGANISASI
                # ========================================
                pegawai_obj.nama_lengkap = request.POST.get('nama_lengkap', '').strip()
                pegawai_obj.email = request.POST.get('email', '').strip()
                pegawai_obj.tanggal_lahir = request.POST.get('tanggal_lahir') or None
                pegawai_obj.nomor_hp = request.POST.get('nomor_hp', '').strip()
                pegawai_obj.alamat = request.POST.get('alamat', '').strip()
                
                _update_pegawai_relations(pegawai_obj, request.POST)
                
                pegawai_obj.tanggal_bergabung = request.POST.get('tanggal_bergabung') or None
                pegawai_obj.tanggal_nonaktif = request.POST.get('tanggal_nonaktif') or None
                pegawai_obj.is_shift_worker = request.POST.get('is_shift_worker') == 'on'
                pegawai_obj.is_active = request.POST.get('is_active') == 'on'
                
                pegawai_obj.save()
                
                # ========================================
                # 2. PROSES MODE ASSIGNMENTS (CRITICAL FIX)
                # ========================================
                
                mode_assignments_json = request.POST.get('mode_assignments', '{}')
                
                print("\n" + "="*60)
                print(f"🔧 EDIT PEGAWAI: {pegawai_obj.nama_lengkap}")
                print("="*60)
                print(f"📥 Raw mode_assignments: {mode_assignments_json[:200]}...")
                
                if not mode_assignments_json or mode_assignments_json == '{}':
                    print("⚠️  WARNING: mode_assignments kosong!")
                    messages.warning(request, '⚠️ Mode jam kerja tidak diubah (data kosong)')
                else:
                    try:
                        mode_assignments = json.loads(mode_assignments_json)
                        print(f"✅ Parsed: {len(mode_assignments)} modes")
                    except json.JSONDecodeError as e:
                        print(f"❌ JSON Error: {str(e)}")
                        messages.warning(request, f'⚠️ Format assignment tidak valid, skip: {str(e)}')
                        mode_assignments = {}
                    
                    if mode_assignments:
                        # ✅ HAPUS assignment lama
                        deleted_count = pegawai_obj.mode_assignments.all().delete()[0]
                        print(f"🗑️  Deleted {deleted_count} old assignments")
                        
                        # ✅ BUAT assignment baru
                        created_count = 0
                        
                        for mode_id_str, assignment_data in mode_assignments.items():
                            try:
                                mode_id = int(mode_id_str)
                                mode = MasterModeJamKerja.objects.get(id=mode_id)
                                
                                print(f"\n   🔧 Mode {mode_id}: {mode.nama}")
                                
                                # ========================================
                                # ✅ CRITICAL: Get jadwal_per_hari
                                # ========================================
                                
                                jadwal_per_hari = assignment_data.get('jadwal_per_hari', {})
                                
                                print(f"   📥 Received jadwal_per_hari: {jadwal_per_hari}")
                                
                                # ✅ VALIDASI: Pastikan tidak kosong
                                if not jadwal_per_hari or not any(v for v in jadwal_per_hari.values()):
                                    print(f"   ⚠️  jadwal_per_hari kosong, skip mode ini")
                                    continue
                                
                                # ✅ VALIDASI: Semua jadwal_id valid
                                valid_jadwal_ids = set(
                                    ModeJamKerjaJadwal.objects.filter(
                                        mode_id=mode_id
                                    ).values_list('id', flat=True)
                                )
                                
                                cleaned_jadwal = {}
                                for hari_str, jadwal_id in jadwal_per_hari.items():
                                    # ✅ Convert ke int jika perlu
                                    try:
                                        jadwal_id_int = int(jadwal_id)
                                    except (ValueError, TypeError):
                                        print(f"   ⚠️  Invalid jadwal_id: {jadwal_id}, skip")
                                        continue
                                    
                                    if jadwal_id_int in valid_jadwal_ids:
                                        cleaned_jadwal[str(hari_str)] = jadwal_id_int
                                    else:
                                        print(f"   ⚠️  Jadwal ID {jadwal_id_int} tidak valid, skip")
                                
                                if not cleaned_jadwal:
                                    print(f"   ❌ No valid jadwal for mode {mode.nama}")
                                    continue
                                
                                print(f"   ✅ Final jadwal: {cleaned_jadwal} ({len(cleaned_jadwal)} hari)")
                                
                                # ✅ CREATE ASSIGNMENT
                                PegawaiModeAssignment.objects.create(
                                    pegawai=pegawai_obj,
                                    mode_id=mode_id,
                                    jadwal_per_hari=cleaned_jadwal,
                                    is_active=True
                                )
                                
                                created_count += 1
                                print(f"   ✅ Created assignment untuk {mode.nama}")
                                
                            except (ValueError, TypeError) as e:
                                print(f"   ❌ Error parsing mode {mode_id_str}: {str(e)}")
                                continue
                            except MasterModeJamKerja.DoesNotExist:
                                print(f"   ❌ Mode {mode_id} tidak ditemukan")
                                continue
                        
                        print(f"\n📊 Total assignments created: {created_count}")
                
                # ========================================
                # 3. CLEAR CACHE (PENTING!)
                # ========================================
                
                from django.core.cache import cache
                from .services import WorkModeService
                
                print(f"\n🧹 Clearing cache for pegawai {pegawai_obj.id}...")
                
                # Clear global cache
                WorkModeService.clear_cache()
                
                # Clear pegawai-specific cache
                cache_key = f"pegawai_mode_detail_{pegawai_obj.id}"
                cache.delete(cache_key)
                
                # Clear jadwal cache (7 hari ke depan)
                for i in range(7):
                    cache_key = f"jadwal_{pegawai_obj.id}_{date.today() + timedelta(days=i)}"
                    cache.delete(cache_key)
                
                print(f"✅ Cache cleared!")
                print("="*60 + "\n")
                
                # ========================================
                # 4. SUCCESS MESSAGE
                # ========================================
                
                messages.success(
                    request,
                    f"✅ Data {pegawai_obj.nama_lengkap} berhasil diperbarui.\n\n"
                    f"💡 Refresh halaman detail untuk melihat perubahan mode jam kerja."
                )
                return redirect('Pegawai_detail', pk=pk)
        
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"\n❌ ERROR in Pegawai_edit:\n{error_detail}")
            messages.error(request, f"❌ Error: {str(e)}")
            return redirect('Pegawai_edit', pk=pk)
    
    # ========================================
    # GET: Load data untuk form
    # ========================================
    
    context = {
        'pegawai': pegawai_obj,
        'departemen_list': MasterDepartemen.objects.filter(is_active=True).order_by('nama'),
        'jabatan_list': MasterJabatan.objects.filter(is_active=True).order_by('nama'),
        'cabang_list': MasterCabang.objects.filter(is_active=True).order_by('nama'),
        'mesin_list': MasterMesin.objects.filter(is_active=True).order_by('nama'),
        'mode_jam_kerja_list': MasterModeJamKerja.objects.filter(is_active=True).order_by('nama'),
    }
    return render(request, 'absensi_app/pegawai/pegawai_edit.html', context)

def _update_pegawai_relations(pegawai, post_data):
    """Helper untuk memperbarui relasi foreign key pegawai"""
    departemen_id = post_data.get('departemen')
    pegawai.departemen = MasterDepartemen.objects.get(id=departemen_id) if departemen_id else None
    
    jabatan_id = post_data.get('jabatan')
    pegawai.jabatan = MasterJabatan.objects.get(id=jabatan_id) if jabatan_id else None
    
    cabang_id = post_data.get('cabang')
    if cabang_id:
        pegawai.cabang = MasterCabang.objects.get(id=cabang_id)
    
    mesin_id = post_data.get('mesin')
    pegawai.mesin = MasterMesin.objects.get(id=mesin_id) if mesin_id else None
    
    mode_id = post_data.get('mode_jam_kerja')
    pegawai.mode_jam_kerja = MasterModeJamKerja.objects.get(id=mode_id) if mode_id else None


@login_required
def Pegawai_hapus(request, pk):
    """Menghapus data pegawai dan data terkait secara permanen."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    pegawai_obj = get_object_or_404(Pegawai, pk=pk)
    
    if request.method == 'POST':
        try:
            pegawai_nama = pegawai_obj.nama_lengkap
            
            with transaction.atomic():
                pegawai_obj.absensi.all().delete()
                pegawai_obj.fingerprint_templates.all().delete()
                pegawai_obj.delete()
            
            messages.success(request, f'Pegawai {pegawai_nama} berhasil dihapus.')
            return redirect('daftar_Pegawai')
        except Exception as e:
            messages.error(request, f'Gagal menghapus: {str(e)}')
            return redirect('Pegawai_detail', pk=pk)
    
    return render(
        request,
        'absensi_app/pegawai/pegawai_hapus.html',
        {'pegawai': pegawai_obj}
    )


@login_required
def toggle_pegawai_status(request, pk):
    """Mengubah status aktif/nonaktif pegawai."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    if request.method != 'POST':
        return redirect('daftar_Pegawai')
    
    try:
        pegawai = get_object_or_404(Pegawai, pk=pk)
        pegawai.is_active = not pegawai.is_active
        pegawai.save()
        
        status_text = "AKTIF" if pegawai.is_active else "NONAKTIF"
        messages.success(
            request,
            f"Status {pegawai.nama_lengkap} diubah menjadi {status_text}"
        )
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect('daftar_Pegawai')


## FUNGSI OPERASI PEGAWAI MASSAL (BULK)
# Kode ini untuk melakukan operasi massal seperti aktivasi, nonaktivasi, dan hapus.

@login_required
def bulk_deactivate_pegawai(request):
    """Endpoint AJAX: Menonaktifkan beberapa pegawai sekaligus."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({"status": "error", "msg": "Method tidak diizinkan"}, status=405)
    
    try:
        pegawai_ids = _parse_bulk_ids(request.POST.get('pegawai_ids', ''))
        
        if not pegawai_ids:
            return JsonResponse(
                {"status": "error", "msg": "Tidak ada pegawai yang dipilih"},
                status=400
            )
        
        deactivated = _bulk_update_status(pegawai_ids, is_active=False)
        
        return JsonResponse({
            "status": "success",
            "msg": f"{deactivated} pegawai dinonaktifkan",
            "deactivated": deactivated
        })
    except Exception as e:
        return JsonResponse({"status": "error", "msg": str(e)}, status=500)


@login_required
def bulk_activate_pegawai(request):
    """Endpoint AJAX: Mengaktifkan beberapa pegawai sekaligus."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({"status": "error", "msg": "Method tidak diizinkan"}, status=405)
    
    try:
        pegawai_ids = _parse_bulk_ids(request.POST.get('pegawai_ids', ''))
        
        if not pegawai_ids:
            return JsonResponse(
                {"status": "error", "msg": "Tidak ada pegawai yang dipilih"},
                status=400
            )
        
        activated = _bulk_update_status(pegawai_ids, is_active=True)
        
        return JsonResponse({
            "status": "success",
            "msg": f"{activated} pegawai diaktifkan",
            "activated": activated
        })
    except Exception as e:
        return JsonResponse({"status": "error", "msg": str(e)}, status=500)


@login_required
def bulk_delete_pegawai(request):
    """Endpoint AJAX: Menghapus beberapa pegawai secara permanen."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({"status": "error", "msg": "Method tidak diizinkan"}, status=405)
    
    try:
        pegawai_ids = _parse_bulk_ids(request.POST.get('pegawai_ids', ''))
        
        if not pegawai_ids:
            return JsonResponse(
                {"status": "error", "msg": "Tidak ada pegawai yang dipilih"},
                status=400
            )
        
        deleted_count = 0
        for pegawai_id in pegawai_ids:
            try:
                pegawai = Pegawai.objects.get(id=pegawai_id)
                
                with transaction.atomic():
                    pegawai.absensi.all().delete()
                    pegawai.fingerprint_templates.all().delete()
                    pegawai.delete()
                
                deleted_count += 1
            except Pegawai.DoesNotExist:
                continue
        
        return JsonResponse({
            "status": "success",
            "msg": f"{deleted_count} pegawai dihapus secara permanen",
            "deleted": deleted_count
        })
    except Exception as e:
        return JsonResponse({"status": "error", "msg": str(e)}, status=500)


def _parse_bulk_ids(ids_string):
    """Helper untuk mem-parsing string ID yang dipisahkan koma."""
    return [id.strip() for id in ids_string.split(',') if id.strip()]


def _bulk_update_status(pegawai_ids, is_active):
    """Helper untuk memperbarui status banyak pegawai."""
    count = 0
    for pegawai_id in pegawai_ids:
        try:
            pegawai = Pegawai.objects.get(id=pegawai_id)
            pegawai.is_active = is_active
            pegawai.save()
            count += 1
        except Pegawai.DoesNotExist:
            continue
    return count


## FUNGSI PENDAFTARAN PEGAWAI
# Kode ini untuk menangani alur pendaftaran pegawai, baik secara manual maupun dari mesin.

@login_required
def register_Pegawai_menu(request):
    """Menu pendaftaran pegawai (manual atau dari mesin)."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')

    cabang_aktif = get_active_cabang(request)
    
    context = {
        'cabang_aktif': cabang_aktif,
    }
    
    return render(request, 'absensi_app/menu/register_Pegawai_menu.html', context)


@login_required
def register_Pegawai(request):
    """
    ✅ FINAL VERSION: Mendaftarkan pegawai secara manual ke database.
    
    CRITICAL FIX:
    - Build jadwal_per_hari LENGKAP untuk semua hari dalam group
    - Tidak hanya 1 hari seperti sebelumnya
    """
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    # ✅ VALIDASI CABANG AKTIF
    cabang_aktif = get_active_cabang(request)
    
    if not cabang_aktif:
        messages.warning(
            request,
            "⚠️ Silakan pilih cabang terlebih dahulu dari menu di pojok kanan atas!"
        )
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = PegawaiForm(request.POST)
        
        mode_assignments_json = request.POST.get('mode_assignments', '{}')
        
        if not mode_assignments_json or mode_assignments_json == '{}':
            messages.error(request, '❌ Minimal pilih 1 grup jam kerja!')
            return _render_pegawai_form_with_modes(request, form, cabang_aktif)
        
        try:
            mode_assignments = json.loads(mode_assignments_json)
        except json.JSONDecodeError as e:
            messages.error(request, f'❌ Format assignment tidak valid: {str(e)}')
            return _render_pegawai_form_with_modes(request, form, cabang_aktif)
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    new_pegawai = form.save(commit=False)
                    new_pegawai.userid = request.POST.get('userid', '').strip()
                    new_pegawai.uid_mesin = 0
                    new_pegawai.is_active = True
                    
                    # ✅ FORCE SET CABANG AKTIF
                    new_pegawai.cabang = cabang_aktif
                    
                    new_pegawai.email = request.POST.get('email', '').strip()
                    
                    tanggal_bergabung = request.POST.get('tanggal_bergabung')
                    if tanggal_bergabung:
                        new_pegawai.tanggal_bergabung = datetime.strptime(
                            tanggal_bergabung, '%Y-%m-%d'
                        ).date()
                    else:
                        new_pegawai.tanggal_bergabung = datetime.now().date()
                    
                    # Set mode_jam_kerja default
                    default_mode_id = None
                    for mode_id_str, assignment_data in mode_assignments.items():
                        jadwal_per_hari = assignment_data.get('jadwal_per_hari', {})
                        group_id = assignment_data.get('group_id')
                        
                        if jadwal_per_hari or group_id:
                            default_mode_id = int(mode_id_str)
                            break
                    
                    if default_mode_id:
                        new_pegawai.mode_jam_kerja_id = default_mode_id
                    
                    new_pegawai.save()
                    
                    # ========================================
                    # ✅ CRITICAL FIX: SIMPAN MODE ASSIGNMENTS DENGAN JADWAL LENGKAP
                    # ========================================
                    
                    assignment_count = 0
                    
                    print("\n" + "="*60)
                    print("💾 SAVING MODE ASSIGNMENTS - REGISTER MANUAL")
                    print("="*60)
                    print(f"👤 Pegawai: {new_pegawai.nama_lengkap} (ID: {new_pegawai.id})")
                    
                    for mode_id_str, assignment_data in mode_assignments.items():
                        try:
                            mode_id = int(mode_id_str)
                            mode = MasterModeJamKerja.objects.get(id=mode_id)
                            
                            print(f"\n📋 Mode: {mode.nama} (ID: {mode_id})")
                            
                            # ========================================
                            # ✅ GET jadwal_per_hari (PRIORITAS TERTINGGI)
                            # ========================================
                            
                            jadwal_per_hari = assignment_data.get('jadwal_per_hari', {})
                            
                            print(f"   📥 Received jadwal_per_hari: {jadwal_per_hari}")
                            
                            # ✅ VALIDASI: Pastikan jadwal_per_hari ada dan tidak kosong
                            if jadwal_per_hari and any(v for v in jadwal_per_hari.values()):
                                # ========================================
                                # ✅ FORMAT 1: Sudah ada jadwal_per_hari lengkap dari frontend
                                # ========================================
                                
                                print(f"   ✅ Format 1: Gunakan jadwal_per_hari langsung ({len(jadwal_per_hari)} hari)")
                                
                                # Validasi semua jadwal_id valid
                                valid_jadwal_ids = set(
                                    ModeJamKerjaJadwal.objects.filter(
                                        mode_id=mode_id
                                    ).values_list('id', flat=True)
                                )
                                
                                cleaned_jadwal = {}
                                for hari_str, jadwal_id in jadwal_per_hari.items():
                                    try:
                                        jadwal_id_int = int(jadwal_id)
                                    except (ValueError, TypeError):
                                        print(f"   ⚠️ Invalid jadwal_id: {jadwal_id}, skip")
                                        continue
                                    
                                    if jadwal_id_int in valid_jadwal_ids:
                                        cleaned_jadwal[str(hari_str)] = jadwal_id_int
                                    else:
                                        print(f"   ⚠️ Jadwal ID {jadwal_id_int} tidak valid, skip")
                                
                                if not cleaned_jadwal:
                                    print(f"   ❌ No valid jadwal after cleaning!")
                                    continue
                                
                                PegawaiModeAssignment.objects.create(
                                    pegawai=new_pegawai,
                                    mode_id=mode_id,
                                    jadwal_per_hari=cleaned_jadwal,
                                    is_active=True
                                )
                                assignment_count += 1
                                print(f"   ✅ Assignment created with {len(cleaned_jadwal)} hari!")
                            
                            else:
                                # ========================================
                                # ✅ FORMAT 2: Build dari group_id (FALLBACK)
                                # ========================================
                                
                                group_id = assignment_data.get('group_id')
                                
                                if not group_id:
                                    print(f"   ❌ No jadwal_per_hari and no group_id, skip!")
                                    continue
                                
                                print(f"   🔄 Format 2: Build dari group_id={group_id}")
                                
                                # ✅ STEP 1: Ambil jadwal pertama untuk dapat group_name
                                first_jadwal = ModeJamKerjaJadwal.objects.filter(
                                    mode_id=mode_id,
                                    id=group_id
                                ).first()
                                
                                if not first_jadwal:
                                    print(f"   ❌ Jadwal ID {group_id} tidak ditemukan!")
                                    continue
                                
                                group_name = first_jadwal.group_name
                                print(f"   ✅ Group name: '{group_name}'")
                                
                                # ✅ STEP 2: Ambil SEMUA jadwal dalam group yang sama
                                all_jadwal_in_group = ModeJamKerjaJadwal.objects.filter(
                                    mode_id=mode_id,
                                    group_name=group_name
                                ).order_by('hari')
                                
                                if not all_jadwal_in_group.exists():
                                    print(f"   ❌ No jadwal found for group '{group_name}'!")
                                    continue
                                
                                # ✅ STEP 3: Build jadwal_per_hari lengkap
                                built_jadwal = {}
                                for jdl in all_jadwal_in_group:
                                    built_jadwal[str(jdl.hari)] = jdl.id
                                    print(f"      • Hari {jdl.hari}: Jadwal ID {jdl.id}")
                                
                                print(f"   ✅ Built jadwal_per_hari: {built_jadwal} ({len(built_jadwal)} hari)")
                                
                                if not built_jadwal:
                                    print(f"   ❌ Built jadwal kosong!")
                                    continue
                                
                                # ✅ STEP 4: Create assignment
                                PegawaiModeAssignment.objects.create(
                                    pegawai=new_pegawai,
                                    mode_id=mode_id,
                                    jadwal_per_hari=built_jadwal,
                                    is_active=True
                                )
                                assignment_count += 1
                                print(f"   ✅ Assignment created dengan {len(built_jadwal)} hari!")
                        
                        except MasterModeJamKerja.DoesNotExist:
                            print(f"   ❌ Mode ID {mode_id} tidak ditemukan!")
                            continue
                        except Exception as e:
                            import traceback
                            error_detail = traceback.format_exc()
                            print(f"   ❌ Error: {str(e)}")
                            print(f"   Debug: {error_detail}")
                            continue
                    
                    print(f"\n{'='*60}")
                    print(f"📊 TOTAL ASSIGNMENTS CREATED: {assignment_count}")
                    print(f"{'='*60}\n")
                    
                    # Clear cache
                    WorkModeService.clear_cache()
                    
                    messages.success(
                        request,
                        f"✅ Pegawai {new_pegawai.nama_lengkap} berhasil didaftarkan!\n\n"
                        f"📋 Detail:\n"
                        f"• User ID: {new_pegawai.userid}\n"
                        f"• Cabang: {cabang_aktif.nama}\n"
                        f"• Departemen: {new_pegawai.departemen.nama if new_pegawai.departemen else '-'}\n"
                        f"• Mode Assignments: {assignment_count} mode\n\n"
                        f"⚠️ Langkah Selanjutnya:\n"
                        f"Silakan lanjutkan ke menu 'Sinkron ke Mesin' untuk registrasi fingerprint!"
                    )
                    
                    return redirect('sinkron_ke_mesin')
                
            except Exception as e:
                import traceback
                error_detail = traceback.format_exc()
                print(f"\n❌ ERROR SAVE PEGAWAI:\n{error_detail}")
                messages.error(request, f'❌ Error: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = PegawaiForm()
    
    return _render_pegawai_form_with_modes(request, form, cabang_aktif)


def _render_pegawai_form_with_modes(request, form, cabang_aktif):
    """Helper untuk render form dengan filtering cabang."""
    from django.db.models import Q
    
    # ✅ FILTER DEPARTEMEN
    # Ambil departemen yang ada pegawainya di cabang ini + departemen kosong
    departemen_ids = Pegawai.objects.filter(
        cabang=cabang_aktif,
        is_active=True
    ).values_list('departemen_id', flat=True).distinct()
    
    departemen_list = MasterDepartemen.objects.filter(
        is_active=True
    ).filter(
        Q(id__in=departemen_ids) | Q(pegawai_list__isnull=True)
    ).distinct().order_by('nama')
    
    # ✅ FILTER MODE: GLOBAL atau sesuai cabang
    mode_list = MasterModeJamKerja.objects.filter(
        is_active=True
    ).filter(
        Q(cabang__isnull=True) | Q(cabang=cabang_aktif)
    ).order_by('-is_default', '-priority', 'nama')
    
    return render(request, 'absensi_app/register/register_pegawai.html', {
        'form': form,
        'departemen_list': departemen_list,
        'jabatan_list': MasterJabatan.objects.filter(is_active=True).order_by('nama'),
        'mode_list': mode_list,
        'cabang_aktif': cabang_aktif,
    })


@login_required
def api_get_mode_jadwal_departemen(request, pk):
    """Redirect ke fungsi ambil grup jam kerja per mode."""
    return api_get_jam_kerja_groups(request, pk)

@login_required
def api_get_jam_kerja_groups(request, pk):
    """
    API untuk mendapatkan grup jam kerja yang BENAR
    Grup = jadwal dengan group_name yang SAMA
    """
    from django.http import JsonResponse
    
    try:
        from .models import MasterModeJamKerja, ModeJamKerjaJadwal
        
        print(f"\n{'='*60}")
        print(f"API GET JAM KERJA GROUPS - Mode ID: {pk}")
        print(f"{'='*60}")
        
        mode = MasterModeJamKerja.objects.get(id=pk, is_active=True)
        print(f"✅ Mode: {mode.nama}")
        
        # Ambil SEMUA jadwal, urutkan berdasarkan group_name dan hari
        jadwal_all = ModeJamKerjaJadwal.objects.filter(mode=mode).order_by('group_name', 'hari')
        
        print(f"📊 Total jadwal di DB: {jadwal_all.count()}")
        
        if jadwal_all.count() == 0:
            return JsonResponse({'status': 'success', 'groups': []})
        
        # Group berdasarkan group_name yang SAMA
        groups_dict = {}
        
        for jadwal in jadwal_all:
            # Gunakan group_name sebagai key
            group_key = jadwal.group_name if jadwal.group_name else f"Grup-{jadwal.id}"
            
            print(f"  Jadwal ID={jadwal.id}, group_name='{group_key}', hari={jadwal.hari}")
            
            # Inisialisasi grup baru jika belum ada
            if group_key not in groups_dict:
                groups_dict[group_key] = {
                    'id': jadwal.id,  # ID jadwal pertama dalam grup
                    'nama': group_key,
                    'jam_masuk': str(jadwal.jam_masuk),
                    'jam_keluar': str(jadwal.jam_keluar),
                    'hari_kerja': 0,
                    'jadwal': []
                }
            
            # Tambahkan jadwal ke grup
            hari_names = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
            groups_dict[group_key]['jadwal'].append({
                'hari': jadwal.hari,
                'jadwal_id': jadwal.id,
                'hari_nama': hari_names[jadwal.hari] if 0 <= jadwal.hari < 7 else f'Hari-{jadwal.hari}',
                'jam_masuk': str(jadwal.jam_masuk),
                'jam_pulang': str(jadwal.jam_keluar)
            })
            
            groups_dict[group_key]['hari_kerja'] += 1
        
        # Convert dict to list
        groups = list(groups_dict.values())
        
        print(f"\n✅ Hasil grouping: {len(groups)} grup")
        for g in groups:
            print(f"   - Grup '{g['nama']}': {g['hari_kerja']} hari")
            for jd in g['jadwal']:
                print(f"      • {jd['hari_nama']} (ID={jd['jadwal_id']}): {jd['jam_masuk']}-{jd['jam_pulang']}")
        print(f"{'='*60}\n")
        
        return JsonResponse({'status': 'success', 'groups': groups})
        
    except MasterModeJamKerja.DoesNotExist:
        return JsonResponse({'status': 'error', 'msg': 'Mode tidak ditemukan'}, status=404)
    except Exception as e:
        import traceback
        print(f"\n❌ ERROR:\n{traceback.format_exc()}\n")
        return JsonResponse({'status': 'error', 'msg': str(e)}, status=500)


@login_required
def api_get_applicable_modes(request):
    """API untuk mendapatkan semua mode jam kerja yang aktif"""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        modes = MasterModeJamKerja.objects.filter(is_active=True)
        modes = modes.order_by('-is_default', '-priority', 'nama')
        
        mode_list = []
        for mode in modes:
            group_names = set()
            for jadwal in mode.jadwal_list.all():
                if jadwal.group_name:
                    group_names.add(jadwal.group_name)
            
            mode_list.append({
                'id': mode.id,
                'nama': mode.nama,
                'kode': mode.kode,
                'warna': mode.warna,
                'icon': mode.icon,
                'is_default': mode.is_default,
                'priority': mode.priority,
                'groups': ', '.join(sorted(group_names)) if group_names else 'Belum ada grup',
            })
        
        return JsonResponse({
            "status": "success",
            "modes": mode_list,
            "total": len(mode_list)
        })
    
    except Exception as e:
        return JsonResponse({"status": "error", "msg": str(e)}, status=500)
    

@login_required
def register_Pegawai_dari_mesin(request):
    """Menampilkan halaman untuk mengambil data pegawai dari mesin sidik jari."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    # ✅ VALIDASI CABANG AKTIF
    cabang_aktif = get_active_cabang(request)
    
    if not cabang_aktif:
        messages.warning(
            request,
            "⚠️ Silakan pilih cabang terlebih dahulu dari menu di pojok kanan atas!"
        )
        return redirect('dashboard')
    
    # ✅ FILTER DATA BERDASARKAN CABANG
    departemen_ids = Pegawai.objects.filter(
        cabang=cabang_aktif,
        is_active=True
    ).values_list('departemen_id', flat=True).distinct()
    
    departemen_list = MasterDepartemen.objects.filter(
        Q(id__in=departemen_ids) | Q(pegawai_list__isnull=True),
        is_active=True
    ).distinct().order_by('id_departemen')
    
    mode_list = MasterModeJamKerja.objects.filter(
        Q(cabang__isnull=True) | Q(cabang=cabang_aktif),
        is_active=True
    ).order_by('nama')
    
    mesin_list = MasterMesin.objects.filter(
        cabang=cabang_aktif,
        is_active=True
    ).select_related('cabang').order_by('nama')
    
    context = {
        'departemen_list': departemen_list,
        'jabatan_list': MasterJabatan.objects.filter(is_active=True).order_by('nama'),
        'mode_jam_kerja_list': mode_list,
        'mesin_list': mesin_list,
        'page_title': 'Ambil Data dari Mesin Sidik Jari',
        'cabang_aktif': cabang_aktif,
    }
    
    return render(request, 'absensi_app/register/register_dari_mesin.html', context)


@login_required
def simpan_Pegawai_dari_mesin(request):
    """
    ✅ FINAL VERSION: Menyimpan data pegawai yang diambil dari mesin ke database.
    
    CRITICAL FIX:
    - Build jadwal_per_hari LENGKAP untuk semua hari dalam group
    - Tidak hanya 1 hari seperti sebelumnya
    """
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    if request.headers.get("X-Requested-With") != "XMLHttpRequest":
        return JsonResponse({"status": "error", "msg": "Permintaan tidak valid"}, status=400)
    
    # ✅ VALIDASI CABANG AKTIF
    cabang_aktif = get_active_cabang(request)
    
    if not cabang_aktif:
        return JsonResponse({
            "status": "error",
            "msg": "⚠️ Silakan pilih cabang terlebih dahulu!"
        }, status=400)
    
    with transaction.atomic():
        try:
            # ========================================
            # 1. VALIDASI MESIN
            # ========================================
            
            mesin_id = request.POST.get("mesin_id")
            if not mesin_id:
                return JsonResponse({
                    "status": "error",
                    "msg": "Pilih mesin terlebih dahulu"
                }, status=400)
            
            mesin = get_object_or_404(MasterMesin, id=mesin_id, is_active=True)
            
            # ✅ VALIDASI MESIN SESUAI CABANG
            if mesin.cabang != cabang_aktif:
                return JsonResponse({
                    "status": "error",
                    "msg": f"Mesin '{mesin.nama}' tidak terdaftar di cabang {cabang_aktif.nama}"
                }, status=403)
            
            # ========================================
            # 2. VALIDASI MODE ASSIGNMENTS
            # ========================================
            
            mode_assignments_json = request.POST.get('mode_assignments', '{}')
            if not mode_assignments_json or mode_assignments_json == '{}':
                return JsonResponse({
                    "status": "error",
                    "msg": "Minimal pilih 1 grup jam kerja untuk salah satu mode!"
                }, status=400)
            
            try:
                mode_assignments = json.loads(mode_assignments_json)
            except json.JSONDecodeError as e:
                return JsonResponse({
                    "status": "error",
                    "msg": f"Format assignment tidak valid: {str(e)}"
                }, status=400)
            
            # ========================================
            # 3. VALIDASI UID
            # ========================================
            
            uid = request.POST.get("uid")
            if not uid:
                return JsonResponse({
                    "status": "error",
                    "msg": "UID dari mesin tidak ditemukan!"
                }, status=400)
            
            # ========================================
            # 4. VALIDASI DATA PEGAWAI
            # ========================================
            
            validation_result = _validate_pegawai_data(request.POST)
            if validation_result:
                return validation_result
            
            # ========================================
            # 5. CREATE PEGAWAI
            # ========================================
            
            new_pegawai = _create_pegawai_from_machine(request.POST, mesin)
            new_pegawai.uid_mesin = int(uid)
            
            # ✅ FORCE SET CABANG AKTIF
            new_pegawai.cabang = cabang_aktif
            
            # ========================================
            # 6. SET DEFAULT MODE
            # ========================================
            
            default_mode_id = None
            for mode_id_str, assignment_data in mode_assignments.items():
                jadwal_per_hari = assignment_data.get('jadwal_per_hari', {})
                group_id = assignment_data.get('group_id')
                if jadwal_per_hari or group_id:
                    default_mode_id = int(mode_id_str)
                    break
            
            if default_mode_id:
                new_pegawai.mode_jam_kerja_id = default_mode_id
            
            new_pegawai.save()
            
            # ========================================
            # ✅ CRITICAL FIX: SIMPAN MODE ASSIGNMENTS DENGAN JADWAL LENGKAP
            # ========================================
            
            assignment_count = 0
            
            print("\n" + "="*60)
            print("💾 SAVING MODE ASSIGNMENTS - REGISTER DARI MESIN")
            print("="*60)
            print(f"👤 Pegawai: {new_pegawai.nama_lengkap} (ID: {new_pegawai.id})")
            
            for mode_id_str, assignment_data in mode_assignments.items():
                try:
                    mode_id = int(mode_id_str)
                    mode = MasterModeJamKerja.objects.get(id=mode_id)
                    
                    print(f"\n📋 Mode: {mode.nama} (ID: {mode_id})")
                    
                    # ========================================
                    # ✅ GET jadwal_per_hari (PRIORITAS TERTINGGI)
                    # ========================================
                    
                    jadwal_per_hari = assignment_data.get('jadwal_per_hari', {})
                    
                    print(f"   📥 Received jadwal_per_hari: {jadwal_per_hari}")
                    
                    # ✅ VALIDASI: Pastikan jadwal_per_hari ada dan tidak kosong
                    if jadwal_per_hari and any(v for v in jadwal_per_hari.values()):
                        # ========================================
                        # ✅ FORMAT 1: Sudah ada jadwal_per_hari lengkap dari frontend
                        # ========================================
                        
                        print(f"   ✅ Format 1: Gunakan jadwal_per_hari langsung ({len(jadwal_per_hari)} hari)")
                        
                        # Validasi semua jadwal_id valid
                        valid_jadwal_ids = set(
                            ModeJamKerjaJadwal.objects.filter(
                                mode_id=mode_id
                            ).values_list('id', flat=True)
                        )
                        
                        cleaned_jadwal = {}
                        for hari_str, jadwal_id in jadwal_per_hari.items():
                            try:
                                jadwal_id_int = int(jadwal_id)
                            except (ValueError, TypeError):
                                print(f"   ⚠️ Invalid jadwal_id: {jadwal_id}, skip")
                                continue
                            
                            if jadwal_id_int in valid_jadwal_ids:
                                cleaned_jadwal[str(hari_str)] = jadwal_id_int
                            else:
                                print(f"   ⚠️ Jadwal ID {jadwal_id_int} tidak valid, skip")
                        
                        if not cleaned_jadwal:
                            print(f"   ❌ No valid jadwal after cleaning!")
                            continue
                        
                        PegawaiModeAssignment.objects.create(
                            pegawai=new_pegawai,
                            mode_id=mode_id,
                            jadwal_per_hari=cleaned_jadwal,
                            is_active=True
                        )
                        assignment_count += 1
                        print(f"   ✅ Assignment created with {len(cleaned_jadwal)} hari!")
                    
                    else:
                        # ========================================
                        # ✅ FORMAT 2: Build dari group_id (FALLBACK)
                        # ========================================
                        
                        group_id = assignment_data.get('group_id')
                        
                        if not group_id:
                            print(f"   ❌ No jadwal_per_hari and no group_id, skip!")
                            continue
                        
                        print(f"   🔄 Format 2: Build dari group_id={group_id}")
                        
                        # ✅ STEP 1: Ambil jadwal pertama untuk dapat group_name
                        first_jadwal = ModeJamKerjaJadwal.objects.filter(
                            mode_id=mode_id,
                            id=group_id
                        ).first()
                        
                        if not first_jadwal:
                            print(f"   ❌ Jadwal ID {group_id} tidak ditemukan!")
                            continue
                        
                        group_name = first_jadwal.group_name
                        print(f"   ✅ Group name: '{group_name}'")
                        
                        # ✅ STEP 2: Ambil SEMUA jadwal dalam group yang sama
                        all_jadwal_in_group = ModeJamKerjaJadwal.objects.filter(
                            mode_id=mode_id,
                            group_name=group_name
                        ).order_by('hari')
                        
                        if not all_jadwal_in_group.exists():
                            print(f"   ❌ No jadwal found for group '{group_name}'!")
                            continue
                        
                        # ✅ STEP 3: Build jadwal_per_hari lengkap
                        built_jadwal = {}
                        for jdl in all_jadwal_in_group:
                            built_jadwal[str(jdl.hari)] = jdl.id
                            print(f"      • Hari {jdl.hari}: Jadwal ID {jdl.id}")
                        
                        print(f"   ✅ Built jadwal_per_hari: {built_jadwal} ({len(built_jadwal)} hari)")
                        
                        if not built_jadwal:
                            print(f"   ❌ Built jadwal kosong!")
                            continue
                        
                        # ✅ STEP 4: Create assignment
                        PegawaiModeAssignment.objects.create(
                            pegawai=new_pegawai,
                            mode_id=mode_id,
                            jadwal_per_hari=built_jadwal,
                            is_active=True
                        )
                        assignment_count += 1
                        print(f"   ✅ Assignment created dengan {len(built_jadwal)} hari!")
                
                except MasterModeJamKerja.DoesNotExist:
                    print(f"   ❌ Mode ID {mode_id} tidak ditemukan!")
                    continue
                except Exception as e:
                    import traceback
                    error_detail = traceback.format_exc()
                    print(f"   ❌ Error: {str(e)}")
                    print(f"   Debug: {error_detail}")
                    continue
            
            print(f"\n{'='*60}")
            print(f"📊 TOTAL ASSIGNMENTS CREATED: {assignment_count}")
            print(f"{'='*60}\n")
            
            # ========================================
            # 7. SYNC FINGERPRINT DARI MESIN
            # ========================================
            
            fingers_count = 0
            try:
                conn = connect_to_fingerprint_machine(mesin.ip_address, mesin.port, timeout=15)
                conn.disable_device()
                
                all_templates = conn.get_templates()
                all_users = conn.get_users()
                
                target_user = next(
                    (u for u in all_users if str(getattr(u, 'user_id', u.uid)) == str(new_pegawai.userid)),
                    None
                )
                
                if target_user:
                    user_templates = [t for t in all_templates if t.uid == target_user.uid]
                    
                    for template in user_templates:
                        FingerprintTemplate.objects.create(
                            pegawai=new_pegawai,
                            uid=template.uid,
                            fid=template.fid,
                            size=template.size,
                            valid=template.valid,
                            template=template.template
                        )
                        fingers_count += 1
                    
                    conn.enable_device()
                    conn.disconnect()
                else:
                    conn.enable_device()
                    conn.disconnect()
                    
                    return JsonResponse({
                        "status": "error",
                        "msg": f"User {new_pegawai.userid} tidak ditemukan di mesin!"
                    }, status=400)
                
            except Exception as e:
                transaction.set_rollback(True)
                return JsonResponse({
                    "status": "error",
                    "msg": f"Gagal sync fingerprint: {str(e)}"
                }, status=500)
            
            # ========================================
            # 8. VALIDASI FINGERPRINT
            # ========================================
            
            if fingers_count == 0:
                transaction.set_rollback(True)
                return JsonResponse({
                    "status": "error",
                    "msg": "❌ Tidak ada data fingerprint!"
                }, status=400)
            
            # ========================================
            # 9. SUCCESS MESSAGE
            # ========================================
            
            msg = f"✅ Pegawai berhasil didaftarkan!\n\n"
            msg += f"Detail:\n"
            msg += f"• Nama: {new_pegawai.nama_lengkap}\n"
            msg += f"• User ID: {new_pegawai.userid}\n"
            msg += f"• Cabang: {cabang_aktif.nama}\n"
            msg += f"• Machine UID: {new_pegawai.uid_mesin}\n"
            msg += f"• Fingerprint: {fingers_count} templates\n"
            msg += f"• Mode Assignments: {assignment_count} mode\n"
            
            if assignment_count > 0:
                msg += f"\n💡 Silakan cek detail pegawai untuk melihat jadwal lengkap!"
            
            return JsonResponse({
                "status": "success",
                "msg": msg,
                "pegawai_id": new_pegawai.id,
                "userid": new_pegawai.userid,
                "fingers_count": fingers_count,
                "assignment_count": assignment_count
            })
            
        except MasterMesin.DoesNotExist:
            return JsonResponse({
                "status": "error",
                "msg": "Mesin tidak ditemukan"
            }, status=404)
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"\n❌ CRITICAL ERROR:\n{error_detail}")
            
            transaction.set_rollback(True)
            return JsonResponse({
                "status": "error", 
                "msg": f"Error: {str(e)}"
            }, status=500)
                
def _validate_pegawai_data(post_data):
    """
    Helper untuk memvalidasi data pegawai dari POST.
    
    Returns:
        JsonResponse jika ada error, None jika valid
    """
    userid = post_data.get("userid_mesin")
    nama = post_data.get("nama_lengkap")
    dept_id = post_data.get("departemen")
    jabatan_id = post_data.get("jabatan")
    
    if not all([userid, nama, dept_id, jabatan_id]):
        return JsonResponse(
            {"status": "error", "msg": "Data wajib tidak lengkap"},
            status=400
        )
    
    if Pegawai.objects.filter(userid=userid).exists():
        return JsonResponse(
            {"status": "error", "msg": f"User ID {userid} sudah terdaftar"},
            status=400
        )
    
    return None


def _create_pegawai_from_machine(post_data, mesin):
    """
    Helper untuk membuat objek pegawai dari data mesin.
    
    Args:
        post_data: Data dari request.POST
        mesin: Object MasterMesin
    
    Returns:
        Pegawai object (belum di-save)
    """
    userid = post_data.get("userid_mesin")
    nama = post_data.get("nama_lengkap")
    
    departemen = MasterDepartemen.objects.get(id=post_data.get("departemen"))
    jabatan = MasterJabatan.objects.get(id=post_data.get("jabatan"))
    
    tanggal_lahir = post_data.get("tanggal_lahir")
    tanggal_bergabung = post_data.get("tanggal_bergabung")
    
    cabang_aktif = mesin.cabang if hasattr(mesin, 'cabang') else None
    
    return Pegawai(
        userid=userid,
        nama_lengkap=nama,
        email=post_data.get("email", ""),
        tanggal_lahir=datetime.strptime(tanggal_lahir, '%Y-%m-%d').date() if tanggal_lahir else None,
        nomor_hp=post_data.get("nomor_hp", ""),
        alamat=post_data.get("alamat", ""),
        departemen=departemen,
        jabatan=jabatan,
        mesin=mesin,
        cabang=cabang_aktif,
        tanggal_bergabung=datetime.strptime(tanggal_bergabung, '%Y-%m-%d').date()
            if tanggal_bergabung else timezone.now().date(),
        is_active=True
    )

## FUNGSI VALIDASI & GENERASI USERID
# Kode ini untuk mengecek ketersediaan dan membuat User ID pegawai secara otomatis.

@login_required
def cek_userid_tersedia(request):
    """Endpoint AJAX: Memeriksa ketersediaan User ID di database dan mesin."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    if request.headers.get("X-Requested-With") != "XMLHttpRequest":
        return JsonResponse({"status": "error", "msg": "Permintaan tidak valid"}, status=400)
    
    try:
        userid = request.POST.get('userid', '').strip()
        mesin_id = request.POST.get('mesin_id')
        
        if not userid or not userid.isdigit():
            return JsonResponse({
                "status": "error",
                "available": False,
                "msg": "User ID harus berupa angka"
            }, status=400)
        
        if Pegawai.objects.filter(userid=userid).exists():
            return JsonResponse({
                "status": "error",
                "available": False,
                "msg": f"User ID {userid} sudah terdaftar di database"
            })
        
        if mesin_id:
            machine_check = _check_userid_in_machine(userid, mesin_id)
            return machine_check
        else:
            return JsonResponse({
                "status": "success",
                "available": True,
                "msg": f"User ID {userid} tersedia di database"
            })
    
    except Exception as e:
        return JsonResponse({"status": "error", "msg": str(e)}, status=500)


def _check_userid_in_machine(userid, mesin_id):
    """Helper untuk memeriksa ketersediaan userid di mesin."""
    try:
        mesin = MasterMesin.objects.get(id=mesin_id, is_active=True)
        conn = connect_to_fingerprint_machine(mesin.ip_address, mesin.port, timeout=5)
        users = conn.get_users()
        
        existing_user = next(
            (u for u in users if str(getattr(u, 'user_id', u.uid)) == str(userid)),
            None
        )
        
        conn.disconnect()
        
        if existing_user:
            return JsonResponse({
                "status": "warning",
                "available": False,
                "msg": f"User ID {userid} sudah ada di mesin {mesin.nama} (Nama: {existing_user.name})"
            })
        else:
            return JsonResponse({
                "status": "success",
                "available": True,
                "msg": f"User ID {userid} tersedia"
            })
    
    except MasterMesin.DoesNotExist:
        return JsonResponse(
            {"status": "error", "msg": "Mesin tidak ditemukan"},
            status=404
        )
    except Exception:
        return JsonResponse({
            "status": "success",
            "available": True,
            "msg": f"User ID {userid} tersedia (Mesin tidak terhubung)"
        })


@login_required
def generate_userid_otomatis(request):
    """Endpoint AJAX: Membuat User ID secara otomatis berdasarkan departemen."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        departemen_id = request.GET.get('departemen_id') or request.POST.get('departemen_id')
        
        if not departemen_id:
            return JsonResponse(
                {"status": "error", "msg": "Departemen wajib dipilih"},
                status=400
            )
        
        departemen = MasterDepartemen.objects.get(id=departemen_id)
        
        if not departemen.id_departemen:
            return JsonResponse(
                {"status": "error", "msg": f"Departemen '{departemen.nama}' belum memiliki ID Departemen."},
                status=400
            )
        
        next_userid = departemen.generate_next_userid()
        total_pegawai = departemen.get_jumlah_pegawai()
        
        return JsonResponse({
            "status": "success",
            "userid": next_userid,
            "departemen_id": departemen.id_departemen,
            "departemen_nama": departemen.nama,
            "total_pegawai": total_pegawai,
            "msg": f"User ID: {next_userid} (Departemen {departemen.nama})"
        })
    
    except Exception as e:
        return JsonResponse({"status": "error", "msg": str(e)}, status=500)


## FUNGSI SINKRONISASI MESIN
# Kode ini untuk mengelola sinkronisasi data pegawai (UID) ke mesin absensi.

@login_required
def register_Pegawai_ke_mesin(request):
    """
    Menampilkan daftar pegawai yang statusnya PENDING (belum sync ke mesin).
    
     Filter cabang diterapkan
    """
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    cabang_aktif = get_active_cabang(request)
    
    #  Filter pegawai berdasarkan cabang
    pegawai_list = Pegawai.objects.filter(
        is_active=True
    ).filter(
        Q(uid_mesin__isnull=True) | Q(uid_mesin=0)
    ).select_related('departemen', 'jabatan', 'cabang').order_by('userid')
    
    #  Filter berdasarkan cabang aktif
    if cabang_aktif:
        pegawai_list = pegawai_list.filter(cabang=cabang_aktif)
    
    #  Filter mesin berdasarkan cabang aktif
    mesin_list = MasterMesin.objects.filter(is_active=True).select_related('cabang')
    
    if cabang_aktif:
        mesin_list = mesin_list.filter(cabang=cabang_aktif)
    
    mesin_list = mesin_list.order_by('nama')
    
    if not mesin_list.exists():
        messages.warning(
            request,
            f"Tidak ada mesin aktif di cabang {cabang_aktif.nama if cabang_aktif else 'yang dipilih'}. "
            f"Silakan tambahkan mesin terlebih dahulu di menu Pengaturan."
        )
    
    context = {
        'pegawai_list': pegawai_list,
        'mesin_list': mesin_list,
        'total_pegawai': pegawai_list.count(),
        'cabang_aktif': cabang_aktif,
    }
    
    return render(request, 'absensi_app/register/sinkron_ke_mesin.html', context)


@login_required
def daftarkan_Pegawai_ke_mesin(request):
    """Endpoint AJAX: Mendaftarkan satu pegawai ke mesin sidik jari."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({"status": "error", "msg": "Method tidak diizinkan"}, status=405)
    
    try:
        pegawai_id = request.POST.get('pegawai_id') or request.POST.get('Pegawai_id')
        ip_address = request.POST.get('ip_address')
        if not ip_address:
            return JsonResponse({
                "status": "error",
                "msg": "IP address mesin wajib diisi"
            }, status=400)
        
        if not pegawai_id:
            return JsonResponse(
                {"status": "error", "msg": "Pegawai ID tidak ditemukan"},
                status=400
            )
        
        pegawai_obj = Pegawai.objects.get(id=pegawai_id)
        
        if pegawai_obj.uid_mesin and pegawai_obj.uid_mesin > 0:
            return JsonResponse({
                "status": "error",
                "msg": f"Pegawai sudah terdaftar (UID: {pegawai_obj.uid_mesin})"
            }, status=400)
        
        conn = connect_to_fingerprint_machine(ip_address)
        uid = register_Pegawai_to_machine(
            conn,
            pegawai_obj.userid,
            pegawai_obj.nama_lengkap
        )
        
        if uid and uid > 0:
            pegawai_obj.uid_mesin = uid
            pegawai_obj.save()
        else:
            raise Exception(f"UID tidak valid: {uid}")
        
        conn.disconnect()
        
        return JsonResponse({
            "status": "success",
            "msg": f"Pegawai {pegawai_obj.nama_lengkap} berhasil didaftarkan",
            "uid": uid,
            "pegawai_id": pegawai_obj.id,
            "userid": pegawai_obj.userid,
            "uid_mesin": pegawai_obj.uid_mesin
        })
    
    except Pegawai.DoesNotExist:
        return JsonResponse(
            {"status": "error", "msg": "Pegawai tidak ditemukan"},
            status=404
        )
    except Exception as e:
        return JsonResponse({"status": "error", "msg": str(e)}, status=500)


@login_required
def ambil_user_dari_mesin(request):
    """Endpoint AJAX: Mengambil daftar user dari mesin (yang belum ada di DB)."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    if request.headers.get("X-Requested-With") != "XMLHttpRequest":
        return JsonResponse({"status": "error", "msg": "Permintaan tidak valid"}, status=400)
    
    try:
        mesin_id = request.POST.get("mesin_id")
        
        if not mesin_id:
            return JsonResponse(
                {"status": "error", "msg": "Pilih mesin terlebih dahulu"},
                status=400
            )
        
        mesin = get_object_or_404(MasterMesin, id=mesin_id, is_active=True)
        
        conn = connect_to_fingerprint_machine(mesin.ip_address, mesin.port)
        users = conn.get_users()
        existing_userids = set(Pegawai.objects.values_list('userid', flat=True))
        
        available_users = []
        for user in users:
            userid = str(getattr(user, 'user_id', user.uid))
            
            if userid not in existing_userids:
                available_users.append({
                    'uid': user.uid,
                    'userid': userid,
                    'name': user.name,
                    'privilege': 'Admin' if user.privilege == (const.USER_ADMIN if const else 14) else 'User'
                })
        
        conn.disconnect()
        
        return JsonResponse({
            "status": "success",
            "data": available_users,
            "total": len(available_users),
            "mesin": {
                "id": mesin.id,
                "nama": mesin.nama,
                "ip": mesin.ip_address,
                "port": mesin.port,
                "cabang": mesin.cabang.nama if mesin.cabang else '-'
            }
        })
    
    except Exception as e:
        return JsonResponse({"status": "error", "msg": str(e)}, status=500)


@login_required
def sync_fingerprint_from_machine(request):
    """Endpoint AJAX: Sinkronisasi template sidik jari dari mesin ke database."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        cabang_aktif = get_active_cabang(request)

        # ✅ Cek mesin_id dari POST dulu, fallback ke mesin pertama di cabang
        mesin_id = request.POST.get('mesin_id')
        if mesin_id:
            mesin = get_object_or_404(MasterMesin, id=mesin_id, is_active=True)
        else:
            mesin = MasterMesin.objects.filter(
                is_active=True,
                cabang=cabang_aktif
            ).first() if cabang_aktif else MasterMesin.objects.filter(is_active=True).first()

        if not mesin:
            return JsonResponse({
                "status": "error",
                "msg": "Tidak ada mesin aktif di cabang ini"
            }, status=404)

        ip_address = request.POST.get('ip_address') or mesin.ip_address
        port = mesin.port or 4370

        pegawai_id = request.POST.get('pegawai_id')

        if pegawai_id:
            pegawai_list = [get_object_or_404(Pegawai, id=pegawai_id)]
        else:
            pegawai_with_fp = get_pegawai_with_fingerprint()

            pegawai_query = Pegawai.objects.filter(
                is_active=True
            ).exclude(id__in=pegawai_with_fp)

            if cabang_aktif:
                pegawai_query = pegawai_query.filter(cabang=cabang_aktif)

            pegawai_list = list(pegawai_query)

        if not pegawai_list:
            return JsonResponse({
                "status": "info",
                "msg": "Tidak ada pegawai yang perlu disinkronisasi",
                "synced_count": 0
            })

        conn = connect_to_fingerprint_machine(ip_address, port=port)
        conn.disable_device()
        all_templates = conn.get_templates()
        all_users = conn.get_users()
        conn.enable_device()

        synced_count = 0
        new_templates_count = 0
        failed_list = []

        for pegawai in pegawai_list:
            try:
                target_user = next(
                    (u for u in all_users if str(getattr(u, 'user_id', u.uid)) == str(pegawai.userid)),
                    None
                )

                if not target_user:
                    failed_list.append({
                        'userid': pegawai.userid,
                        'nama': pegawai.nama_lengkap,
                        'reason': 'User tidak ditemukan di mesin'
                    })
                    continue

                if not pegawai.uid_mesin or pegawai.uid_mesin != target_user.uid:
                    pegawai.uid_mesin = target_user.uid
                    pegawai.save(update_fields=['uid_mesin'])

                pegawai_templates = [t for t in all_templates if t.uid == target_user.uid]

                if pegawai_templates:
                    if pegawai.fingerprint_templates.count() > 0:
                        pegawai.fingerprint_templates.all().delete()

                    for template in pegawai_templates:
                        FingerprintTemplate.objects.create(
                            pegawai=pegawai,
                            uid=template.uid,
                            fid=template.fid,
                            size=template.size,
                            valid=template.valid,
                            template=template.template
                        )
                        new_templates_count += 1

                    synced_count += 1
                else:
                    failed_list.append({
                        'userid': pegawai.userid,
                        'nama': pegawai.nama_lengkap,
                        'reason': 'Sidik jari belum discan di mesin'
                    })
            except Exception as e:
                failed_list.append({
                    'userid': pegawai.userid,
                    'nama': pegawai.nama_lengkap,
                    'reason': str(e)
                })
                continue

        conn.disconnect()

        msg = f"Berhasil sinkronisasi {synced_count} pegawai dengan {new_templates_count} template sidik jari"

        if failed_list:
            msg += f"\n\nGagal sinkronisasi {len(failed_list)} pegawai:\n"
            for item in failed_list[:5]:
                msg += f"- {item['nama']} ({item['userid']}): {item['reason']}\n"
            if len(failed_list) > 5:
                msg += f"... dan {len(failed_list) - 5} pegawai lainnya"

        return JsonResponse({
            "status": "success",
            "msg": msg,
            "synced_count": synced_count,
            "templates_count": new_templates_count,
            "failed_count": len(failed_list),
            "failed_details": failed_list
        })

    except Exception as e:
        return JsonResponse({"status": "error", "msg": str(e)}, status=500)

@login_required
def batalkan_pegawai_pending(request):
    """Endpoint AJAX: Menghapus pegawai yang statusnya PENDING (belum sync ke mesin) secara permanen."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({"status": "error", "msg": "Method tidak diizinkan"}, status=405)
    
    try:
        pegawai_id = request.POST.get('pegawai_id')
        
        if not pegawai_id:
            return JsonResponse({
                "status": "error",
                "msg": "Pegawai ID tidak ditemukan"
            }, status=400)
        
        pegawai_obj = get_object_or_404(Pegawai, id=pegawai_id)
        
        if pegawai_obj.uid_mesin is not None and pegawai_obj.uid_mesin > 0:
            return JsonResponse({
                "status": "error",
                "msg": f"Tidak bisa dibatalkan! {pegawai_obj.nama_lengkap} sudah terdaftar di mesin (UID: {pegawai_obj.uid_mesin})."
            }, status=400)
        
        pegawai_nama = pegawai_obj.nama_lengkap
        pegawai_userid = pegawai_obj.userid
        pegawai_dept = pegawai_obj.departemen.nama if pegawai_obj.departemen else '-'
        
        try:
            with transaction.atomic():
                # ✅ Hapus relasi-relasi yang ada
                pegawai_obj.absensi.all().delete()
                pegawai_obj.fingerprint_templates.all().delete()
                pegawai_obj.mode_assignments.all().delete()
                
                # ❌ HAPUS BARIS INI (karena relasi tidak ada)
                # pegawai_obj.excluded_from_modes.all().delete()
                
                # Hapus pegawai
                pegawai_obj.delete()
            
            return JsonResponse({
                "status": "success",
                "msg": f"✅ {pegawai_nama} berhasil dihapus PERMANEN!\n\n"
                       f"Detail:\n"
                       f"• User ID: {pegawai_userid}\n"
                       f"• Departemen: {pegawai_dept}\n"
                       f"• Status: PENDING (belum sync)\n\n"
                       f"Data telah dihapus dari database dan tidak dapat dikembalikan.",
                "pegawai_id": pegawai_id,
                "userid": pegawai_userid,
                "nama": pegawai_nama
            })
        
        except Exception as e:
            return JsonResponse({
                "status": "error",
                "msg": f"❌ Error saat menghapus: {str(e)}"
            }, status=500)
            
    except Pegawai.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "msg": "Pegawai tidak ditemukan"
        }, status=404)
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "msg": f"Error: {str(e)}"
        }, status=500)

## FUNGSI SINKRONISASI ABSENSI
# Kode ini untuk mengambil data absensi (tap log) dari mesin dan menyimpannya ke database.

@login_required
def sync_absensi(request):
    """Endpoint AJAX: Mengambil data absensi dari mesin dan memprosesnya (sinkronisasi)."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    cabang_aktif = get_active_cabang(request)
    
    lock_id = "sync_absensi_in_progress"
    lock_value = cache.get(lock_id)
    
    if lock_value:
        sync_by = lock_value.get('username', 'admin lain')
        started_at = lock_value.get('started_at', 'baru saja')
        return JsonResponse({
            "status": "info",
            "msg": f"Sinkronisasi sedang berjalan oleh {sync_by} (dimulai {started_at})\n\nHarap tunggu."
        })
    
    lock_data = {
        'username': request.user.username,
        'started_at': timezone.now().strftime('%H:%M:%S')
    }
    cache.set(lock_id, lock_data, 600)
    
    try:
        date_filter = _parse_date_filter(request)
        if isinstance(date_filter, JsonResponse):
            cache.delete(lock_id)
            return date_filter
        
        tanggal_mulai, tanggal_akhir = date_filter
        
        mesin_list = MasterMesin.objects.filter(is_active=True)
        
        if cabang_aktif:
            mesin_list = mesin_list.filter(cabang=cabang_aktif)
        
        if not mesin_list.exists():
            cache.delete(lock_id)
            return JsonResponse({
                "status": "error",
                "msg": f"Tidak ada mesin aktif di cabang {cabang_aktif.nama if cabang_aktif else 'yang dipilih'}"
            }, status=404)
        
        mode_info = WorkModeService.get_mode_today()
        
        all_attendances = []
        mesin_results = []
        
        for mesin in mesin_list:
            result = _fetch_attendance_from_machine(mesin, tanggal_mulai, tanggal_akhir)
            mesin_results.append(result)
            
            if result['status'] == 'success' and result['attendances']:
                all_attendances.extend(result['attendances'])
        
        if not all_attendances:
            cache.delete(lock_id)
            summary = _create_no_data_summary(mesin_list, mesin_results)
            return JsonResponse({
                "status": "info",
                "msg": summary,
                "stats": {
                    "created": 0,
                    "updated": 0,
                    "skipped": 0,
                    "total_machines": len(mesin_list),
                    "online_machines": sum(1 for m in mesin_results if m['status'] == 'success'),
                    "offline_machines": sum(1 for m in mesin_results if m['status'] == 'error')
                }
            })
        
        stats = _process_attendance_data(all_attendances)
        
        msg = _create_sync_success_message(
            mesin_list,
            mesin_results,
            stats,
            tanggal_mulai,
            tanggal_akhir
        )
        
        if mode_info and mode_info.get('nama_mode'):
            msg += f"\n\nMode Jam Kerja: {mode_info['nama_mode']}"
            if mode_info.get('nama_periode'):
                msg += f"\nPeriode: {mode_info['nama_periode']}"
            
        cache.delete(lock_id)
        
        return JsonResponse({
            "status": "success",
            "msg": msg,
            "stats": {
                "created": stats['created'],
                "updated": stats['updated'],
                "skipped": stats['skipped'],
                "total_machines": len(mesin_list),
                "online_machines": sum(1 for m in mesin_results if m['status'] == 'success'),
                "offline_machines": sum(1 for m in mesin_results if m['status'] == 'error'),
                "filter": {
                    "tanggal_mulai": tanggal_mulai.strftime('%Y-%m-%d') if tanggal_mulai else None,
                    "tanggal_akhir": tanggal_akhir.strftime('%Y-%m-%d') if tanggal_akhir else None
                },
                "mode": {
                "nama": mode_info.get('nama_mode', 'N/A'),
                "kode": mode_info.get('mode_kode', 'N/A'),
                "periode": mode_info.get('nama_periode')
            }
            }
        })
        
    except Exception as e:
        cache.delete(lock_id)
        return JsonResponse({
            "status": "error",
            "msg": f"Sinkronisasi gagal: {str(e)}"
        }, status=500)


def _parse_date_filter(request):
    """Helper untuk mem-parsing dan memvalidasi filter tanggal."""
    tanggal_mulai = request.GET.get('tanggal_mulai')
    tanggal_akhir = request.GET.get('tanggal_akhir')
    
    if tanggal_mulai:
        try:
            tanggal_mulai = datetime.strptime(tanggal_mulai, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({
                "status": "error",
                "msg": "Format tanggal_mulai tidak valid (gunakan YYYY-MM-DD)"
            }, status=400)
    
    if tanggal_akhir:
        try:
            tanggal_akhir = datetime.strptime(tanggal_akhir, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({
                "status": "error",
                "msg": "Format tanggal_akhir tidak valid (gunakan YYYY-MM-DD)"
            }, status=400)
    
    if tanggal_mulai and tanggal_akhir and tanggal_mulai > tanggal_akhir:
        return JsonResponse({
            "status": "error",
            "msg": "Tanggal mulai tidak boleh lebih besar dari tanggal akhir"
        }, status=400)
    
    return tanggal_mulai, tanggal_akhir


def _fetch_attendance_from_machine(mesin, tanggal_mulai, tanggal_akhir):
    """Helper untuk mengambil data absensi dari satu mesin."""
    mesin_result = {
        'nama': mesin.nama,
        'ip': mesin.ip_address,
        'cabang': mesin.cabang.nama if mesin.cabang else '-',
        'status': 'processing',
        'attendances': []
    }
    
    try:
        conn = connect_to_fingerprint_machine(
            mesin.ip_address,
            mesin.port,
            timeout=15
        )
        conn.disable_device()
        
        attendances = conn.get_attendance()
        
        conn.enable_device()
        conn.disconnect()
        
        if not attendances:
            mesin_result['status'] = 'no_data'
            mesin_result['msg'] = 'Tidak ada data di mesin'
            mesin_result['total_records'] = 0
            return mesin_result
        
        if tanggal_mulai or tanggal_akhir:
            filtered = []
            for att in attendances:
                att_date = att.timestamp.date()
                
                if tanggal_mulai and att_date < tanggal_mulai:
                    continue
                if tanggal_akhir and att_date > tanggal_akhir:
                    continue
                
                filtered.append(att)
            
            attendances = filtered
        
        if not attendances:
            mesin_result['status'] = 'no_data_in_range'
            mesin_result['msg'] = 'Tidak ada data di periode yang dipilih'
            mesin_result['total_records'] = 0
            return mesin_result
        
        mesin_result['attendances'] = attendances
        mesin_result['status'] = 'success'
        mesin_result['total_records'] = len(attendances)
        mesin_result['msg'] = f'Berhasil mengambil {len(attendances)} data'
        
        return mesin_result
    
    except Exception as e:
        mesin_result['status'] = 'error'
        mesin_result['msg'] = str(e)
        mesin_result['total_records'] = 0
        return mesin_result


def _process_attendance_data(all_attendances):
    """Memproses data absensi dan menyimpannya ke database."""
    grouped_data = {}
    for att in all_attendances:
        key = (att.timestamp.date(), str(att.user_id))
        if key not in grouped_data:
            grouped_data[key] = []
        grouped_data[key].append(att)
    
    created_count = 0
    updated_count = 0
    skipped_count = 0
    
    skipped_details = {
        'pegawai_not_found': [],
        'no_tap_masuk': [],
        'other_errors': []
    }
    
    for (tanggal, userid), records in grouped_data.items():
        try:
            pegawai = Pegawai.objects.filter(userid=userid).first()
            
            if not pegawai:
                skipped_count += 1
                skipped_details['pegawai_not_found'].append({
                    'userid': userid,
                    'tanggal': tanggal.strftime('%Y-%m-%d'),
                    'total_tap': len(records)
                })
                continue
            
            attendance_data = _calculate_attendance_times(records, tanggal, pegawai)
            
            if not attendance_data:
                skipped_count += 1
                skipped_details['no_tap_masuk'].append({
                    'userid': userid,
                    'nama': pegawai.nama_lengkap,
                    'tanggal': tanggal.strftime('%Y-%m-%d'),
                    'total_tap': len(records)
                })
                continue
            
            # Keterangan diisi manual oleh admin, bukan otomatis
            attendance_data['keterangan'] = ''
            
            with transaction.atomic():
                absensi, created = Absensi.objects.update_or_create(
                    pegawai=pegawai,
                    tanggal=tanggal,
                    defaults=attendance_data
                )
            
            if created:
                created_count += 1
            else:
                updated_count += 1
        
        except Exception as e:
            skipped_count += 1
            skipped_details['other_errors'].append({
                'userid': userid,
                'tanggal': tanggal.strftime('%Y-%m-%d'),
                'error': str(e)
            })
            continue
    
    return {
        "created": created_count,
        "updated": updated_count,
        "skipped": skipped_count,
        "skipped_details": skipped_details
    }


def _calculate_attendance_times(records, tanggal, pegawai):
    """Menghitung waktu absensi dari log mesin (masuk/pulang/istirahat)."""
    sorted_records = sorted(records, key=lambda x: x.timestamp)
    
    # Pisahkan berdasarkan punch type
    punch_0 = [r for r in sorted_records if r.punch == 0]  
    punch_1 = [r for r in sorted_records if r.punch == 1] 
    punch_2 = [r for r in sorted_records if r.punch == 2] 
    punch_3 = [r for r in sorted_records if r.punch == 3] 
    
    # Ambil waktu tap dari mesin
    tap_masuk = punch_0[0].timestamp.time() if punch_0 else None
    tap_pulang = punch_1[-1].timestamp.time() if punch_1 else None
    tap_istirahat_keluar = punch_2[0].timestamp.time() if punch_2 else None
    tap_istirahat_masuk = punch_3[0].timestamp.time() if punch_3 else None
    
    if not tap_masuk:
        return None
    
    # Validasi keterlambatan/pulang cepat
    validation = _validate_attendance(tap_masuk, tap_pulang, tanggal, pegawai)
    
    return {
        'tap_masuk': tap_masuk,
        'tap_pulang': tap_pulang,
        'tap_istirahat_keluar': tap_istirahat_keluar,
        'tap_istirahat_masuk': tap_istirahat_masuk,
        'status': 'Hadir',
        'is_late': validation['is_late'],
        'is_early_departure': validation['is_early'],
        'keterangan': '',  \
        'updated_at': timezone.now()
    }

def _validate_attendance(tap_masuk, tap_pulang, tanggal, pegawai):
    """
    ✅ FIXED: Validasi absensi berdasarkan MODE AKTIF pada tanggal
    
    Returns:
        dict: {is_late, is_early, keterangan}
    """
    from .services import LayananModeKerja
    from datetime import datetime, timedelta
    
    is_late = False
    is_early = False
    keterangan = []
    
    # ✅ STEP 1: Ambil mode aktif pada tanggal ini
    mode_info = LayananModeKerja.ambil_mode_aktif(tanggal)
    
    if not mode_info or not mode_info['mode']:
        return {
            'is_late': False,
            'is_early': False,
            'keterangan': ''
        }
    
    mode = mode_info['mode']
    periode = mode_info['periode']
    
    # ✅ STEP 2: Cek apakah hari kerja
    if not LayananModeKerja.cek_hari_kerja(pegawai, tanggal):
        return {
            'is_late': False,
            'is_early': False,
            'keterangan': f'Hari libur ({mode.nama})'
        }
    
    # ✅ STEP 3: Ambil jadwal pegawai
    jadwal_info = LayananModeKerja.ambil_jadwal_pegawai(pegawai, tanggal)
    
    if not jadwal_info or not jadwal_info.get('jadwal'):
        return {
            'is_late': False,
            'is_early': False,
            'keterangan': ''
        }
    
    jadwal = jadwal_info['jadwal']
    
    # ✅ STEP 4: Validasi TERLAMBAT
    if jadwal.jam_masuk and tap_masuk:
        jam_masuk_batas = (
            datetime.combine(tanggal, jadwal.jam_masuk) +
            timedelta(minutes=jadwal.toleransi_terlambat)
        ).time()
        
        if tap_masuk > jam_masuk_batas:
            is_late = True
            selisih = datetime.combine(tanggal, tap_masuk) - datetime.combine(tanggal, jadwal.jam_masuk)
            menit_terlambat = int(selisih.total_seconds() / 60)
            keterangan.append(f"Terlambat {menit_terlambat} menit")
    
    # ✅ STEP 5: Validasi PULANG CEPAT
    if jadwal.jam_keluar and tap_pulang:
        jam_keluar_batas = (
            datetime.combine(tanggal, jadwal.jam_keluar) -
            timedelta(minutes=jadwal.toleransi_pulang_cepat)
        ).time()
        
        if tap_pulang < jam_keluar_batas:
            is_early = True
            selisih = datetime.combine(tanggal, jadwal.jam_keluar) - datetime.combine(tanggal, tap_pulang)
            menit_lebih_cepat = int(selisih.total_seconds() / 60)
            keterangan.append(f"Pulang {menit_lebih_cepat} menit lebih cepat")
    
    # ✅ STEP 6: Tambahkan info mode khusus
    if periode:
        keterangan.append(f"Mode: {mode.nama} ({periode.nama})")
    
    return {
        'is_late': is_late,
        'is_early': is_early,
        'keterangan': ' | '.join(keterangan)
    }

def _create_no_data_summary(mesin_list, mesin_results):
    """Helper untuk membuat ringkasan jika tidak ada data absensi dari mesin."""
    summary = f"Tidak ada data absensi dari {len(mesin_list)} mesin:\n\n"
    for mr in mesin_results:
        summary += f"• {mr['nama']} ({mr['cabang']}): {mr['msg']}\n"
    return summary


def _create_sync_success_message(mesin_list, mesin_results, stats, tgl_mulai, tgl_akhir):
    """Helper untuk membuat pesan sukses setelah sinkronisasi absensi."""
    filter_info = ""
    if tgl_mulai and tgl_akhir:
        filter_info = f" (periode: {tgl_mulai} hingga {tgl_akhir})"
    elif tgl_mulai:
        filter_info = f" (dari: {tgl_mulai})"
    elif tgl_akhir:
        filter_info = f" (hingga: {tgl_akhir})"
    
    msg = f"Sinkronisasi selesai dari {len(mesin_list)} mesin{filter_info}\n\n"
    msg += f"Hasil:\n"
    msg += f"• Baru: {stats['created']} data\n"
    msg += f"• Diperbarui: {stats['updated']} data\n"
    msg += f"• Dilewati: {stats['skipped']} data\n\n"
    msg += f"Status Mesin:\n"
    
    success_machines = [m for m in mesin_results if m['status'] == 'success']
    error_machines = [m for m in mesin_results if m['status'] == 'error']
    
    msg += f"• Online: {len(success_machines)} mesin\n"
    if error_machines:
        msg += f"• Offline/Error: {len(error_machines)} mesin\n"
    
    return msg


## FUNGSI MANAJEMEN ABSENSI
# Kode ini untuk menampilkan riwayat absensi dan mengelola input absensi manual oleh admin.


@login_required
def riwayat_absensi(request):
    """Menampilkan riwayat absensi dengan filter tanggal dan pencarian."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    cabang_aktif = get_active_cabang(request)
    mode_info = WorkModeService.get_mode_today()
    
    form = LaporanFilterForm(request.GET)
    
    # ✅ SELECT RELATED
    absensi_list = Absensi.objects.select_related(
        'pegawai',
        'pegawai__departemen',
        'pegawai__jabatan',
        'pegawai__cabang'
    ).all()
    
    # ✅ FILTER CABANG (OPTIONAL)
    if cabang_aktif:
        absensi_list = absensi_list.filter(pegawai__cabang=cabang_aktif)
    
    # ✅ SORTING: Data terbaru di atas
    absensi_list = absensi_list.order_by('-tanggal', '-created_at')
    
    # ✅ FILTER dari form
    if form.is_valid():
        tgl_mulai = form.cleaned_data.get('tanggal_mulai')
        tgl_akhir = form.cleaned_data.get('tanggal_akhir')
        
        if tgl_mulai:
            absensi_list = absensi_list.filter(tanggal__gte=tgl_mulai)
        if tgl_akhir:
            absensi_list = absensi_list.filter(tanggal__lte=tgl_akhir)
        
        search_employee = form.cleaned_data.get('search_employee')
        if search_employee:
            absensi_list = absensi_list.filter(
                Q(pegawai__nama_lengkap__icontains=search_employee) |
                Q(pegawai__userid__icontains=search_employee)
            )
    
    mesin_list = MasterMesin.objects.filter(is_active=True).select_related('cabang')
    if cabang_aktif:
        mesin_list = mesin_list.filter(cabang=cabang_aktif)
    mesin_list = mesin_list.order_by('nama')
    
    # ✅ HITUNG TOTAL JAM KERJA + AMBIL INFO MODE PER ABSENSI
# HITUNG TOTAL JAM KERJA + CEK VIOLATIONS
    absensi_with_stats = []
    
    for absensi in absensi_list:
        total_jam = absensi.calculate_total_jam_kerja()
        mode_pada_tanggal = WorkModeService.get_mode_for_date(absensi.tanggal)

        # Cek violation istirahat melebihi 1 jam
        is_istirahat_violation = False
        if absensi.tap_istirahat_keluar and absensi.tap_istirahat_masuk:
            from datetime import datetime as dt
            keluar = dt.combine(absensi.tanggal, absensi.tap_istirahat_keluar)
            masuk  = dt.combine(absensi.tanggal, absensi.tap_istirahat_masuk)
            durasi_istirahat = (masuk - keluar).total_seconds() / 60
            if durasi_istirahat > 60:
                is_istirahat_violation = True
        elif absensi.tap_masuk and absensi.tap_pulang:
            # Tidak ada tap istirahat sama sekali
            is_istirahat_violation = True

        # Gabungan violation
        has_violation = absensi.is_late or absensi.is_early_departure or is_istirahat_violation

        absensi.total_jam_kerja       = total_jam.get('formatted', '-') if total_jam else '-'
        absensi.has_violation         = has_violation
        absensi.is_istirahat_violation = is_istirahat_violation
        absensi.mode_warna            = mode_pada_tanggal.get('warna_mode', '#6b7280')
        absensi.mode_nama             = mode_pada_tanggal.get('nama_mode', 'Normal')
        absensi.is_mode_khusus        = mode_pada_tanggal.get('is_mode_khusus', False)

        absensi_with_stats.append(absensi)

    context = {
        'absensi_list': absensi_with_stats,
        'form': form,
        'mesin_list': mesin_list,
        'cabang_aktif': cabang_aktif,
        'mode_info': mode_info,
    }

    return render(request, 'absensi_app/absensi/riwayat_absensi.html', context)

@login_required
def absensi_admin(request):
    """Form input absensi manual (izin, sakit, atau kehadiran manual) oleh admin."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    cabang_aktif = get_active_cabang(request)
    
    if request.method == 'POST':
        status = request.POST.get('status')
        pegawai_id = request.POST.get('pegawai')
        keterangan = request.POST.get('keterangan', '').strip()
        
        # ✅ 1. VALIDASI BASIC
        if not pegawai_id:
            messages.error(request, '❌ Pilih pegawai terlebih dahulu!')
            return redirect('absensi_admin')
        
        if not status:
            messages.error(request, '❌ Status absensi wajib dipilih!')
            return redirect('absensi_admin')
        
        try:
            pegawai_obj = Pegawai.objects.get(id=pegawai_id)
        except Pegawai.DoesNotExist:
            messages.error(request, '❌ Pegawai tidak ditemukan!')
            return redirect('absensi_admin')
        
        # ✅ 2. HANDLE STATUS HADIR/ABSEN (SINGLE DATE)
        if status in ['Hadir', 'Absen']:
            tanggal_str = request.POST.get('tanggal')
            
            if not tanggal_str:
                messages.error(request, '❌ Tanggal wajib diisi!')
                return redirect('absensi_admin')
            
            try:
                tanggal = datetime.strptime(tanggal_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, '❌ Format tanggal tidak valid!')
                return redirect('absensi_admin')
            
            # ✅ Cek duplikat
            if Absensi.objects.filter(pegawai=pegawai_obj, tanggal=tanggal).exists():
                messages.warning(
                    request,
                    f'⚠️ Absensi untuk {pegawai_obj.nama_lengkap} pada {tanggal.strftime("%d %b %Y")} sudah ada!'
                )
                return redirect('absensi_admin')
            
            # ✅ Ambil waktu tap
            tap_masuk = None
            tap_pulang = None
            tap_istirahat_keluar = None
            tap_istirahat_masuk = None
            
            if status == 'Hadir':
                tap_masuk_str = request.POST.get('tap_masuk')
                if not tap_masuk_str:
                    messages.error(request, '❌ Waktu masuk wajib diisi untuk status Hadir!')
                    return redirect('absensi_admin')
                
                try:
                    tap_masuk = datetime.strptime(tap_masuk_str, '%H:%M').time()
                    
                    tap_pulang_str = request.POST.get('tap_pulang')
                    if tap_pulang_str:
                        tap_pulang = datetime.strptime(tap_pulang_str, '%H:%M').time()
                    
                    tap_istirahat_keluar_str = request.POST.get('tap_istirahat_keluar')
                    if tap_istirahat_keluar_str:
                        tap_istirahat_keluar = datetime.strptime(tap_istirahat_keluar_str, '%H:%M').time()
                    
                    tap_istirahat_masuk_str = request.POST.get('tap_istirahat_masuk')
                    if tap_istirahat_masuk_str:
                        tap_istirahat_masuk = datetime.strptime(tap_istirahat_masuk_str, '%H:%M').time()
                    
                except ValueError:
                    messages.error(request, '❌ Format waktu tidak valid!')
                    return redirect('absensi_admin')
            
            # ✅ 3. SIMPAN KE DATABASE (TANPA TRANSACTION - LANGSUNG COMMIT)
            try:
                absensi = Absensi(
                    pegawai=pegawai_obj,
                    tanggal=tanggal,
                    status=status,
                    tap_masuk=tap_masuk,
                    tap_pulang=tap_pulang,
                    tap_istirahat_keluar=tap_istirahat_keluar,
                    tap_istirahat_masuk=tap_istirahat_masuk,
                    keterangan=keterangan
                )
                
                # ✅ LANGSUNG SAVE (auto-commit)
                absensi.save()
                
                messages.success(
                    request,
                    f'✅ Absensi {status} untuk {pegawai_obj.nama_lengkap} berhasil disimpan!'
                )
                
                # ✅ Redirect dengan delay kecil
                return redirect('riwayat_absensi')
                
            except Exception as e:
                messages.error(request, f'❌ Error menyimpan: {str(e)}')
                return redirect('absensi_admin')
        
        # ✅ 4. HANDLE STATUS IZIN/SAKIT (MULTIPLE DATES)
        elif status in ['Izin', 'Sakit']:
            tanggal_mulai_str = request.POST.get('tanggal_mulai')
            tanggal_selesai_str = request.POST.get('tanggal_selesai')
            
            if not tanggal_mulai_str or not tanggal_selesai_str:
                messages.error(request, '❌ Tanggal mulai dan selesai wajib diisi!')
                return redirect('absensi_admin')
            
            try:
                tanggal_mulai = datetime.strptime(tanggal_mulai_str, '%Y-%m-%d').date()
                tanggal_selesai = datetime.strptime(tanggal_selesai_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, '❌ Format tanggal tidak valid!')
                return redirect('absensi_admin')
            
            if tanggal_selesai < tanggal_mulai:
                messages.error(request, '❌ Tanggal selesai tidak boleh lebih awal dari tanggal mulai!')
                return redirect('absensi_admin')
            
            # ✅ Ambil tipe
            tipe_izin = request.POST.get('tipe_izin', 'full')
            
            tap_masuk = None
            tap_pulang = None
            
            if tipe_izin == 'half':
                tap_masuk_str = request.POST.get('tap_masuk_half')
                tap_pulang_str = request.POST.get('tap_pulang_half')
                
                if tap_masuk_str:
                    try:
                        tap_masuk = datetime.strptime(tap_masuk_str, '%H:%M').time()
                    except ValueError:
                        pass
                
                if tap_pulang_str:
                    try:
                        tap_pulang = datetime.strptime(tap_pulang_str, '%H:%M').time()
                    except ValueError:
                        pass
            
            # ✅ Simpan untuk setiap tanggal
            created_count = 0
            skipped_count = 0
            
            current_date = tanggal_mulai
            
            try:
                while current_date <= tanggal_selesai:
                    if not Absensi.objects.filter(pegawai=pegawai_obj, tanggal=current_date).exists():
                        absensi = Absensi(
                            pegawai=pegawai_obj,
                            tanggal=current_date,
                            status=status,
                            tap_masuk=tap_masuk,
                            tap_pulang=tap_pulang,
                            keterangan=keterangan
                        )
                        absensi.save()
                        created_count += 1
                    else:
                        skipped_count += 1
                    
                    current_date += timedelta(days=1)
                
                msg = f'✅ Berhasil! {created_count} hari {status} disimpan'
                if skipped_count > 0:
                    msg += f' ({skipped_count} dilewati karena sudah ada)'
                
                messages.success(request, msg)
                return redirect('riwayat_absensi')
                
            except Exception as e:
                messages.error(request, f'❌ Error: {str(e)}')
                return redirect('absensi_admin')
        
        else:
            messages.error(request, '❌ Status tidak valid!')
            return redirect('absensi_admin')
    
    # ✅ GET: Tampilkan form
    form = AbsensiAdminForm()
    
    pegawai_with_fp = get_pegawai_with_fingerprint()
    
    queryset = Pegawai.objects.filter(
        is_active=True,
        uid_mesin__isnull=False,
        id__in=pegawai_with_fp
    ).exclude(uid_mesin=0)
    
    if cabang_aktif:
        queryset = queryset.filter(cabang=cabang_aktif)
    
    form.fields['pegawai'].queryset = queryset.order_by('nama_lengkap')
    
    context = {
        'form': form,
        'cabang_aktif': cabang_aktif,
    }
    
    return render(request, 'absensi_app/absensi/absensi_admin_form.html', context)
    
    # ========================================
    # ✅ GET: Tampilkan form
    # ========================================
    form = AbsensiAdminForm()
    
    pegawai_with_fp = get_pegawai_with_fingerprint()
    
    # ✅ Filter pegawai berdasarkan cabang aktif
    queryset = Pegawai.objects.filter(
        is_active=True,
        uid_mesin__isnull=False,
        id__in=pegawai_with_fp
    ).exclude(uid_mesin=0)
    
    if cabang_aktif:
        queryset = queryset.filter(cabang=cabang_aktif)
    
    form.fields['pegawai'].queryset = queryset.order_by('nama_lengkap')
    
    context = {
        'form': form,
        'cabang_aktif': cabang_aktif,
    }
    
    return render(request, 'absensi_app/absensi/absensi_admin_form.html', context)

## FUNGSI MANAJEMEN MONITORING MESIN
# Kode ini untuk memantau status mesin dan log absensi dari mesin.

@login_required
def monitor_absensi_mesin(request):
    """Menampilkan halaman untuk memantau data absensi dari mesin sidik jari."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    cabang_aktif = get_active_cabang(request)
    
    mesin_list = MasterMesin.objects.filter(is_active=True)
    if cabang_aktif:
        mesin_list = mesin_list.filter(cabang=cabang_aktif)
    mesin_list = mesin_list.order_by('nama')
    
    context = {
        'mesin_list': mesin_list,
        'page_title': 'Monitor Absensi Mesin',
        'cabang_aktif': cabang_aktif,
    }
    
    return render(request, 'absensi_app/data_mesin.html', context)


@login_required
def get_absensi_dari_mesin(request):
    """Endpoint AJAX: Mengambil semua data tap (log) dari satu mesin sidik jari."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        mesin_id = request.GET.get('mesin_id')
        tanggal_mulai = request.GET.get('tanggal_mulai')
        tanggal_akhir = request.GET.get('tanggal_akhir')
        
        if not mesin_id:
            return JsonResponse(
                {"status": "error", "msg": "Parameter mesin_id wajib diisi"},
                status=400
            )
        
        mesin = get_object_or_404(MasterMesin, id=mesin_id, is_active=True)
        
        try:
            conn = connect_to_fingerprint_machine(mesin.ip_address, mesin.port, timeout=10)
        except Exception as e:
            return JsonResponse({
                "status": "error",
                "msg": f"Gagal terhubung ke mesin '{mesin.nama}' ({mesin.ip_address}:{mesin.port}). Error: {str(e)}"
            }, status=500)
        
        try:
            conn.disable_device()
            all_taps = conn.get_attendance()
            all_users = conn.get_users()
            conn.enable_device()
            conn.disconnect()
        except Exception as e:
            try:
                conn.disconnect()
            except:
                pass
            return JsonResponse({
                "status": "error",
                "msg": f"Gagal mengambil data dari mesin '{mesin.nama}'. Error: {str(e)}"
            }, status=500)
        
        user_map = {
            str(getattr(user, 'user_id', user.uid)): user.name
            for user in all_users
        }
        
        filtered_taps = _filter_taps_by_date(all_taps, tanggal_mulai, tanggal_akhir)
        
        filtered_taps.sort(key=lambda x: x.timestamp, reverse=True)
        
        data_list = _build_tap_data_list(filtered_taps, user_map)
        
        return JsonResponse({
            "status": "success",
            "data": data_list,
            "total": len(data_list),
            "mesin": {
                "id": mesin.id,
                "nama": mesin.nama,
                "ip_address": mesin.ip_address,
                "port": mesin.port
            },
            "filter": {
                "tanggal_mulai": tanggal_mulai,
                "tanggal_akhir": tanggal_akhir
            }
        })
    
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "msg": f"Terjadi error: {str(e)}"
        }, status=500)


def _filter_taps_by_date(all_taps, tanggal_mulai, tanggal_akhir):
    """Helper untuk memfilter data tap berdasarkan rentang tanggal."""
    if tanggal_mulai and tanggal_akhir:
        try:
            start = datetime.strptime(f"{tanggal_mulai} 00:00:00", '%Y-%m-%d %H:%M:%S')
            end = datetime.strptime(f"{tanggal_akhir} 23:59:59", '%Y-%m-%d %H:%M:%S')
            return [tap for tap in all_taps if start <= tap.timestamp <= end]
        except ValueError:
            return all_taps
    return all_taps


def _build_tap_data_list(filtered_taps, user_map):
    """Helper untuk membuat daftar data tap (log) untuk respon API."""
    punch_names = {
        0: 'Masuk',
        1: 'Keluar',
        2: 'Istirahat Keluar',
        3: 'Istirahat Masuk',
        4: 'Lembur Masuk',
        5: 'Lembur Keluar'
    }
    
    data_list = []
    for index, tap in enumerate(filtered_taps):
        userid = str(tap.user_id)
        pegawai = Pegawai.objects.filter(userid=userid).first()
        
        if pegawai:
            nama = pegawai.nama_lengkap
            dept = pegawai.departemen.nama if pegawai.departemen else '-'
            is_registered = True
        else:
            nama = user_map.get(userid, f"Tidak Diketahui ({userid})")
            dept = '-'
            is_registered = False
        
        data_list.append({
            'no': index + 1,
            'userid': userid,
            'nama_display': nama,
            'departemen': dept,
            'is_registered': is_registered,
            'tanggal': tap.timestamp.strftime('%Y-%m-%d'),
            'tanggal_display': tap.timestamp.strftime('%d %b %Y'),
            'waktu': tap.timestamp.strftime('%H:%M:%S'),
            'punch_type': punch_names.get(tap.punch, f'Tidak Diketahui ({tap.punch})'),
            'punch_code': tap.punch,
        })
    
    return data_list


@login_required
def check_all_machines_status(request):
    """Endpoint AJAX: Memeriksa status koneksi semua mesin sidik jari."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        cabang_aktif = get_active_cabang(request)
        
        mesin_list = MasterMesin.objects.filter(is_active=True)
        if cabang_aktif:
            mesin_list = mesin_list.filter(cabang=cabang_aktif)
        mesin_list = mesin_list.order_by('nama')
        
        if not mesin_list.exists():
            return JsonResponse(
                {"status": "error", "msg": f"Tidak ada mesin aktif di cabang {cabang_aktif.nama if cabang_aktif else 'yang dipilih'}"},
                status=404
            )
        
        mesin_status_list = []
        total_online = 0
        total_offline = 0
        
        for mesin in mesin_list:
            status_data = _check_single_machine_status(mesin)
            mesin_status_list.append(status_data)
            
            if status_data['online']:
                total_online += 1
            else:
                total_offline += 1
        
        return JsonResponse({
            "status": "success",
            "machines": mesin_status_list,
            "summary": {
                "total": len(mesin_list),
                "online": total_online,
                "offline": total_offline,
                "online_percentage": round((total_online / len(mesin_list)) * 100, 2)
                    if len(mesin_list) > 0 else 0
            },
            "cabang": {
                "id": cabang_aktif.id if cabang_aktif else None,
                "nama": cabang_aktif.nama if cabang_aktif else None
            }
        })
    
    except Exception as e:
        return JsonResponse({"status": "error", "msg": f"Error: {str(e)}"}, status=500)


def _check_single_machine_status(mesin):
    """Helper untuk memeriksa status koneksi satu mesin."""
    status_data = {
        'id': mesin.id,
        'nama': mesin.nama,
        'ip_address': mesin.ip_address,
        'port': mesin.port,
        'cabang': mesin.cabang.nama if mesin.cabang else '-',
        'lokasi': mesin.lokasi or '-',
        'status': 'checking'
    }
    
    try:
        conn = connect_to_fingerprint_machine(mesin.ip_address, mesin.port, timeout=5)
        
        users = conn.get_users()
        attendances = conn.get_attendance()
        
        try:
            firmware = conn.get_firmware_version()
        except:
            firmware = 'N/A'
        
        try:
            serial = conn.get_serialnumber()
        except:
            serial = 'N/A'
        
        conn.disconnect()
        
        status_data.update({
            'status': 'online',
            'online': True,
            'total_users': len(users),
            'total_attendance': len(attendances),
            'firmware': firmware,
            'serial': serial,
            'msg': f'Online - {len(users)} user, {len(attendances)} data'
        })
    
    except Exception as e:
        status_data.update({
            'status': 'offline',
            'online': False,
            'total_users': 0,
            'total_attendance': 0,
            'firmware': 'N/A',
            'serial': 'N/A',
            'msg': f'Offline - {str(e)}',
            'error': str(e)
        })
    
    return status_data


@login_required
def get_absensi_all_machines(request):
    """Endpoint AJAX: Mengambil data absensi dari SEMUA mesin sidik jari."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        tanggal_mulai = request.GET.get('tanggal_mulai')
        tanggal_akhir = request.GET.get('tanggal_akhir')
        
        mesin_list = MasterMesin.objects.filter(is_active=True)
        
        if not mesin_list.exists():
            return JsonResponse({
                "status": "error",
                "msg": "Tidak ada mesin aktif"
            }, status=404)
        
        all_taps = []
        mesin_status = []
        
        for mesin in mesin_list:
            try:
                conn = connect_to_fingerprint_machine(mesin.ip_address, mesin.port, timeout=10)
                conn.disable_device()
                
                taps = conn.get_attendance()
                users = conn.get_users()
                
                conn.enable_device()
                conn.disconnect()
                
                user_map = {
                    str(getattr(u, 'user_id', u.uid)): u.name
                    for u in users
                }
                
                filtered_taps = []
                if tanggal_mulai and tanggal_akhir:
                    start = datetime.strptime(f"{tanggal_mulai} 00:00:00", '%Y-%m-%d %H:%M:%S')
                    end = datetime.strptime(f"{tanggal_akhir} 23:59:59", '%Y-%m-%d %H:%M:%S')
                    filtered_taps = [t for t in taps if start <= t.timestamp <= end]
                else:
                    filtered_taps = taps
                
                punch_names = {
                    0: 'Masuk', 1: 'Keluar', 2: 'Istirahat Keluar', 3: 'Istirahat Masuk',
                    4: 'Lembur Masuk', 5: 'Lembur Keluar'
                }
                
                for tap in filtered_taps:
                    userid = str(tap.user_id)
                    pegawai = Pegawai.objects.filter(userid=userid).first()
                    
                    all_taps.append({
                        'mesin_nama': mesin.nama,
                        'mesin_ip': mesin.ip_address,
                        'cabang': mesin.cabang.nama if mesin.cabang else '-',
                        'userid': userid,
                        'nama_display': pegawai.nama_lengkap if pegawai else user_map.get(userid, f"Tidak Diketahui ({userid})"),
                        'departemen': pegawai.departemen.nama if pegawai and pegawai.departemen else '-',
                        'is_registered': bool(pegawai),
                        'tanggal': tap.timestamp.strftime('%Y-%m-%d'),
                        'tanggal_display': tap.timestamp.strftime('%d %b %Y'),
                        'waktu': tap.timestamp.strftime('%H:%M:%S'),
                        'punch_type': punch_names.get(tap.punch, f'Tidak Diketahui ({tap.punch})'),
                        'punch_code': tap.punch,
                        'timestamp': tap.timestamp.isoformat()
                    })
                
                mesin_status.append({
                    'nama': mesin.nama,
                    'ip': mesin.ip_address,
                    'cabang': mesin.cabang.nama if mesin.cabang else '-',
                    'status': 'online',
                    'total_taps': len(filtered_taps)
                })
                
            except Exception as e:
                mesin_status.append({
                    'nama': mesin.nama,
                    'ip': mesin.ip_address,
                    'cabang': mesin.cabang.nama if mesin.cabang else '-',
                    'status': 'offline',
                    'error': str(e)
                })
        
        all_taps.sort(key=lambda x: x['timestamp'], reverse=True)
        
        return JsonResponse({
            "status": "success",
            "data": all_taps,
            "total": len(all_taps),
            "mesin_status": mesin_status,
            "filter": {
                "tanggal_mulai": tanggal_mulai,
                "tanggal_akhir": tanggal_akhir
            }
        })
        
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "msg": str(e)
        }, status=500)


## FUNGSI LAPORAN & EKSPOR
# Kode ini untuk mengelola menu dan proses ekspor data absensi ke format Excel/CSV.
@login_required
def export_menu(request):
    """
    Menampilkan halaman menu ekspor laporan absensi.
    
     Filter cabang diterapkan
    """
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    cabang_aktif = get_active_cabang(request)
    
    #  Filter departemen berdasarkan pegawai yang ada di cabang
    if cabang_aktif:
        departemen_ids = Pegawai.objects.filter(
            cabang=cabang_aktif,
            is_active=True
        ).values_list('departemen_id', flat=True).distinct()
        
        departemen_list = MasterDepartemen.objects.filter(
            id__in=departemen_ids,
            is_active=True
        ).order_by('nama')
    else:
        departemen_list = MasterDepartemen.objects.filter(is_active=True).order_by('nama')
    
    pegawai_with_fp = get_pegawai_with_fingerprint()
    
    #  Filter pegawai berdasarkan cabang
    pegawai_list = Pegawai.objects.filter(
        is_active=True,
        uid_mesin__isnull=False,
        id__in=pegawai_with_fp
    ).exclude(uid_mesin=0)
    
    if cabang_aktif:
        pegawai_list = pegawai_list.filter(cabang=cabang_aktif)
    
    pegawai_list = pegawai_list.order_by('nama_lengkap')
    
    context = {
        'departemen_list': departemen_list,
        'pegawai_list': pegawai_list,
        'cabang_aktif': cabang_aktif,
    }
    return render(request, 'absensi_app/menu/export_menu.html', context)


def get_filtered_absensi(request):
    """Helper untuk mengambil data absensi yang telah difilter."""
    cabang_aktif = get_active_cabang(request)
    
    #AMBIL SEMUA ABSENSI (tidak filter pegawai dulu)
    absensi_list = Absensi.objects.select_related(
        'pegawai',
        'pegawai__departemen',
        'pegawai__jabatan',
        'pegawai__cabang'
    ).all()
    
    # FILTER BERDASARKAN CABANG AKTIF
    if cabang_aktif:
        absensi_list = absensi_list.filter(pegawai__cabang=cabang_aktif)
    
    # FILTER BERDASARKAN PEGAWAI TERTENTU
    pegawai_id = request.GET.get('pegawai_id')
    if pegawai_id:
        absensi_list = absensi_list.filter(pegawai_id=pegawai_id)
    
    # FILTER BERDASARKAN DEPARTEMEN
    departemen_nama = request.GET.get('departemen')
    if departemen_nama:
        absensi_list = absensi_list.filter(pegawai__departemen__nama=departemen_nama)
    
    # FILTER BERDASARKAN PERIODE
    periode = request.GET.get('periode', 'custom')
    today = timezone.now().date()
    
    if periode == 'hari_ini':
        absensi_list = absensi_list.filter(tanggal=today)
    elif periode == 'minggu_ini':
        start_week = today - timedelta(days=today.weekday())
        absensi_list = absensi_list.filter(tanggal__gte=start_week, tanggal__lte=today)
    elif periode == 'bulan_ini':
        start_month = today.replace(day=1)
        absensi_list = absensi_list.filter(tanggal__gte=start_month, tanggal__lte=today)
    elif periode == 'tahun_ini':
        start_year = today.replace(month=1, day=1)
        absensi_list = absensi_list.filter(tanggal__gte=start_year, tanggal__lte=today)
    elif periode == 'custom':
        tgl_mulai = request.GET.get('tanggal_mulai')
        tgl_akhir = request.GET.get('tanggal_akhir')
        
        if tgl_mulai:
            absensi_list = absensi_list.filter(tanggal__gte=tgl_mulai)
        if tgl_akhir:
            absensi_list = absensi_list.filter(tanggal__lte=tgl_akhir)
    
    return absensi_list.order_by('-tanggal', 'pegawai__userid')

def _build_export_filter_info(request):
    """Helper untuk membuat informasi filter untuk laporan ekspor."""
    filter_info = []
    
    pegawai_id = request.GET.get('pegawai_id')
    if pegawai_id:
        try:
            pegawai = Pegawai.objects.get(id=pegawai_id)
            filter_info.append(f"Pegawai: {pegawai.nama_lengkap} ({pegawai.userid})")
        except Pegawai.DoesNotExist:
            pass
    
    departemen_nama = request.GET.get('departemen')
    if departemen_nama:
        filter_info.append(f"Departemen: {departemen_nama}")
    
    periode = request.GET.get('periode', 'custom')
    if periode == 'custom':
        tgl_mulai = request.GET.get('tanggal_mulai', '-')
        tgl_akhir = request.GET.get('tanggal_akhir', '-')
        filter_info.append(f"Periode: {tgl_mulai} sampai {tgl_akhir}")
    else:
        periode_map = {
            'hari_ini': 'Hari Ini',
            'minggu_ini': 'Minggu Ini',
            'bulan_ini': 'Bulan Ini',
            'tahun_ini': 'Tahun Ini'
        }
        filter_info.append(f"Periode: {periode_map.get(periode, periode)}")
    
    cabang_aktif = get_active_cabang(request)
    if cabang_aktif:
        filter_info.append(f"Cabang: {cabang_aktif.nama}")
    
    return filter_info


def _build_export_filename(request, ext):
    """Helper untuk membuat nama file ekspor."""
    filename_parts = ['Absensi']
    
    pegawai_id = request.GET.get('pegawai_id')
    if pegawai_id:
        try:
            pegawai = Pegawai.objects.get(id=pegawai_id)
            filename_parts.append(pegawai.userid)
        except Pegawai.DoesNotExist:
            pass
    
    departemen = request.GET.get('departemen')
    if departemen:
        dept_clean = departemen.replace(' ', '_').replace('/', '-')
        filename_parts.append(dept_clean)
    
    periode = request.GET.get('periode', 'custom')
    if periode == 'custom':
        tgl_mulai = request.GET.get('tanggal_mulai')
        tgl_akhir = request.GET.get('tanggal_akhir')
        if tgl_mulai and tgl_akhir:
            filename_parts.append(f"{tgl_mulai}_to_{tgl_akhir}")
    else:
        filename_parts.append(periode)
    
    filename_parts.append(timezone.now().strftime("%Y%m%d_%H%M%S"))
    
    filename = '_'.join(filename_parts)
    
    return f"{filename}.{ext}"


@login_required
def export_absensi_csv_advanced(request):
    """Mengekspor laporan absensi ke file CSV."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    print("\n========== DEBUG EXPORT CSV ==========")
    print("GET Parameters:")
    for key, value in request.GET.items():
        print(f"  {key}: {value}")
    
    absensi_list = get_filtered_absensi(request)
    total_data = absensi_list.count()
    
    print(f"Total data ditemukan: {total_data}")
    print("======================================\n")
    
    if total_data == 0:
        messages.warning(request, 'Tidak ada data absensi untuk periode yang dipilih!')
        return redirect('export_menu')
    
    response = HttpResponse(content_type='text/csv')
    filename = _build_export_filename(request, 'csv')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    writer = csv.writer(response)
    
    writer.writerow([f"LAPORAN ABSENSI - Diekspor pada {timezone.now().strftime('%d %B %Y %H:%M:%S')}"])
    writer.writerow([])
    
    filter_info = _build_export_filter_info(request)
    for info in filter_info:
        writer.writerow([info])
    
    writer.writerow([f"Total Data: {total_data} record"])
    writer.writerow([])
    
    headers = [
        'No', 'User ID', 'Nama', 'Departemen', 'Jabatan',
        'Tanggal', 'Masuk', 'Istirahat Keluar', 'Istirahat Masuk', 'Keluar',
        'Total Jam Kerja', 'Status', 'Catatan'
    ]
    writer.writerow(headers)
    
    for idx, absensi in enumerate(absensi_list, 1):
        keterangan = absensi.keterangan or '-'
        if absensi.is_late:
            keterangan += " (Terlambat)"
        if absensi.is_early_departure:
            keterangan += " (Pulang Cepat)"
        
        jam_kerja = absensi.calculate_total_jam_kerja()
        
        writer.writerow([
            idx,
            absensi.pegawai.userid,
            absensi.pegawai.nama_lengkap,
            absensi.pegawai.departemen.nama if absensi.pegawai.departemen else '-',
            absensi.pegawai.jabatan.nama if absensi.pegawai.jabatan else '-',
            absensi.tanggal.strftime('%d-%m-%Y'),
            absensi.tap_masuk.strftime('%H:%M') if absensi.tap_masuk else '-',
            absensi.tap_istirahat_keluar.strftime('%H:%M') if absensi.tap_istirahat_keluar else '-',
            absensi.tap_istirahat_masuk.strftime('%H:%M') if absensi.tap_istirahat_masuk else '-',
            absensi.tap_pulang.strftime('%H:%M') if absensi.tap_pulang else '-',
            jam_kerja['formatted'] if jam_kerja else '-',
            absensi.status,
            keterangan
        ])
    
    return response

@login_required
def preview_export_data(request):
    """
    Endpoint AJAX: Menampilkan pratinjau data sebelum diekspor.
    
    ✅ OPTIMIZED: Menggunakan select_related dan values untuk performa
    """
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        # ✅ Query optimized dengan select_related
        absensi_list = Absensi.objects.select_related(
            'pegawai',
            'pegawai__departemen',
            'pegawai__jabatan'
        ).all()
        
        # ✅ Filter cabang
        cabang_aktif = get_active_cabang(request)
        if cabang_aktif:
            absensi_list = absensi_list.filter(pegawai__cabang=cabang_aktif)
        
        # ✅ Filter pegawai
        pegawai_id = request.GET.get('pegawai_id')
        if pegawai_id:
            absensi_list = absensi_list.filter(pegawai_id=pegawai_id)
        
        # ✅ Filter departemen
        departemen = request.GET.get('departemen')
        if departemen:
            absensi_list = absensi_list.filter(pegawai__departemen__nama=departemen)
        
        # ✅ Filter periode
        periode = request.GET.get('periode', 'custom')
        today = date.today()
        
        if periode == 'hari_ini':
            absensi_list = absensi_list.filter(tanggal=today)
        elif periode == 'minggu_ini':
            start_week = today - timedelta(days=today.weekday())
            absensi_list = absensi_list.filter(tanggal__gte=start_week, tanggal__lte=today)
        elif periode == 'bulan_ini':
            start_month = today.replace(day=1)
            absensi_list = absensi_list.filter(tanggal__gte=start_month, tanggal__lte=today)
        elif periode == 'tahun_ini':
            start_year = today.replace(month=1, day=1)
            absensi_list = absensi_list.filter(tanggal__gte=start_year, tanggal__lte=today)
        elif periode == 'custom':
            tgl_mulai = request.GET.get('tanggal_mulai')
            tgl_akhir = request.GET.get('tanggal_akhir')
            
            if tgl_mulai:
                absensi_list = absensi_list.filter(tanggal__gte=tgl_mulai)
            if tgl_akhir:
                absensi_list = absensi_list.filter(tanggal__lte=tgl_akhir)
        
        # ✅ Order by tanggal (terbaru dulu)
        absensi_list = absensi_list.order_by('-tanggal', 'pegawai__userid')
        
        # ✅ Hitung total SEBELUM limit
        total_data = absensi_list.count()
        
        print(f"\n========== PREVIEW EXPORT ==========")
        print(f"Total data: {total_data}")
        print(f"Filter periode: {periode}")
        print(f"Cabang: {cabang_aktif.nama if cabang_aktif else 'All'}")
        print(f"====================================\n")
        
        # ✅ Apply limit
        limit = request.GET.get('limit', '10')
        if limit != 'all':
            try:
                limit_num = int(limit)
                absensi_list = absensi_list[:limit_num]
            except ValueError:
                absensi_list = absensi_list[:10]
        
        # ✅ Build preview data (minimal processing)
        preview_data = []
        for absensi in absensi_list:
            # ✅ Hitung total jam kerja
            jam_kerja = absensi.calculate_total_jam_kerja()
            total_jam_kerja = jam_kerja.get('formatted', '-') if jam_kerja else '-'
            
            # ✅ Format shift
            shift_text = '-'
            if absensi.tap_masuk and absensi.tap_pulang:
                shift_text = f"{absensi.tap_masuk.strftime('%H:%M')} - {absensi.tap_pulang.strftime('%H:%M')}"
            elif absensi.tap_masuk:
                shift_text = f"{absensi.tap_masuk.strftime('%H:%M')} - ..."
            
            preview_data.append({
                'userid': absensi.pegawai.userid,
                'nama': absensi.pegawai.nama_lengkap,
                'departemen': absensi.pegawai.departemen.nama if absensi.pegawai.departemen else '-',
                'jabatan': absensi.pegawai.jabatan.nama if absensi.pegawai.jabatan else '-',
                'tanggal': absensi.tanggal.strftime('%d-%m-%Y'),
                'shift': shift_text,
                'tap_masuk': absensi.tap_masuk.strftime('%H:%M') if absensi.tap_masuk else '-',
                'tap_pulang': absensi.tap_pulang.strftime('%H:%M') if absensi.tap_pulang else '-',
                'total_jam_kerja': total_jam_kerja,
                'status': absensi.status,
                'is_late': absensi.is_late,
                'is_early_departure': absensi.is_early_departure
            })
        
        return JsonResponse({
            "status": "success",
            "total": total_data,
            "preview": preview_data,
            "limit": limit,
            "showing": len(preview_data)
        })
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"ERROR Preview: {error_detail}")
        
        return JsonResponse({
            "status": "error",
            "msg": str(e)
        }, status=500)
    
@login_required
def export_statistik_absensi(request):
    """Mengekspor laporan statistik absensi ke Excel."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    try:
        tgl_mulai = request.GET.get('tanggal_mulai')
        tgl_akhir = request.GET.get('tanggal_akhir')
        
        if not tgl_mulai or not tgl_akhir:
            messages.error(request, "Tanggal mulai dan akhir wajib diisi")
            return redirect('export_menu')
        
        start_date = datetime.strptime(tgl_mulai, '%Y-%m-%d').date()
        end_date = datetime.strptime(tgl_akhir, '%Y-%m-%d').date()

        cabang_aktif = get_active_cabang(request)
        
        pegawai_list = Pegawai.objects.filter(
            is_active=True
        ).select_related('departemen', 'jabatan', 'cabang')
        
        if cabang_aktif:
            pegawai_list = pegawai_list.filter(cabang=cabang_aktif)
        
        pegawai_list = pegawai_list.order_by('nama_lengkap')
        
        total_pegawai = pegawai_list.count()
        

        print(f"\n========== EXPORT STATISTIK ==========")
        print(f"Total Pegawai: {total_pegawai}")
        print(f"Periode: {start_date} s/d {end_date}")
        print(f"======================================\n")
        
        if total_pegawai == 0:
            messages.warning(request, 'Tidak ada data pegawai yang ditemukan!')
            return redirect('export_menu')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Statistik Absensi"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row_num = 1
        ws.merge_cells('A1:K1')
        ws['A1'].value = f"LAPORAN STATISTIK ABSENSI"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        row_num += 1
        
        ws.merge_cells(f'A{row_num}:K{row_num}')
        ws[f'A{row_num}'].value = f"Periode: {start_date.strftime('%d %B %Y')} - {end_date.strftime('%d %B %Y')}"
        ws[f'A{row_num}'].font = Font(italic=True)
        row_num += 1
        
        ws.merge_cells(f'A{row_num}:K{row_num}')
        ws[f'A{row_num}'].value = f"Total Pegawai: {total_pegawai}"
        ws[f'A{row_num}'].font = Font(italic=True, bold=True)
        row_num += 2
        
        headers = [
            'No', 'User ID', 'Nama', 'Departemen', 'Jabatan',
            'Hadir', 'Sakit', 'Izin', 'Absen', 'Terlambat', 'Pulang Cepat'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        row_num += 1
        
        for idx, pegawai in enumerate(pegawai_list, 1):
            absensi_range = Absensi.objects.filter(
                pegawai=pegawai,
                tanggal__gte=start_date,
                tanggal__lte=end_date
            )
            
            hadir = absensi_range.filter(status='Hadir').count()
            sakit = absensi_range.filter(status='Sakit').count()
            izin = absensi_range.filter(status='Izin').count()
            absen = absensi_range.filter(status='Absen').count()
            late = absensi_range.filter(is_late=True).count()
            early = absensi_range.filter(is_early_departure=True).count()
            
            ws.cell(row=row_num, column=1, value=idx).border = border
            ws.cell(row=row_num, column=2, value=pegawai.userid).border = border
            ws.cell(row=row_num, column=3, value=pegawai.nama_lengkap).border = border
            ws.cell(row=row_num, column=4, value=pegawai.departemen.nama if pegawai.departemen else '-').border = border
            ws.cell(row=row_num, column=5, value=pegawai.jabatan.nama if pegawai.jabatan else '-').border = border
            ws.cell(row=row_num, column=6, value=hadir).border = border
            ws.cell(row=row_num, column=7, value=sakit).border = border
            ws.cell(row=row_num, column=8, value=izin).border = border
            ws.cell(row=row_num, column=9, value=absen).border = border
            ws.cell(row=row_num, column=10, value=late).border = border
            ws.cell(row=row_num, column=11, value=early).border = border
            
            row_num += 1
        
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"Statistik_Absensi_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        return response
        
    except Exception as e:
        print(f"ERROR Export Statistik: {str(e)}")
        import traceback
        traceback.print_exc()
        messages.error(request, f"Ekspor gagal: {str(e)}")
        return redirect('export_menu')


@login_required
def preview_export_data(request):
    """Endpoint AJAX: Menampilkan pratinjau data sebelum diekspor."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        absensi_list = get_filtered_absensi(request)
        
        limit = request.GET.get('limit', '10')
        total_data = absensi_list.count()
        
        print(f"\n========== PREVIEW EXPORT ==========")
        print(f"Total data: {total_data}")
        print(f"Limit: {limit}")
        print(f"GET params: {dict(request.GET)}")
        print(f"====================================\n")
        
        if limit == 'all':
            preview_list = absensi_list
        else:
            try:
                limit_num = int(limit)
                preview_list = absensi_list[:limit_num]
            except ValueError:
                preview_list = absensi_list[:10]
        
        preview_data = []
        for absensi in preview_list:
            # TAMBAH FIELD SHIFT
            shift_text = '-'
            if absensi.tap_masuk and absensi.tap_pulang:
                shift_text = f"{absensi.tap_masuk.strftime('%H:%M')} - {absensi.tap_pulang.strftime('%H:%M')}"
            elif absensi.tap_masuk:
                shift_text = f"{absensi.tap_masuk.strftime('%H:%M')} - ..."
            
            preview_data.append({
                'userid': absensi.pegawai.userid,
                'nama': absensi.pegawai.nama_lengkap,
                'departemen': absensi.pegawai.departemen.nama if absensi.pegawai.departemen else '-',
                'jabatan': absensi.pegawai.jabatan.nama if absensi.pegawai.jabatan else '-',
                'tanggal': absensi.tanggal.strftime('%d-%m-%Y'),
                'shift': shift_text, 
                'tap_masuk': absensi.tap_masuk.strftime('%H:%M') if absensi.tap_masuk else '-',
                'tap_pulang': absensi.tap_pulang.strftime('%H:%M') if absensi.tap_pulang else '-',
                'status': absensi.status,
                'is_late': absensi.is_late,
                'is_early_departure': absensi.is_early_departure
            })
        
        return JsonResponse({
            "status": "success",
            "total": total_data,
            "preview": preview_data,
            "limit": limit,
            "showing": len(preview_data)
        })
        
    except Exception as e:
        print(f"ERROR Preview: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            "status": "error",
            "msg": str(e)
        }, status=500)

## FUNGSI MANAJEMEN DATA MASTER
# Kode ini untuk mengelola data master: Departemen, Jabatan, Cabang, dan Mesin.

@login_required
@user_passes_test(is_staff_or_superuser)
def menu_pengaturan(request):
    """Menampilkan halaman menu Pengaturan Sistem."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    context = {
        'total_departemen': MasterDepartemen.objects.filter(is_active=True).count(),
        'total_jabatan': MasterJabatan.objects.filter(is_active=True).count(),
        'total_cabang': MasterCabang.objects.filter(is_active=True).count(),
        'total_mesin': MasterMesin.objects.filter(is_active=True).count(),
        'total_admin': User.objects.filter(is_staff=True).count(),
        'total_mode_jam_kerja': MasterModeJamKerja.objects.filter(is_active=True).count(),
        'page_title': 'Pengaturan Sistem',
    }
    
    return render(request, 'absensi_app/menu/menu_pengaturan.html', context)


@login_required
def daftar_departemen(request):
    """
    Menampilkan daftar departemen.
    
     Tidak perlu filter cabang (departemen bersifat global)
    """
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    cabang_aktif = get_active_cabang(request)
    
    departemen_list = MasterDepartemen.objects.all().order_by('id_departemen', 'nama')
    
    departemen_stats = []
    for dept in departemen_list:
        #  Hitung pegawai per cabang
        total_pegawai = dept.pegawai_list.filter(is_active=True)
        
        if cabang_aktif:
            total_pegawai = total_pegawai.filter(cabang=cabang_aktif)
        
        departemen_stats.append({
            'departemen': dept,
            'total_pegawai': total_pegawai.count(),
        })
    
    context = {
        'departemen_list': departemen_list,
        'departemen_stats': departemen_stats,
        'cabang_aktif': cabang_aktif,
    }
    return render(request, 'absensi_app/pengaturan/departemen/daftar.html', context)

@login_required
def tambah_departemen(request):
    """Menambah departemen baru."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = MasterDepartemenForm(request.POST)
        if form.is_valid():
            departemen = form.save(commit=False)
            kode_dept = departemen.id_departemen
            
            if MasterDepartemen.objects.filter(id_departemen=kode_dept).exists():
                messages.error(
                    request,
                    f'ID Departemen "{kode_dept}" sudah digunakan. Gunakan ID yang berbeda.'
                )
                context = {'form': form, 'judul': 'Tambah Departemen'}
                return render(request, 'absensi_app/pengaturan/departemen/form.html', context)
            
            departemen.save()
            messages.success(
                request,
                f'Departemen "{departemen.nama}" berhasil ditambahkan dengan ID: {kode_dept}'
            )
            return redirect('daftar_departemen')
        else:
            _show_form_errors(request, form)
    else:
        form = MasterDepartemenForm()
    
    context = {'form': form, 'judul': 'Tambah Departemen'}
    return render(request, 'absensi_app/pengaturan/departemen/form.html', context)


@login_required
def edit_departemen(request, pk):
    """Mengedit data departemen."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    departemen = get_object_or_404(MasterDepartemen, pk=pk)
    
    if request.method == 'POST':
        form = MasterDepartemenForm(request.POST, instance=departemen)
        if form.is_valid():
            updated_dept = form.save(commit=False)
            kode_dept = updated_dept.id_departemen
            
            if MasterDepartemen.objects.filter(id_departemen=kode_dept).exclude(pk=pk).exists():
                messages.error(
                    request,
                    f'ID Departemen "{kode_dept}" sudah digunakan oleh departemen lain. Gunakan ID yang berbeda.'
                )
                context = {'form': form, 'judul': 'Edit Departemen', 'departemen': departemen}
                return render(request, 'absensi_app/pengaturan/departemen/form.html', context)
            
            updated_dept.save()
            messages.success(
                request,
                f'Departemen "{updated_dept.nama}" berhasil diperbarui dengan ID: {kode_dept}'
            )
            return redirect('daftar_departemen')
        else:
            _show_form_errors(request, form)
    else:
        form = MasterDepartemenForm(instance=departemen)
    
    context = {'form': form, 'judul': 'Edit Departemen', 'departemen': departemen}
    return render(request, 'absensi_app/pengaturan/departemen/form.html', context)


@login_required
def hapus_departemen(request, pk):
    """Menghapus departemen."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    departemen = get_object_or_404(MasterDepartemen, pk=pk)
    total_pegawai = departemen.pegawai_list.count()
    
    if request.method == 'POST':
        if total_pegawai > 0:
            messages.error(
                request,
                f'Tidak bisa dihapus. Departemen "{departemen.nama}" masih memiliki {total_pegawai} pegawai. Pindahkan atau hapus pegawai terlebih dahulu.'
            )
            return redirect('daftar_departemen')
        
        nama = departemen.nama
        kode_dept = departemen.id_departemen
        departemen.delete()
        messages.success(request, f'Departemen "{nama}" (Kode: {kode_dept}) berhasil dihapus')
        return redirect('daftar_departemen')
    
    context = {'departemen': departemen, 'total_pegawai': total_pegawai}
    return render(request, 'absensi_app/pengaturan/departemen/hapus.html', context)


@login_required
def daftar_jabatan(request):
    """
    Menampilkan daftar jabatan.
    
     Tidak perlu filter cabang (jabatan bersifat global)
    """
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    cabang_aktif = get_active_cabang(request)
    
    jabatan_list = MasterJabatan.objects.all().order_by('nama')
    
    #  Hitung pegawai per jabatan per cabang
    jabatan_stats = []
    for jabatan in jabatan_list:
        total_pegawai = jabatan.pegawai_list.filter(is_active=True)
        
        if cabang_aktif:
            total_pegawai = total_pegawai.filter(cabang=cabang_aktif)
        
        jabatan_stats.append({
            'jabatan': jabatan,
            'total_pegawai': total_pegawai.count(),
        })
    
    context = {
        'jabatan_list': jabatan_list,
        'jabatan_stats': jabatan_stats,
        'cabang_aktif': cabang_aktif,
    }
    return render(request, 'absensi_app/pengaturan/jabatan/daftar.html', context)

@login_required
def tambah_jabatan(request):
    """Menambah jabatan baru."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = MasterJabatanForm(request.POST)
        if form.is_valid():
            try:
                jabatan = form.save()
                messages.success(request, f'Jabatan "{jabatan.nama}" berhasil ditambahkan')
                return redirect('daftar_jabatan')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
        else:
            _show_form_errors(request, form)
    else:
        form = MasterJabatanForm()
    
    context = {'form': form, 'judul': 'Tambah Jabatan'}
    return render(request, 'absensi_app/pengaturan/jabatan/form.html', context)


@login_required
def edit_jabatan(request, pk):
    """Mengedit data jabatan."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    jabatan = get_object_or_404(MasterJabatan, pk=pk)
    
    if request.method == 'POST':
        form = MasterJabatanForm(request.POST, instance=jabatan)
        if form.is_valid():
            form.save()
            messages.success(request, 'Jabatan berhasil diperbarui')
            return redirect('daftar_jabatan')
        else:
            messages.error(request, 'Gagal memperbarui jabatan')
    else:
        form = MasterJabatanForm(instance=jabatan)
    
    context = {'form': form, 'judul': 'Edit Jabatan', 'jabatan': jabatan}
    return render(request, 'absensi_app/pengaturan/jabatan/form.html', context)


@login_required
def hapus_jabatan(request, pk):
    """Menghapus jabatan."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    jabatan = get_object_or_404(MasterJabatan, pk=pk)
    
    if request.method == 'POST':
        nama = jabatan.nama
        jabatan.delete()
        messages.success(request, f'Jabatan "{nama}" berhasil dihapus')
        return redirect('daftar_jabatan')
    
    context = {'jabatan': jabatan}
    return render(request, 'absensi_app/pengaturan/jabatan/hapus.html', context)


@login_required
def daftar_cabang(request):
    """Menampilkan daftar cabang."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    cabang_list = MasterCabang.objects.all().order_by('nama')
    
    context = {'cabang_list': cabang_list}
    return render(request, 'absensi_app/pengaturan/cabang/daftar.html', context)


@login_required
def tambah_cabang(request):
    """Menambah cabang baru."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = MasterCabangForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cabang berhasil ditambahkan')
            return redirect('daftar_cabang')
        else:
            messages.error(request, 'Gagal menambah cabang')
    else:
        form = MasterCabangForm()
    
    context = {'form': form, 'judul': 'Tambah Cabang'}
    return render(request, 'absensi_app/pengaturan/cabang/form.html', context)


@login_required
def edit_cabang(request, pk):
    """Mengedit data cabang."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    cabang = get_object_or_404(MasterCabang, pk=pk)
    
    if request.method == 'POST':
        form = MasterCabangForm(request.POST, instance=cabang)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cabang berhasil diperbarui')
            return redirect('daftar_cabang')
        else:
            messages.error(request, 'Gagal memperbarui cabang')
    else:
        form = MasterCabangForm(instance=cabang)
    
    context = {'form': form, 'judul': 'Edit Cabang', 'cabang': cabang}
    return render(request, 'absensi_app/pengaturan/cabang/form.html', context)


@login_required
def hapus_cabang(request, pk):
    """Menghapus cabang."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    cabang = get_object_or_404(MasterCabang, pk=pk)
    
    if request.method == 'POST':
        nama = cabang.nama
        cabang.delete()
        messages.success(request, f'Cabang "{nama}" berhasil dihapus')
        return redirect('daftar_cabang')
    
    context = {'cabang': cabang}
    return render(request, 'absensi_app/pengaturan/cabang/hapus.html', context)


@login_required
def daftar_mesin(request):
    """Menampilkan daftar mesin sidik jari dengan filter cabang."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    cabang_aktif = get_active_cabang(request)
    
    mesin_list = MasterMesin.objects.select_related('cabang').filter(is_active=True)
    
    if cabang_aktif:
        mesin_list = mesin_list.filter(cabang=cabang_aktif)
    
    mesin_list = mesin_list.order_by('nama')
    
    context = {
        'mesin_list': mesin_list,
        'cabang_aktif': cabang_aktif,
    }
    return render(request, 'absensi_app/pengaturan/mesin/daftar.html', context)


@login_required
def tambah_mesin(request):
    """Menambah mesin baru."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = MasterMesinForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Mesin berhasil ditambahkan')
            return redirect('daftar_mesin')
        else:
            messages.error(request, 'Gagal menambah mesin')
    else:
        form = MasterMesinForm()
    
    context = {'form': form, 'judul': 'Tambah Mesin'}
    return render(request, 'absensi_app/pengaturan/mesin/form.html', context)


@login_required
def edit_mesin(request, pk):
    """Mengedit data mesin."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    mesin = get_object_or_404(MasterMesin, pk=pk)
    
    if request.method == 'POST':
        form = MasterMesinForm(request.POST, instance=mesin)
        if form.is_valid():
            form.save()
            messages.success(request, 'Mesin berhasil diperbarui')
            return redirect('daftar_mesin')
        else:
            messages.error(request, 'Gagal memperbarui mesin')
    else:
        form = MasterMesinForm(instance=mesin)
    
    context = {'form': form, 'judul': 'Edit Mesin', 'mesin': mesin}
    return render(request, 'absensi_app/pengaturan/mesin/form.html', context)


@login_required
def hapus_mesin(request, pk):
    """Menghapus mesin."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    mesin = get_object_or_404(MasterMesin, pk=pk)
    
    if request.method == 'POST':
        nama = mesin.nama
        mesin.delete()
        messages.success(request, f'Mesin "{nama}" berhasil dihapus')
        return redirect('daftar_mesin')
    
    context = {'mesin': mesin}
    return render(request, 'absensi_app/pengaturan/mesin/hapus.html', context)


@login_required
def test_mesin(request, pk):
    """Endpoint AJAX: Menguji koneksi ke mesin sidik jari."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    mesin = get_object_or_404(MasterMesin, pk=pk)
    
    try:
        conn = connect_to_fingerprint_machine(mesin.ip_address, mesin.port, timeout=5)
        users = conn.get_users()
        conn.disconnect()
        
        return JsonResponse({
            "status": "success",
            "online": True,
            "msg": f"Koneksi berhasil! Ditemukan {len(users)} user.",
            "total_users": len(users)
        })
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "online": False,
            "msg": f"Koneksi gagal: {str(e)}"
        })


## FUNGSI MANAJEMEN ADMIN
# Kode ini untuk mengelola user admin (staf dan superuser).

@login_required
def daftar_admin(request):
    """Menampilkan daftar user admin."""
    if not request.user.is_superuser:
        messages.error(request, "Akses ditolak. Hanya untuk super admin.")
        return redirect('dashboard')
    
    admin_list = User.objects.filter(is_staff=True).order_by('-is_superuser', 'username')
    
    context = {
        'admin_list': admin_list,
        'total_admin': admin_list.count(),
        'total_superuser': admin_list.filter(is_superuser=True).count(),
        'total_staff': admin_list.filter(is_superuser=False).count(),
    }
    return render(request, 'absensi_app/pengaturan/admin/daftar.html', context)


@login_required
def tambah_admin(request):
    """Menambah user admin baru."""
    if not request.user.is_superuser:
        messages.error(request, "Akses ditolak. Hanya untuk super admin.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        password_confirm = request.POST.get('password_confirm', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        is_superuser = request.POST.get('is_superuser') == 'on'
        
        if not username:
            messages.error(request, 'Username wajib diisi')
            return redirect('tambah_admin')
        
        if not password or password != password_confirm or len(password) < 8:
            messages.error(
                request,
                'Password wajib diisi, minimal 8 karakter, dan harus sama'
            )
            return redirect('tambah_admin')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, f'Username "{username}" sudah ada')
            return redirect('tambah_admin')
        
        if email and User.objects.filter(email=email).exists():
            messages.error(request, f'Email "{email}" sudah terdaftar')
            return redirect('tambah_admin')
        
        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                is_staff=True,
                is_superuser=is_superuser,
                is_active=True
            )
            
            role_text = "Super Admin" if is_superuser else "Admin/HR"
            messages.success(
                request,
                f'Admin "{username}" berhasil dibuat! Role: {role_text}'
            )
            return redirect('daftar_admin')
            
        except Exception as e:
            messages.error(request, f'Gagal membuat admin: {str(e)}')
            return redirect('tambah_admin')
    
    return render(request, 'absensi_app/pengaturan/admin/form.html', {'judul': 'Tambah Admin Baru'})


@login_required
def edit_admin(request, pk):
    """Mengedit data user admin."""
    if not request.user.is_superuser:
        messages.error(request, "Akses ditolak. Hanya untuk super admin.")
        return redirect('dashboard')
    
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        is_superuser = request.POST.get('is_superuser') == 'on'
        is_active = request.POST.get('is_active') == 'on'
        
        if email and User.objects.filter(email=email).exclude(pk=pk).exists():
            messages.error(request, f'Email "{email}" sudah terdaftar')
            return redirect('edit_admin', pk=pk)
        
        try:
            user.email = email
            user.first_name = first_name
            user.last_name = last_name
            user.is_superuser = is_superuser
            user.is_active = is_active
            user.save()
            
            messages.success(request, f'Admin "{user.username}" berhasil diperbarui')
            return redirect('daftar_admin')
            
        except Exception as e:
            messages.error(request, f'Gagal memperbarui admin: {str(e)}')
            return redirect('edit_admin', pk=pk)
    
    context = {'user': user, 'judul': f'Edit Admin: {user.username}'}
    return render(request, 'absensi_app/pengaturan/admin/form_edit.html', context)


@login_required
def reset_password_admin(request, pk):
    """Endpoint AJAX: Mengatur ulang password user admin."""
    if not request.user.is_superuser:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({"status": "error", "msg": "Method tidak diizinkan"}, status=405)
    
    user = get_object_or_404(User, pk=pk)
    
    try:
        new_password = request.POST.get('new_password', '').strip()
        
        if not new_password or len(new_password) < 8:
            return JsonResponse(
                {"status": "error", "msg": "Password minimal 8 karakter"},
                status=400
            )
        
        user.set_password(new_password)
        user.save()
        
        return JsonResponse({
            "status": "success",
            "msg": f"Password untuk admin '{user.username}' berhasil diatur ulang",
            "new_password": new_password
        })
    
    except Exception as e:
        return JsonResponse(
            {"status": "error", "msg": f"Gagal mengatur ulang password: {str(e)}"},
            status=500
        )


@login_required
def hapus_admin(request, pk):
    """Menghapus user admin."""
    if not request.user.is_superuser:
        messages.error(request, "Akses ditolak. Hanya untuk super admin.")
        return redirect('dashboard')
    
    user = get_object_or_404(User, pk=pk)
    
    if user.pk == request.user.pk:
        messages.error(request, "Tidak dapat menghapus akun Anda sendiri")
        return redirect('daftar_admin')
    
    if request.method == 'POST':
        username = user.username
        
        try:
            user.delete()
            messages.success(request, f'Admin "{username}" berhasil dihapus')
        except Exception as e:
            messages.error(request, f'Gagal menghapus admin: {str(e)}')
        
        return redirect('daftar_admin')
    
    context = {'user': user}
    return render(request, 'absensi_app/pengaturan/admin/hapus.html', context)


## FUNGSI MANAJEMEN RIWAYAT ABSENSI
# Kode ini untuk melihat, mengedit, dan menghapus data absensi.

@login_required
def riwayat_absensi_per_pegawai(request, pk):
    """Menampilkan riwayat absensi detail per pegawai."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    pegawai = get_object_or_404(Pegawai, pk=pk)
    form = LaporanFilterForm(request.GET)
    
    today = date.today()
    jam_kerja_info = WorkModeService.get_jam_kerja_for_pegawai(pegawai, today)
    jadwal_hari_ini = jam_kerja_info.get('jadwal')
    mode_info = jam_kerja_info.get('mode')
    
    absensi_list = Absensi.objects.filter(
        pegawai=pegawai
    ).select_related(
        'pegawai',
        'pegawai__departemen', 
        'pegawai__jabatan'
    ).order_by('-tanggal', '-tap_masuk')
    
    if form.is_valid():
        tgl_mulai = form.cleaned_data.get('tanggal_mulai')
        tgl_akhir = form.cleaned_data.get('tanggal_akhir')
        
        if tgl_mulai:
            absensi_list = absensi_list.filter(tanggal__gte=tgl_mulai)
        if tgl_akhir:
            absensi_list = absensi_list.filter(tanggal__lte=tgl_akhir)
    
    period = request.GET.get('period', '')
    today_date = timezone.now().date()
    
    if period == 'week':
        start_date = today_date - timedelta(days=today_date.weekday())
        absensi_list = absensi_list.filter(tanggal__gte=start_date)
    elif period == 'month':
        start_date = today_date.replace(day=1)
        absensi_list = absensi_list.filter(tanggal__gte=start_date)
    elif period == 'year':
        start_date = today_date.replace(month=1, day=1)
        absensi_list = absensi_list.filter(tanggal__gte=start_date)
    
    total_hadir = absensi_list.filter(status='Hadir').count()
    total_sakit = absensi_list.filter(status='Sakit').count()
    total_izin = absensi_list.filter(status='Izin').count()
    total_absen = absensi_list.filter(status='Absen').count()
    total_terlambat = absensi_list.filter(is_late=True).count()
    total_pulang_cepat = absensi_list.filter(is_early_departure=True).count()
    
    # ✅ HITUNG TOTAL JAM KERJA
    total_jam_kerja_minutes = 0
    absensi_with_stats = []
    
    for absensi in absensi_list:
        jam_kerja_data = absensi.calculate_total_jam_kerja()
        
        if jam_kerja_data and jam_kerja_data.get('total_minutes'):
            total_jam_kerja_minutes += jam_kerja_data['total_minutes']
        
        # Tambahkan attribute untuk template
        absensi.total_jam_kerja = jam_kerja_data.get('formatted', '-') if jam_kerja_data else '-'
        
        # Cek violation istirahat
        absensi.has_violation = False
        if absensi.tap_masuk and absensi.tap_pulang:
            if not absensi.tap_istirahat_keluar or not absensi.tap_istirahat_masuk:
                absensi.has_violation = True
        
        absensi_with_stats.append(absensi)
    
    total_jam_kerja_hours = total_jam_kerja_minutes // 60
    total_jam_kerja_mins = total_jam_kerja_minutes % 60
    total_jam_kerja_formatted = f"{total_jam_kerja_hours}j {total_jam_kerja_mins}m"
    
    total_records = absensi_list.count()
    if total_records > 0:
        persentase_hadir = round((total_hadir / total_records) * 100, 2)
    else:
        persentase_hadir = 0
    
    context = {
        'pegawai': pegawai,
        'absensi_list': absensi_with_stats,
        'form': form,
        'period': period,
        'total_hadir': total_hadir,
        'total_sakit': total_sakit,
        'total_izin': total_izin,
        'total_absen': total_absen,
        'total_terlambat': total_terlambat,
        'total_pulang_cepat': total_pulang_cepat,
        'total_records': total_records,
        'persentase_hadir': persentase_hadir,
        'total_jam_kerja_minutes': total_jam_kerja_minutes,
        'total_jam_kerja_hours': total_jam_kerja_hours,
        'total_jam_kerja_mins': total_jam_kerja_mins,
        'total_jam_kerja_formatted': total_jam_kerja_formatted,
        'jadwal_hari_ini': jadwal_hari_ini,
        'mode_info': {
            'mode_nama': mode_info.nama if mode_info else None,
            'mode_kode': mode_info.kode if mode_info else None,
        } if mode_info else None,
        'today': today,
    }
    return render(request, 'absensi_app/absensi/riwayat_per_Pegawai.html', context)


@login_required
def absensi_edit(request, pk):
    """Mengedit data absensi secara manual."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    absensi = get_object_or_404(Absensi, pk=pk)
    
    if request.method == 'POST':
        try:
            absensi.status = request.POST.get('status', 'Hadir')
            absensi.keterangan = request.POST.get('keterangan', '')
            
            tap_masuk_str = request.POST.get('tap_masuk')
            if tap_masuk_str:
                absensi.tap_masuk = datetime.strptime(tap_masuk_str, '%H:%M').time()
            else:
                absensi.tap_masuk = None
            
            tap_pulang_str = request.POST.get('tap_pulang')
            if tap_pulang_str:
                absensi.tap_pulang = datetime.strptime(tap_pulang_str, '%H:%M').time()
            else:
                absensi.tap_pulang = None
            
            tap_istirahat_keluar_str = request.POST.get('tap_istirahat_keluar')
            if tap_istirahat_keluar_str:
                absensi.tap_istirahat_keluar = datetime.strptime(tap_istirahat_keluar_str, '%H:%M').time()
            else:
                absensi.tap_istirahat_keluar = None
            
            tap_istirahat_masuk_str = request.POST.get('tap_istirahat_masuk')
            if tap_istirahat_masuk_str:
                absensi.tap_istirahat_masuk = datetime.strptime(tap_istirahat_masuk_str, '%H:%M').time()
            else:
                absensi.tap_istirahat_masuk = None
            
            if absensi.tap_pulang and absensi.tap_masuk:
                if absensi.tap_pulang <= absensi.tap_masuk:
                    messages.error(request, 'Waktu keluar harus lebih besar dari waktu masuk')
                    return redirect('absensi_edit', pk=pk)
            
            if absensi.status == 'Hadir' and absensi.tap_masuk:
                validation = _validate_attendance(
                    absensi.tap_masuk,
                    absensi.tap_pulang,
                    absensi.tanggal,
                    absensi.pegawai
                )
                absensi.is_late = validation['is_late']
                absensi.is_early_departure = validation['is_early']
                
                if validation['keterangan']:
                    if absensi.keterangan:
                        absensi.keterangan += f" | {validation['keterangan']}"
                    else:
                        absensi.keterangan = validation['keterangan']
            
            absensi.updated_at = timezone.now()
            absensi.save()
            
            messages.success(request, f"Data absensi {absensi.pegawai.nama_lengkap} berhasil diupdate")
            return redirect('riwayat_absensi')
            
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return redirect('absensi_edit', pk=pk)
    
    context = {
        'absensi': absensi,
        'judul': 'Edit Data Absensi'
    }
    return render(request, 'absensi_app/absensi/absensi_edit_form.html', context)


@login_required
def absensi_hapus(request, pk):
    """Menghapus data absensi."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    absensi = get_object_or_404(Absensi, pk=pk)
    
    if request.method == 'POST':
        pegawai_nama = absensi.pegawai.nama_lengkap
        tanggal_absensi = absensi.tanggal.strftime('%d %B %Y')
        
        try:
            absensi.delete()
            messages.success(
                request,
                f'Absensi untuk {pegawai_nama} pada {tanggal_absensi} berhasil dihapus'
            )
        except Exception as e:
            messages.error(request, f'Gagal menghapus: {str(e)}')
        
        return redirect('riwayat_absensi')
    
    context = {'absensi': absensi}
    return render(request, 'absensi_app/absensi/absensi_hapus.html', context)


## FUNGSI TRANSFER PEGAWAI ANTAR MESIN
# Kode ini untuk mengelola pemindahan data user dan sidik jari antar mesin absensi.

@login_required
def transfer_pegawai_ke_mesin(request):
    """
    Menampilkan halaman untuk transfer pegawai antar mesin.
    
     Filter cabang diterapkan
    """
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    cabang_aktif = get_active_cabang(request)
    
    #  Filter pegawai berdasarkan cabang
    pegawai_list = Pegawai.objects.filter(
        is_active=True,
        uid_mesin__isnull=False
    ).exclude(uid_mesin=0).select_related(
        'departemen', 'jabatan', 'cabang', 'mesin'
    ).order_by('nama_lengkap')
    
    if cabang_aktif:
        pegawai_list = pegawai_list.filter(cabang=cabang_aktif)
    
    cabang_list = MasterCabang.objects.filter(is_active=True).order_by('nama')
    
    #  Filter mesin berdasarkan cabang aktif
    mesin_list = MasterMesin.objects.select_related('cabang').filter(
        is_active=True
    )
    
    if cabang_aktif:
        mesin_list = mesin_list.filter(cabang=cabang_aktif)
    
    mesin_list = mesin_list.order_by('nama')
    
    context = {
        'pegawai_list': pegawai_list,
        'cabang_list': cabang_list,
        'mesin_list': mesin_list,
        'total_pegawai': pegawai_list.count(),
        'cabang_aktif': cabang_aktif,
        'page_title': 'Transfer Pegawai Antar Mesin'
    }
    
    return render(request, 'absensi_app/register/transfer_mesin.html', context)


@login_required
def proses_transfer_pegawai(request):
    """Endpoint AJAX: Memproses transfer user dan sidik jari antar mesin."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({"status": "error", "msg": "Method tidak diizinkan"}, status=405)
    
    try:
        pegawai_id = request.POST.get('pegawai_id')
        ip_mesin_asal = request.POST.get('ip_mesin_asal')
        ip_mesin_tujuan = request.POST.get('ip_mesin_tujuan')
        hapus_dari_asal = request.POST.get('hapus_dari_asal') == 'true'
        
        if not all([pegawai_id, ip_mesin_asal, ip_mesin_tujuan]):
            return JsonResponse({
                "status": "error",
                "msg": "Data tidak lengkap"
            }, status=400)
        
        if ip_mesin_asal == ip_mesin_tujuan:
            return JsonResponse({
                "status": "error",
                "msg": "Mesin asal dan tujuan tidak boleh sama"
            }, status=400)
        
        pegawai_obj = get_object_or_404(Pegawai, id=pegawai_id)
        
        # STEP 1: AMBIL DATA DARI MESIN ASAL
        try:
            conn_asal = connect_to_fingerprint_machine(ip_mesin_asal, timeout=15)
            conn_asal.disable_device()
            
            users_asal = conn_asal.get_users()
            
            target_user = next(
                (u for u in users_asal if str(getattr(u, 'user_id', u.uid)) == str(pegawai_obj.userid)),
                None
            )
            
            if not target_user:
                conn_asal.enable_device()
                conn_asal.disconnect()
                return JsonResponse({
                    "status": "error",
                    "msg": f"User {pegawai_obj.userid} tidak ditemukan di mesin asal"
                }, status=404)
            
            fingerprint_templates = conn_asal.get_templates()
            
            user_templates = [
                t for t in fingerprint_templates
                if t.uid == target_user.uid
            ]
            
            conn_asal.enable_device()
            conn_asal.disconnect()
            
        except Exception as e:
            return JsonResponse({
                "status": "error",
                "msg": f"Gagal mengambil data dari mesin asal: {str(e)}"
            }, status=500)
        
        # STEP 2: REGISTER KE MESIN TUJUAN
        try:
            conn_tujuan = connect_to_fingerprint_machine(ip_mesin_tujuan, timeout=15)
            conn_tujuan.disable_device()
            
            users_tujuan = conn_tujuan.get_users()
            existing_user = next(
                (u for u in users_tujuan if str(getattr(u, 'user_id', u.uid)) == str(pegawai_obj.userid)),
                None
            )
            
            if existing_user:
                uid_tujuan = existing_user.uid
                
                conn_tujuan.set_user(
                    uid=uid_tujuan,
                    name=target_user.name,
                    privilege=target_user.privilege,
                    password=target_user.password or '',
                    group_id=target_user.group_id or '',
                    user_id=target_user.user_id,
                    card=target_user.card
                )
            else:
                uid_tujuan = register_Pegawai_to_machine(
                    conn_tujuan,
                    pegawai_obj.userid,
                    pegawai_obj.nama_lengkap
                )
                
                users_updated = conn_tujuan.get_users()
                new_user = next(
                    (u for u in users_updated if u.uid == uid_tujuan),
                    None
                )
                
                if not new_user:
                    raise Exception("Gagal membuat user di mesin tujuan")
                
                target_user = new_user
            
        except Exception as e:
            try:
                conn_tujuan.enable_device()
                conn_tujuan.disconnect()
            except:
                pass
            
            return JsonResponse({
                "status": "error",
                "msg": f"Gagal register user ke mesin tujuan: {str(e)}"
            }, status=500)
        
        # STEP 3: SAVE FINGERPRINT TEMPLATE
        if user_templates:
            try:
                from zk import user as usr
                
                user_obj = usr.User(
                    uid=uid_tujuan,
                    name=target_user.name,
                    privilege=target_user.privilege,
                    password=target_user.password or '',
                    group_id=target_user.group_id or '',
                    user_id=target_user.user_id,
                    card=target_user.card
                )
                
                from zk import finger
                updated_templates = []
                for t in user_templates:
                    updated_templates.append(
                        finger.Finger(
                            uid=uid_tujuan,
                            fid=t.fid,
                            valid=t.valid,
                            template=t.template
                        )
                    )
                
                conn_tujuan.save_user_template(user_obj, updated_templates)
                
            except Exception as e:
                pass
        
        conn_tujuan.enable_device()
        conn_tujuan.disconnect()
        
        # STEP 4: DELETE FROM SOURCE (OPSIONAL)
        deleted_from_source = False
        if hapus_dari_asal:
            try:
                conn_asal = connect_to_fingerprint_machine(ip_mesin_asal, timeout=10)
                conn_asal.disable_device()
                
                conn_asal.delete_user(uid=target_user.uid)
                deleted_from_source = True
                
                conn_asal.enable_device()
                conn_asal.disconnect()
            except Exception as e:
                pass
        
        # STEP 5: UPDATE DATABASE
        try:
            with transaction.atomic():
                pegawai_obj.uid_mesin = uid_tujuan
                
                mesin_tujuan = MasterMesin.objects.filter(
                    ip_address=ip_mesin_tujuan
                ).first()
                
                if mesin_tujuan:
                    pegawai_obj.mesin = mesin_tujuan
                    if mesin_tujuan.cabang:
                        pegawai_obj.cabang = mesin_tujuan.cabang
                
                pegawai_obj.save()
                
                pegawai_obj.fingerprint_templates.all().delete()
                
                for template in user_templates:
                    FingerprintTemplate.objects.create(
                        pegawai=pegawai_obj,
                        uid=uid_tujuan,
                        fid=template.fid,
                        size=template.size,
                        valid=template.valid,
                        template=template.template
                    )
                
        except Exception as e:
            return JsonResponse({
                "status": "error",
                "msg": f"Transfer berhasil ke mesin, tapi gagal update database: {str(e)}"
            }, status=500)
        
        # SUCCESS MESSAGE
        mesin_tujuan_obj = MasterMesin.objects.filter(ip_address=ip_mesin_tujuan).first()
        
        msg = f"Transfer Berhasil!\n\n"
        msg += f"Detail:\n"
        msg += f"• Pegawai: {pegawai_obj.nama_lengkap}\n"
        msg += f"• User ID: {pegawai_obj.userid}\n"
        msg += f"• UID Baru: {uid_tujuan}\n"
        msg += f"• Mesin Tujuan: {mesin_tujuan_obj.nama if mesin_tujuan_obj else ip_mesin_tujuan}\n"
        msg += f"• Fingerprint: {len(user_templates)} templates\n"
        
        if deleted_from_source:
            msg += f"• Data dihapus dari mesin asal\n"
        
        msg += f"\nPegawai sudah bisa langsung TAP di mesin tujuan!"
        
        return JsonResponse({
            "status": "success",
            "msg": msg,
            "data": {
                "pegawai_id": pegawai_obj.id,
                "uid_baru": uid_tujuan,
                "fingerprint_count": len(user_templates),
                "hapus_asal": deleted_from_source
            }
        })
        
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "msg": f"Transfer gagal: {str(e)}"
        }, status=500)


@login_required
def bulk_transfer_pegawai(request):
    """Endpoint AJAX: Transfer massal pegawai (hanya User ID) ke mesin lain."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({"status": "error", "msg": "Method tidak diizinkan"}, status=405)
    
    try:
        pegawai_ids = request.POST.getlist('pegawai_ids[]')
        ip_mesin_tujuan = request.POST.get('ip_mesin_tujuan')
        
        if not pegawai_ids or not ip_mesin_tujuan:
            return JsonResponse({
                "status": "error",
                "msg": "Data tidak lengkap"
            }, status=400)
        
        conn_tujuan = connect_to_fingerprint_machine(ip_mesin_tujuan)
        
        results = {
            'success': [],
            'failed': [],
            'total': len(pegawai_ids)
        }
        
        for pegawai_id in pegawai_ids:
            pegawai_obj = None
            try:
                pegawai_obj = Pegawai.objects.get(id=pegawai_id)
                
                users_tujuan = conn_tujuan.get_users()
                existing = next(
                    (u for u in users_tujuan if str(getattr(u, 'user_id', u.uid)) == str(pegawai_obj.userid)),
                    None
                )
                
                if existing:
                    results['failed'].append({
                        'userid': pegawai_obj.userid,
                        'nama': pegawai_obj.nama_lengkap,
                        'reason': 'Sudah terdaftar di mesin tujuan'
                    })
                    continue
                
                uid_tujuan = register_Pegawai_to_machine(
                    conn_tujuan,
                    pegawai_obj.userid,
                    pegawai_obj.nama_lengkap
                )
                
                pegawai_obj.uid_mesin = uid_tujuan
                pegawai_obj.save()
                
                results['success'].append({
                    'userid': pegawai_obj.userid,
                    'nama': pegawai_obj.nama_lengkap,
                    'uid_baru': uid_tujuan
                })
                
            except Pegawai.DoesNotExist:
                results['failed'].append({
                    'userid': pegawai_id,
                    'nama': 'Tidak Diketahui',
                    'reason': 'Pegawai tidak ditemukan'
                })
            except Exception as e:
                pegawai_nama = pegawai_obj.nama_lengkap if pegawai_obj else 'Tidak Diketahui'
                pegawai_id_str = pegawai_obj.userid if pegawai_obj else pegawai_id
                results['failed'].append({
                    'userid': pegawai_id_str,
                    'nama': pegawai_nama,
                    'reason': str(e)
                })
        
        conn_tujuan.disconnect()
        
        msg = f"Transfer massal selesai\n\n"
        msg += f"Berhasil: {len(results['success'])} pegawai\n"
        msg += f"Gagal: {len(results['failed'])} pegawai"
        
        if results['failed'] and len(results['failed']) <= 5:
            msg += "\n\nDetail Kegagalan:\n"
            for item in results['failed']:
                msg += f"- {item['nama']} ({item['userid']}): {item['reason']}\n"
        
        return JsonResponse({
            "status": "success",
            "msg": msg,
            "results": results
        })
        
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "msg": f"Transfer massal gagal: {str(e)}"
        }, status=500)


@login_required
def get_pegawai_by_cabang(request):
    """Endpoint AJAX: Mengambil daftar pegawai berdasarkan cabang untuk menu transfer."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        cabang_id = request.GET.get('cabang_id')
        
        if not cabang_id:
            return JsonResponse({
                "status": "error",
                "msg": "Cabang ID tidak ditemukan"
            }, status=400)
        
        cabang = get_object_or_404(MasterCabang, id=cabang_id)
        
        pegawai_list = Pegawai.objects.filter(
            is_active=True,
            uid_mesin__isnull=False,
            cabang=cabang
        ).exclude(uid_mesin=0).select_related(
            'departemen', 'mesin'
        ).order_by('nama_lengkap')
        
        data = []
        for p in pegawai_list:
            data.append({
                'id': p.id,
                'userid': p.userid,
                'nama': p.nama_lengkap,
                'departemen': p.departemen.nama if p.departemen else '-',
                'uid_mesin': p.uid_mesin,
                'mesin': p.mesin.nama if p.mesin else '-',
                'fingerprint_count': p.fingerprint_templates.count()
            })
        
        return JsonResponse({
            "status": "success",
            "cabang": {
                "id": cabang.id,
                "nama": cabang.nama,
                "kode": cabang.kode
            },
            "pegawai": data,
            "total": len(data)
        })
        
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "msg": str(e)
        }, status=500)


@login_required
def cek_pegawai_di_mesin(request):
    """Endpoint AJAX: Memeriksa apakah pegawai sudah ada di mesin tertentu."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        userid = request.GET.get('userid')
        ip_mesin = request.GET.get('ip_mesin')
        
        if not all([userid, ip_mesin]):
            return JsonResponse({
                "status": "error",
                "msg": "Data tidak lengkap"
            }, status=400)
        
        conn = connect_to_fingerprint_machine(ip_mesin, timeout=5)
        users = conn.get_users()
        
        existing_user = next(
            (u for u in users if str(getattr(u, 'user_id', u.uid)) == str(userid)),
            None
        )
        
        conn.disconnect()
        
        if existing_user:
            return JsonResponse({
                "status": "success",
                "exists": True,
                "data": {
                    "uid": existing_user.uid,
                    "userid": getattr(existing_user, 'user_id', existing_user.uid),
                    "nama": existing_user.name,
                    "privilege": existing_user.privilege
                }
            })
        else:
            return JsonResponse({
                "status": "success",
                "exists": False,
                "msg": f"Pegawai dengan User ID {userid} tidak ditemukan di mesin"
            })
            
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "msg": str(e)
        }, status=500)


@login_required
def cek_status_mesin_bulk(request):
    """Endpoint AJAX: Memeriksa status koneksi mesin untuk operasi massal."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        ip_address = request.GET.get('ip_address')
        
        if not ip_address:
            return JsonResponse({
                "status": "error",
                "msg": "IP address wajib diisi"
            }, status=400)
        
        conn = connect_to_fingerprint_machine(ip_address, timeout=5)
        users = conn.get_users()
        
        pegawai_belum_sync = Pegawai.objects.filter(
            is_active=True
        ).filter(
            Q(uid_mesin__isnull=True) | Q(uid_mesin=0)
        ).count()
        
        conn.disconnect()
        
        return JsonResponse({
            "status": "success",
            "mesin_online": True,
            "total_user_di_mesin": len(users),
            "pegawai_belum_sync": pegawai_belum_sync,
            "msg": f"Mesin online. Total {len(users)} user terdaftar."
        })
        
    except Exception as e:
        return JsonResponse({
            "status": "warning",
            "mesin_online": False,
            "msg": f"Mesin tidak dapat dijangkau: {str(e)}"
        })


@login_required
def sync_semua_uid_dari_mesin(request):
    """Menyinkronkan semua UID dari mesin ke database (untuk pegawai yang UID-nya NULL)."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    try:
        mesin_id = request.POST.get('mesin_id')
        
        if not mesin_id:
            messages.error(request, "Pilih mesin terlebih dahulu")
            return redirect('daftar_Pegawai')
        
        mesin = get_object_or_404(MasterMesin, id=mesin_id, is_active=True)
        
        conn = connect_to_fingerprint_machine(mesin.ip_address, mesin.port, timeout=10)
        mesin_users = conn.get_users()
        conn.disconnect()
        
        pegawai_list = Pegawai.objects.filter(
            Q(uid_mesin__isnull=True) | Q(uid_mesin=0)
        )
        
        synced_count = 0
        not_found_list = []
        
        for pegawai_obj in pegawai_list:
            mesin_user = next(
                (u for u in mesin_users if str(getattr(u, 'user_id', u.uid)) == str(pegawai_obj.userid)),
                None
            )
            
            if mesin_user:
                pegawai_obj.uid_mesin = mesin_user.uid
                pegawai_obj.mesin = mesin
                pegawai_obj.save()
                synced_count += 1
            else:
                not_found_list.append(f"{pegawai_obj.userid} - {pegawai_obj.nama_lengkap}")
        
        msg = f"Sinkronisasi UID berhasil dari {mesin.nama}! Total tersinkronisasi: {synced_count} pegawai\n"
        
        if not_found_list:
            msg += f"\nTidak ditemukan di mesin ({len(not_found_list)}):\n"
            for item in not_found_list[:5]:
                msg += f" - {item}\n"
            if len(not_found_list) > 5:
                msg += f" ... dan {len(not_found_list) - 5} lainnya\n"
        
        messages.success(request, msg)
        
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect('daftar_Pegawai')


## FUNGSI MANAJEMEN MODE JAM KERJA
# Kode ini untuk mengelola data master Mode Jam Kerja, Periode Khusus, dan Pegawai yang Dikecualikan.

@login_required
def daftar_mode_jam_kerja(request):
    """
    Menampilkan daftar semua mode jam kerja yang aktif.
    
     Hitung pegawai per cabang + Fix status aktif
    """
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    cabang_aktif = get_active_cabang(request)
    mode_today = WorkModeService.get_mode_today()
    
    all_modes = MasterModeJamKerja.objects.filter(
        is_active=True
    ).prefetch_related(
        'periode_list',
        'pegawai_list',
    )
    
    #  Annotate pegawai count berdasarkan cabang
    if cabang_aktif:
        all_modes = all_modes.annotate(
            pegawai_count=Count(
                'pegawai_list', 
                filter=Q(pegawai_list__is_active=True, pegawai_list__cabang=cabang_aktif)
            ),
            jadwal_count=Count('jadwal_list'),
            periode_count=Count('periode_list')
        )
    else:
        all_modes = all_modes.annotate(
            pegawai_count=Count('pegawai_list', filter=Q(pegawai_list__is_active=True)),
            jadwal_count=Count('jadwal_list'),
            periode_count=Count('periode_list')
        )
    
    all_modes = sorted(
        all_modes,
        key=lambda m: (not m.is_default, -m.priority, m.nama)
    )
    
    today = date.today()
    mode_data = []
    
    for mode in all_modes:
        # Cari periode aktif untuk hari ini
        periode_aktif = next(
            (p for p in mode.periode_list.all()
             if p.is_active and p.tanggal_mulai <= today <= p.tanggal_selesai),
            None
        )
        
        #  Mode aktif jika adalah DEFAULT atau ada PERIODE AKTIF
        is_applicable = mode.is_default or periode_aktif is not None
        
        mode_data.append({
            'mode': mode,
            'periode_aktif': periode_aktif,
            'is_applicable': is_applicable,  # ← TAMBAHAN PENTING
            'pegawai_count': mode.pegawai_count,
        })
    
    context = {
        'mode_today': mode_today,
        'mode_data': mode_data,
        'total_mode': len(all_modes),
        'cabang_aktif': cabang_aktif,
    }
    return render(request, 'absensi_app/pengaturan/mode_jam_kerja/daftar.html', context)



@login_required
def tambah_mode_jam_kerja(request):
    """Menambah mode jam kerja baru dengan auto-generate kode."""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                nama = request.POST.get('nama')
                
                # ✅ AUTO-GENERATE KODE dari nama
                kode = ''.join([word[0].upper() for word in nama.split()[:3]])
                
                # ✅ Pastikan kode unik
                base_kode = kode
                counter = 1
                while MasterModeJamKerja.objects.filter(kode=kode).exists():
                    kode = f"{base_kode}{counter}"
                    counter += 1
                
                mode = MasterModeJamKerja.objects.create(
                    nama=nama,
                    kode=kode,  # Auto-generated
                    warna=request.POST.get('warna', '#3B82F6'),
                    priority=int(request.POST.get('priority', 1)),
                )
                
                schedules_json = request.POST.get('schedules_json', '{}')
                
                if not schedules_json or schedules_json == '{}':
                    mode.delete()
                    messages.error(request, '❌ Jadwal belum diatur!')
                    return redirect('tambah_mode_jam_kerja')
                
                try:
                    schedules = json.loads(schedules_json)
                except json.JSONDecodeError as e:
                    mode.delete()
                    messages.error(request, f'❌ Format JSON tidak valid: {str(e)}')
                    return redirect('tambah_mode_jam_kerja')
                
                jadwal_count = 0
                for group_id_str, group_data in schedules.items():
                    if not isinstance(group_data, dict):
                        continue
                    
                    group_name = group_data.get('name', f'Group {group_id_str}')
                    days_data = group_data.get('days', {})
                    
                    if not isinstance(days_data, dict):
                        continue
                    
                    for hari_str, shifts_list in days_data.items():
                        try:
                            hari = int(hari_str)
                        except (ValueError, TypeError):
                            continue
                        
                        if not isinstance(shifts_list, list):
                            shifts_list = [shifts_list]
                        
                        for shift_idx, shift_data in enumerate(shifts_list):
                            if not isinstance(shift_data, dict):
                                continue
                            
                            jam_masuk = shift_data.get('masuk')
                            jam_keluar = shift_data.get('keluar')
                            
                            if not jam_masuk or not jam_keluar:
                                continue
                            
                            ModeJamKerjaJadwal.objects.create(
                                mode=mode,
                                group_name=group_name,
                                hari=hari,
                                jam_masuk=jam_masuk,
                                jam_keluar=jam_keluar,
                                jam_istirahat_keluar=shift_data.get('break_out') or None,
                                jam_istirahat_masuk=shift_data.get('break_in') or None,
                                toleransi_terlambat=5,
                                toleransi_pulang_cepat=5,
                                urutan=shift_idx + 1,
                            )
                            jadwal_count += 1
                
                if jadwal_count == 0:
                    mode.delete()
                    messages.error(request, '❌ Tidak ada jadwal yang valid!')
                    return redirect('tambah_mode_jam_kerja')
                
                WorkModeService.clear_cache()
                messages.success(request, f'✅ Mode "{mode.nama}" berhasil ditambahkan dengan {jadwal_count} jadwal!')
                return redirect('detail_mode_jam_kerja', pk=mode.id)
            
        except Exception as e:
            messages.error(request, f'❌ Error: {str(e)}')
            return redirect('tambah_mode_jam_kerja')
    
    context = {
        'judul': 'Tambah Mode Jam Kerja',
        'priority_choices': MasterModeJamKerja.PRIORITY_CHOICES,
        'schedules_json': '{}',
    }
    return render(request, 'absensi_app/pengaturan/mode_jam_kerja/form.html', context)


@login_required
def edit_mode_jam_kerja(request, pk):
    """
    ✅ FINAL: Edit mode jam kerja dengan auto-update kode + CLEAR CACHE
    """
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    mode = get_object_or_404(MasterModeJamKerja, pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                nama_baru = request.POST.get('nama')
                
                # ✅ Jika nama berubah, generate kode baru
                if nama_baru != mode.nama:
                    kode = ''.join([word[0].upper() for word in nama_baru.split()[:3]])
                    
                    # ✅ Pastikan kode unik (exclude mode saat ini)
                    base_kode = kode
                    counter = 1
                    while MasterModeJamKerja.objects.filter(kode=kode).exclude(pk=pk).exists():
                        kode = f"{base_kode}{counter}"
                        counter += 1
                    
                    mode.kode = kode
                
                mode.nama = nama_baru
                mode.warna = request.POST.get('warna', '#3B82F6')
                mode.icon = request.POST.get('icon', 'fas fa-clock')
                mode.priority = int(request.POST.get('priority', 1))
                mode.save()
                
                # ✅ Hapus jadwal lama
                mode.jadwal_list.all().delete()
                
                # ✅ Parse jadwal baru dari JSON
                schedules_json = request.POST.get('schedules_json', '{}')
                
                if not schedules_json or schedules_json == '{}':
                    messages.error(request, '❌ Jadwal belum diatur!')
                    return redirect('edit_mode_jam_kerja', pk=pk)
                
                try:
                    schedules = json.loads(schedules_json)
                except json.JSONDecodeError as e:
                    messages.error(request, f'❌ Format JSON tidak valid: {str(e)}')
                    return redirect('edit_mode_jam_kerja', pk=pk)
                
                shift_count = 0
                
                for group_id_str, group_data in schedules.items():
                    if not isinstance(group_data, dict):
                        continue
                    
                    group_name = group_data.get('name', f'Group {group_id_str}')
                    days_data = group_data.get('days', {})
                    
                    if not isinstance(days_data, dict):
                        continue
                    
                    for hari_str, shifts_list in days_data.items():
                        try:
                            hari = int(hari_str)
                        except (ValueError, TypeError):
                            continue
                        
                        # ✅ Validasi hari (0-6)
                        if not (0 <= hari <= 6):
                            continue
                        
                        # ✅ Support list of shifts
                        if not isinstance(shifts_list, list):
                            shifts_list = [shifts_list]
                        
                        for shift_idx, shift_data in enumerate(shifts_list):
                            if not isinstance(shift_data, dict):
                                continue
                            
                            jam_masuk = shift_data.get('masuk')
                            jam_keluar = shift_data.get('keluar')
                            
                            if not jam_masuk or not jam_keluar:
                                continue
                            
                            # ✅ Create jadwal baru
                            ModeJamKerjaJadwal.objects.create(
                                mode=mode,
                                group_name=group_name,
                                hari=hari,
                                jam_masuk=jam_masuk,
                                jam_keluar=jam_keluar,
                                jam_istirahat_keluar=shift_data.get('break_out') or None,
                                jam_istirahat_masuk=shift_data.get('break_in') or None,
                                toleransi_terlambat=int(shift_data.get('toleransi', 5)),
                                toleransi_pulang_cepat=int(shift_data.get('toleransi', 5)),
                                urutan=shift_idx + 1,
                            )
                            shift_count += 1
                
                if shift_count == 0:
                    messages.error(request, '❌ Tidak ada jadwal yang valid!')
                    return redirect('edit_mode_jam_kerja', pk=pk)
                
                # ✅ CLEAR ALL CACHE (PENTING!)
                from django.core.cache import cache
                
                print("\n" + "="*60)
                print("🧹 CLEARING ALL CACHE AFTER EDIT MODE")
                print("="*60)
                
                # 1. Clear Django cache
                cache.clear()
                
                # 2. Clear WorkModeService cache
                WorkModeService.clear_cache()
                
                # 3. Clear pegawai detail cache (individual)
                pegawai_list = Pegawai.objects.filter(
                    mode_assignments__mode=mode,
                    mode_assignments__is_active=True
                ).distinct()
                
                for pegawai in pegawai_list:
                    cache_key = f"pegawai_mode_detail_{pegawai.id}"
                    cache.delete(cache_key)
                    print(f"✅ Cleared cache for pegawai: {pegawai.nama_lengkap}")
                
                print("="*60 + "\n")
                
                messages.success(
                    request, 
                    f'✅ Mode "{mode.nama}" berhasil diperbarui dengan {shift_count} jadwal!\n\n'
                    f'💡 Cache telah di-clear. Refresh halaman detail pegawai untuk melihat perubahan.'
                )
                return redirect('detail_mode_jam_kerja', pk=pk)
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"\n❌ ERROR edit_mode_jam_kerja:\n{error_detail}")
            
            messages.error(request, f'❌ Error: {str(e)}')
            return redirect('edit_mode_jam_kerja', pk=pk)
    
    # ========================================
    #  GET: Build schedules_json untuk edit form
    # ========================================
    
    schedules_by_group = {}
    
    for jadwal in mode.jadwal_list.all().order_by('group_name', 'hari', 'urutan'):
        group_name = jadwal.group_name
        hari = jadwal.hari
        
        if group_name not in schedules_by_group:
            schedules_by_group[group_name] = {
                'name': group_name,
                'days': {}
            }
        
        if hari not in schedules_by_group[group_name]['days']:
            schedules_by_group[group_name]['days'][hari] = []
        
        schedules_by_group[group_name]['days'][hari].append({
            'shift': group_name,
            'masuk': jadwal.jam_masuk.strftime('%H:%M') if jadwal.jam_masuk else '',
            'keluar': jadwal.jam_keluar.strftime('%H:%M') if jadwal.jam_keluar else '',
            'break_out': jadwal.jam_istirahat_keluar.strftime('%H:%M') if jadwal.jam_istirahat_keluar else '',
            'break_in': jadwal.jam_istirahat_masuk.strftime('%H:%M') if jadwal.jam_istirahat_masuk else '',
            'toleransi': jadwal.toleransi_terlambat,
        })
    
    context = {
        'mode': mode,
        'judul': f'Edit Mode: {mode.nama}',
        'priority_choices': MasterModeJamKerja.PRIORITY_CHOICES,
        'schedules_json': json.dumps(schedules_by_group),
    }
    return render(request, 'absensi_app/pengaturan/mode_jam_kerja/form.html', context)

@login_required
def detail_mode_jam_kerja(request, pk):
    """Menampilkan detail mode jam kerja."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')

    mode = get_object_or_404(MasterModeJamKerja, pk=pk)
    all_jadwal = mode.jadwal_list.all().order_by('group_name', 'hari', 'urutan')

    from collections import defaultdict
    jadwal_by_group = defaultdict(lambda: {'group_name': '', 'days': {}})
    hari_names = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']

    # Inisialisasi group
    for jadwal in all_jadwal:
        group_name = jadwal.group_name
        if not jadwal_by_group[group_name]['group_name']:
            jadwal_by_group[group_name]['group_name'] = group_name
            for hari_idx in range(7):
                jadwal_by_group[group_name]['days'][hari_idx] = {
                    'hari_idx': hari_idx,
                    'hari_nama': hari_names[hari_idx],
                    'jadwal_list': []
                }

    # Isi jadwal per hari
    for jadwal in all_jadwal:
        jadwal_by_group[jadwal.group_name]['days'][jadwal.hari]['jadwal_list'].append(jadwal)

    # Hitung hanya hari yang benar-benar ada jadwalnya
    for group_name, group_data in jadwal_by_group.items():
        group_data['jumlah_hari_aktif'] = sum(
            1 for d in group_data['days'].values()
            if d['jadwal_list']
        )

    jadwal_by_group = dict(jadwal_by_group)

    today = date.today()
    periode_aktif = mode.periode_list.filter(
        is_active=True,
        tanggal_mulai__lte=today,
        tanggal_selesai__gte=today
    ).first()

    periode_mendatang = mode.periode_list.filter(
        is_active=True,
        tanggal_mulai__gt=today
    ).order_by('tanggal_mulai')[:5]

    pegawai_count = mode.pegawai_list.filter(is_active=True).count()

    context = {
        'mode': mode,
        'jadwal_by_dept': jadwal_by_group,
        'hari_names': hari_names,
        'hari_names_enum': enumerate(hari_names),
        'periode_aktif': periode_aktif,
        'periode_mendatang': periode_mendatang,
        'pegawai_count': pegawai_count,
        'today': today,
    }
    return render(request, 'absensi_app/pengaturan/mode_jam_kerja/detail.html', context)


@login_required
def hapus_mode_jam_kerja(request, pk):
    """Menghapus mode jam kerja."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    mode = get_object_or_404(MasterModeJamKerja, pk=pk)
    
    if mode.is_default:
        messages.error(request, 'Tidak bisa menghapus mode default!')
        return redirect('daftar_mode_jam_kerja')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                nama_mode = mode.nama
                mode.jadwal_list.all().delete()
                mode.periode_list.all().delete()
                mode.delete()
                
                WorkModeService.clear_cache()
                messages.success(request, f'Mode "{nama_mode}" berhasil dihapus!')
                return redirect('daftar_mode_jam_kerja')
            
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return redirect('detail_mode_jam_kerja', pk=pk)
    
    context = {'mode': mode}
    return render(request, 'absensi_app/pengaturan/mode_jam_kerja/hapus.html', context)

@login_required
def hapus_periode_mode(request, pk):
    """Menghapus periode khusus"""
    if not request.user.is_staff:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    periode = get_object_or_404(ModeJamKerjaPeriode, pk=pk)
    mode_id = periode.mode.id
    
    if request.method == 'POST':
        try:
            nama_periode = periode.nama
            tanggal_mulai = periode.tanggal_mulai
            tanggal_selesai = periode.tanggal_selesai
            
            periode.delete()
            WorkModeService.clear_cache(tanggal_mulai, tanggal_selesai)
            
            # Check if AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    "status": "success",
                    "msg": f'✅ Periode "{nama_periode}" berhasil dihapus!'
                })
            
            messages.success(request, f'✅ Periode "{nama_periode}" berhasil dihapus!')
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    "status": "error",
                    "msg": f'❌ Error: {str(e)}'
                }, status=500)
            messages.error(request, f'❌ Error: {str(e)}')
        
        return redirect('detail_mode_jam_kerja', pk=mode_id)
    
    # GET request - tampilkan halaman konfirmasi
    context = {'periode': periode}
    return render(request, 'absensi_app/pengaturan/mode_jam_kerja/periode_hapus.html', context)

@login_required
def duplicate_mode_jam_kerja(request, pk):
    """Menduplikasi mode jam kerja yang sudah ada."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    source_mode = get_object_or_404(MasterModeJamKerja, pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                schedules_json = request.POST.get('schedules_json', '{}')
                
                if not schedules_json or schedules_json == '{}':
                    messages.error(request, 'Jadwal belum diatur!')
                    return redirect('duplicate_mode_jam_kerja', pk=pk)
                
                try:
                    schedules = json.loads(schedules_json)
                except json.JSONDecodeError as e:
                    messages.error(request, f'Format JSON tidak valid: {str(e)}')
                    return redirect('duplicate_mode_jam_kerja', pk=pk)
                
                new_mode = MasterModeJamKerja.objects.create(
                    nama=request.POST.get('nama'),
                    kode=request.POST.get('kode', '').upper(),
                    warna=request.POST.get('warna', source_mode.warna),
                    icon=request.POST.get('icon', source_mode.icon),
                    priority=int(request.POST.get('priority', source_mode.priority)),
                    is_default=False,
                )
                
                jadwal_count = 0
                for group_id_str, group_data in schedules.items():
                    group_name = group_data.get('name', f'Group {group_id_str}')
                    
                    for hari_str, shifts_list in group_data.get('days', {}).items():
                        hari = int(hari_str)
                        
                        for shift_idx, shift_data in enumerate(shifts_list):
                            jam_masuk = shift_data.get('masuk')
                            jam_keluar = shift_data.get('keluar')
                            
                            if not jam_masuk or not jam_keluar:
                                continue
                            
                            ModeJamKerjaJadwal.objects.create(
                                mode=new_mode,
                                group_name=group_name,
                                hari=hari,
                                jam_masuk=jam_masuk,
                                jam_keluar=jam_keluar,
                                jam_istirahat_keluar=shift_data.get('break_out') or None,
                                jam_istirahat_masuk=shift_data.get('break_in') or None,
                                toleransi_terlambat=int(shift_data.get('toleransi', 5)),
                                toleransi_pulang_cepat=int(shift_data.get('toleransi', 5)),
                                urutan=shift_idx + 1,
                            )
                            jadwal_count += 1
                
                if jadwal_count == 0:
                    new_mode.delete()
                    messages.error(request, 'Tidak ada jadwal yang valid!')
                    return redirect('duplicate_mode_jam_kerja', pk=pk)
                
                WorkModeService.clear_cache()
                messages.success(
                    request,
                    f'Mode "{new_mode.nama}" berhasil dibuat dari "{source_mode.nama}" dengan {jadwal_count} jadwal!'
                )
                return redirect('detail_mode_jam_kerja', pk=new_mode.id)
            
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return redirect('duplicate_mode_jam_kerja', pk=pk)
    
    schedules_by_group = {}
    
    for jadwal in source_mode.jadwal_list.all():
        group_name = jadwal.group_name
        hari = jadwal.hari
        
        if group_name not in schedules_by_group:
            schedules_by_group[group_name] = {
                'name': group_name,
                'days': {}
            }
        
        if hari not in schedules_by_group[group_name]['days']:
            schedules_by_group[group_name]['days'][hari] = []
        
        schedules_by_group[group_name]['days'][hari].append({
            'shift': group_name,
            'masuk': jadwal.jam_masuk.strftime('%H:%M') if jadwal.jam_masuk else '',
            'keluar': jadwal.jam_keluar.strftime('%H:%M') if jadwal.jam_keluar else '',
            'break_out': jadwal.jam_istirahat_keluar.strftime('%H:%M') if jadwal.jam_istirahat_keluar else '',
            'break_in': jadwal.jam_istirahat_masuk.strftime('%H:%M') if jadwal.jam_istirahat_masuk else '',
            'toleransi': jadwal.toleransi_terlambat,
        })
    
    context = {
        'source_mode': source_mode,
        'mode': None,
        'judul': f'Duplikasi Mode: {source_mode.nama}',
        'priority_choices': MasterModeJamKerja.PRIORITY_CHOICES,
        'schedules_json': json.dumps(schedules_by_group),
    }
    return render(request, 'absensi_app/pengaturan/mode_jam_kerja/form.html', context)


@login_required
def daftar_periode_mode(request, mode_id):
    """Menampilkan daftar periode khusus untuk suatu mode jam kerja."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    mode = get_object_or_404(MasterModeJamKerja, pk=mode_id)
    periode_list = mode.periode_list.order_by('-tanggal_mulai')
    
    context = {
        'mode': mode,
        'periode_list': periode_list,
        'today': date.today(),
    }
    return render(request, 'absensi_app/pengaturan/mode_jam_kerja/periode_form.html', context)


@login_required
def tambah_periode_mode(request, mode_id):
    """Menambah periode khusus baru untuk mode jam kerja."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    from .models import MasterModeJamKerja, ModeJamKerjaPeriode
    from datetime import datetime
    
    mode = get_object_or_404(MasterModeJamKerja, pk=mode_id)
    
    form_data = {
        'nama': '',
        'tanggal_mulai': '',
        'tanggal_selesai': '',
        'catatan': '',
    }
    
    if request.method == 'POST':
        form_data = {
            'nama': request.POST.get('nama', '').strip(),
            'tanggal_mulai': request.POST.get('tanggal_mulai', '').strip(),
            'tanggal_selesai': request.POST.get('tanggal_selesai', '').strip(),
            'catatan': request.POST.get('catatan', '').strip(),
        }
        
        if not form_data['nama']:
            messages.error(request, 'Nama periode wajib diisi!')
        elif not form_data['tanggal_mulai'] or not form_data['tanggal_selesai']:
            messages.error(request, 'Tanggal mulai dan selesai wajib diisi!')
        else:
            try:
                tanggal_mulai = datetime.strptime(form_data['tanggal_mulai'], '%Y-%m-%d').date()
                tanggal_selesai = datetime.strptime(form_data['tanggal_selesai'], '%Y-%m-%d').date()
                
                if tanggal_selesai < tanggal_mulai:
                    messages.error(request, 'Tanggal selesai harus >= tanggal mulai!')
                else:
                    periode = ModeJamKerjaPeriode.objects.create(
                        mode=mode,
                        nama=form_data['nama'],
                        tanggal_mulai=tanggal_mulai,
                        tanggal_selesai=tanggal_selesai,
                        tahun=tanggal_mulai.year,
                        catatan=form_data['catatan'],
                        is_active=True,
                    )
                    
                    from .services import WorkModeService
                    WorkModeService.clear_cache(tanggal_mulai, tanggal_selesai)
                    
                    messages.success(
                        request,
                        f'Periode "{periode.nama}" berhasil ditambahkan!'
                    )
                    return redirect('detail_mode_jam_kerja', pk=mode_id)
                    
            except ValueError:
                messages.error(request, 'Format tanggal tidak valid! Gunakan format YYYY-MM-DD')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
    
    context = {
        'mode': mode,
        'periode': None,
        'form_data': form_data,
        'judul': f'Tambah Periode - {mode.nama}',
    }
    return render(request, 'absensi_app/pengaturan/mode_jam_kerja/periode_form.html', context)


@login_required
def edit_periode_mode(request, pk):
    """Mengedit periode khusus."""
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    periode = get_object_or_404(ModeJamKerjaPeriode, pk=pk)
    
    if request.method == 'POST':
        try:
            periode.nama = request.POST.get('nama')
            periode.tanggal_mulai = datetime.strptime(request.POST.get('tanggal_mulai'), '%Y-%m-%d').date()
            periode.tanggal_selesai = datetime.strptime(request.POST.get('tanggal_selesai'), '%Y-%m-%d').date()
            periode.tahun = periode.tanggal_mulai.year
            periode.catatan = request.POST.get('catatan', '')
            periode.save()
            
            WorkModeService.clear_cache()
            messages.success(request, f'Periode berhasil diperbarui!')
            return redirect('detail_mode_jam_kerja', pk=periode.mode.id)
            
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'periode': periode,
        'mode': periode.mode,
        'judul': f'Edit Periode: {periode.nama}',
    }
    return render(request, 'absensi_app/pengaturan/mode_jam_kerja/periode_form.html', context)


@login_required
def hapus_periode_mode(request, pk):
    """Menghapus periode khusus"""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    periode = get_object_or_404(ModeJamKerjaPeriode, pk=pk)
    mode_id = periode.mode.id
    
    if request.method == 'POST':
        try:
            nama_periode = periode.nama
            tanggal_mulai = periode.tanggal_mulai
            tanggal_selesai = periode.tanggal_selesai
            
            periode.delete()
            WorkModeService.clear_cache(tanggal_mulai, tanggal_selesai)
            
            # Check if AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    "status": "success",
                    "msg": f'Periode "{nama_periode}" berhasil dihapus!'
                })
            
            messages.success(request, f'Periode "{nama_periode}" berhasil dihapus!')
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    "status": "error",
                    "msg": str(e)
                }, status=500)
            messages.error(request, f'Error: {str(e)}')
        
        return redirect('detail_mode_jam_kerja', pk=mode_id)
    
    context = {'periode': periode}
    return render(request, 'absensi_app/pengaturan/mode_jam_kerja/periode_hapus.html', context)

## FUNGSI API TAMBAHAN
# Kode ini berisi endpoint API tambahan untuk dashboard dan fitur-fitur lainnya.

@login_required
def get_pegawai_status_summary(request):
    """Endpoint AJAX: Mendapatkan ringkasan status pegawai untuk dashboard."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        pegawai_with_fp = get_pegawai_with_fingerprint()
        
        active = Pegawai.objects.filter(
            is_active=True,
            uid_mesin__isnull=False,
            id__in=pegawai_with_fp
        ).exclude(uid_mesin=0).count()
        
        pending = Pegawai.objects.filter(
            is_active=True,
            uid_mesin__isnull=False
        ).exclude(uid_mesin=0).exclude(id__in=pegawai_with_fp).count()
        
        inactive = Pegawai.objects.filter(is_active=False).count()
        
        not_registered = Pegawai.objects.filter(
            is_active=True
        ).filter(
            Q(uid_mesin__isnull=True) | Q(uid_mesin=0)
        ).count()
        
        return JsonResponse({
            "status": "success",
            "data": {
                "active": active,
                "pending": pending,
                "inactive": inactive,
                "not_registered": not_registered,
                "total": active + pending + inactive + not_registered
            }
        })
        
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "msg": str(e)
        }, status=500)


@login_required
def api_get_mode_today(request):
    """Endpoint API untuk mengambil mode jam kerja hari ini dan yang akan datang."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        from .services import WorkModeService
        from datetime import date
        
        # ✅ 1. Get mode info hari ini
        mode_info = WorkModeService.get_mode_today()
        
        print("\n========== DEBUG API GET MODE TODAY ==========")
        print(f"Mode Info: {mode_info}")
        
        # ✅ 2. Get upcoming modes (14 hari ke depan)
        upcoming = WorkModeService.get_upcoming_modes(14)
        
        print(f"Upcoming count: {len(upcoming)}")
        print("==============================================\n")
        
        # ✅ 3. Build response
        upcoming_data = []
        for p in upcoming:
            upcoming_data.append({
                "nama": p.nama,
                "mode": p.mode.nama,
                "warna": p.mode.warna,
                "icon": p.mode.icon,
                "tanggal_mulai": p.tanggal_mulai.strftime('%d %b %Y'),
                "tanggal_selesai": p.tanggal_selesai.strftime('%d %b %Y'),
            })
        
        return JsonResponse({
            "status": "success",
            "mode": {
                "nama": mode_info.get('nama_mode', 'Normal'),
                "kode": mode_info.get('mode_kode', 'N/A'),
                "warna": mode_info.get('mode_warna', '#3b82f6'),
                "icon": mode_info.get('mode_icon', 'fa-clock'),
                "is_libur": mode_info.get('is_libur', False),
                "periode": mode_info.get('nama_periode'),
            },
            "upcoming": upcoming_data
        })
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"❌ ERROR api_get_mode_today:\n{error_detail}")
        
        # ✅ Fallback response
        return JsonResponse({
            "status": "success",
            "mode": {
                "nama": "Normal",
                "kode": "N/A",
                "warna": "#3b82f6",
                "icon": "fa-clock",
                "is_libur": False,
                "periode": None,
            },
            "upcoming": []
        })
    
@login_required
def api_list_cabang(request):
    """Endpoint API: Daftar semua cabang yang aktif."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        cabang_aktif_id = request.session.get('cabang_aktif_id')
        cabang_list = MasterCabang.objects.filter(is_active=True).order_by('nama')
        
        data_cabang = []
        for cabang in cabang_list:
            total_pegawai = Pegawai.objects.filter(
                cabang=cabang,
                is_active=True
            ).count()
            
            total_mesin = MasterMesin.objects.filter(
                cabang=cabang,
                is_active=True
            ).count()
            
            data_cabang.append({
                'id': cabang.id,
                'nama': cabang.nama,
                'kode': cabang.kode,
                'alamat': cabang.alamat,
                'is_current': cabang.id == cabang_aktif_id,
                'total_pegawai': total_pegawai,
                'total_mesin': total_mesin
            })
        
        return JsonResponse({
            "status": "success",
            "cabang_list": data_cabang,
            "total": len(data_cabang)
        })
        
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "msg": str(e)
        }, status=500)

@login_required
def get_cabang_aktif(request):
    """Endpoint API: Mengambil informasi cabang yang sedang aktif di sesi user."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        cabang_aktif = get_active_cabang(request)
        
        if not cabang_aktif:
            return JsonResponse({
                "status": "warning",
                "msg": "Tidak ada cabang aktif",
                "cabang": None
            })
        
        total_pegawai = Pegawai.objects.filter(
            cabang=cabang_aktif,
            is_active=True
        ).count()
        
        total_mesin = MasterMesin.objects.filter(
            cabang=cabang_aktif,
            is_active=True
        ).count()
        
        return JsonResponse({
            "status": "success",
            "cabang": {
                "id": cabang_aktif.id,
                "nama": cabang_aktif.nama,
                "kode": cabang_aktif.kode,
                "alamat": cabang_aktif.alamat,
                "total_pegawai": total_pegawai,
                "total_mesin": total_mesin,
            }
        })
        
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "msg": f"Error: {str(e)}"
        }, status=500)
    
@login_required
def api_get_mode_jadwal(request, pk):
    """Endpoint API untuk mengambil jadwal lengkap mode (7 hari)."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        mode = get_object_or_404(MasterModeJamKerja, pk=pk)
        
        jadwal_list = []
        hari_names = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
        
        for hari in range(7):
            jadwal = mode.get_jadwal_hari(hari)
            
            if jadwal:
                jadwal_list.append({
                    'hari': hari,
                    'hari_nama': hari_names[hari],
                    'is_hari_kerja': True,
                    'jam_masuk': jadwal.jam_masuk.strftime('%H:%M') if jadwal.jam_masuk else None,
                    'jam_keluar': jadwal.jam_keluar.strftime('%H:%M') if jadwal.jam_keluar else None,
                    'jam_istirahat_keluar': jadwal.jam_istirahat_keluar.strftime('%H:%M') if jadwal.jam_istirahat_keluar else None,
                    'jam_istirahat_masuk': jadwal.jam_istirahat_masuk.strftime('%H:%M') if jadwal.jam_istirahat_masuk else None,
                    'toleransi_terlambat': jadwal.toleransi_terlambat,
                    'toleransi_pulang_cepat': jadwal.toleransi_pulang_cepat,
                })
            else:
                jadwal_list.append({
                    'hari': hari,
                    'hari_nama': hari_names[hari],
                    'is_hari_kerja': False,
                    'jam_masuk': None, 'jam_keluar': None, 'jam_istirahat_keluar': None,
                    'jam_istirahat_masuk': None, 'toleransi_terlambat': 0, 'toleransi_pulang_cepat': 0,
                })
        
        return JsonResponse({
            "status": "success",
            "mode": {
                "id": mode.id, "nama": mode.nama, "kode": mode.kode, "warna": mode.warna,
                "icon": mode.icon, "is_libur": mode.is_libur, "is_default": mode.is_default,
            },
            "jadwal": jadwal_list
        })
        
    except MasterModeJamKerja.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "msg": "Mode tidak ditemukan"
        }, status=404)
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "msg": str(e)
        }, status=500)


@login_required
def api_get_mesin_by_pegawai(request):
    """
    API: Mendapatkan daftar mesin yang memiliki data pegawai tertentu (untuk transfer).
    
    ✅ FIXED: Proper JSON response + error handling
    """
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        pegawai_id = request.GET.get('pegawai_id')
        
        print(f"\n🔍 DEBUG: pegawai_id = {pegawai_id}")  # ← Debug log
        
        if not pegawai_id:
            return JsonResponse({
                "status": "error",
                "msg": "Pegawai ID wajib diisi"
            }, status=400)
        
        pegawai = get_object_or_404(Pegawai, id=pegawai_id)
        
        print(f"✅ Pegawai: {pegawai.nama_lengkap} (userid: {pegawai.userid})")  # Debug
        
        mesin_list = MasterMesin.objects.filter(is_active=True).select_related('cabang')
        
        print(f"📡 Total mesin aktif: {mesin_list.count()}")  # Debug
        
        available_machines = []
        
        for mesin in mesin_list:
            try:
                print(f"\n🔌 Connecting to: {mesin.nama} ({mesin.ip_address}:{mesin.port})")
                
                conn = connect_to_fingerprint_machine(mesin.ip_address, mesin.port, timeout=5)
                users = conn.get_users()
                
                # ✅ CEK: Apakah userid pegawai ada di mesin?
                user_exists = any(
                    str(getattr(u, 'user_id', u.uid)) == str(pegawai.userid)
                    for u in users
                )
                
                if user_exists:
                    # ✅ Ambil fingerprint count
                    templates = sync_fingerprint_template_from_machine(conn, pegawai.userid)
                    
                    available_machines.append({
                        'id': mesin.id,
                        'nama': mesin.nama,
                        'ip_address': mesin.ip_address,
                        'port': mesin.port,
                        'cabang': mesin.cabang.nama if mesin.cabang else '-',
                        'lokasi': mesin.lokasi or '-',
                        'fingerprint_count': len(templates)
                    })
                    
                    print(f"✅ User FOUND in {mesin.nama}: {len(templates)} templates")
                else:
                    print(f"⏭️ User NOT found in {mesin.nama}")
                
                conn.disconnect()
                
            except Exception as e:
                print(f"❌ Error connecting to {mesin.nama}: {str(e)}")
                continue
        
        print(f"\n📊 RESULT: {len(available_machines)} mesin ditemukan")
        
        # ✅ RETURN JSON (bukan HTML)
        if not available_machines:
            return JsonResponse({
                "status": "warning",
                "msg": f"Pegawai {pegawai.nama_lengkap} tidak ditemukan di mesin manapun yang online",
                "machines": [],
                "pegawai": {
                    "id": pegawai.id, 
                    "userid": pegawai.userid, 
                    "nama": pegawai.nama_lengkap
                }
            })
        
        return JsonResponse({
            "status": "success",
            "pegawai": {
                "id": pegawai.id, 
                "userid": pegawai.userid, 
                "nama": pegawai.nama_lengkap
            },
            "machines": available_machines,
            "total": len(available_machines)
        })
        
    except Pegawai.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "msg": f"Pegawai ID {pegawai_id} tidak ditemukan"
        }, status=404)
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"\n❌ CRITICAL ERROR:\n{error_detail}")
        
        return JsonResponse({
            "status": "error",
            "msg": f"Server error: {str(e)}"
        }, status=500)


@login_required
@require_http_methods(["GET"])
def assign_mode_jam_kerja_pegawai(request, pegawai_id):
    """View assign mode - filter by cabang"""
    
    try:
        pegawai = Pegawai.objects.get(id=pegawai_id)
    except Pegawai.DoesNotExist:
        messages.error(request, 'Pegawai tidak ditemukan')
        return redirect('daftar_Pegawai')
    
    cabang_aktif = get_active_cabang(request)
    if not cabang_aktif:
        messages.error(request, 'Silakan pilih cabang terlebih dahulu')
        return redirect('dashboard')
    
    if str(pegawai.cabang_id) != str(cabang_aktif.id):
        messages.error(request, 'Pegawai tidak terdaftar di cabang aktif')
        return redirect('daftar_Pegawai')
    
    # ✅ Filter: GLOBAL + mode sesuai cabang
    from django.db import models as django_models
    modes = MasterModeJamKerja.objects.filter(
        is_active=True
    ).filter(
        django_models.Q(cabang__isnull=True) | django_models.Q(cabang_id=pegawai.cabang_id)
    ).order_by('-is_default', '-priority', 'nama')
    
    modes_with_assignment = []
    
    for mode in modes:
        jadwal_list = ModeJamKerjaJadwal.objects.filter(
            mode_id=mode.id
        ).order_by('hari')
        
        jadwal_groups = {}
        for jadwal in jadwal_list:
            group_name = jadwal.group_name
            if group_name not in jadwal_groups:
                jadwal_groups[group_name] = {
                    'first_jadwal_id': jadwal.id,
                    'jam_masuk_display': jadwal.jam_masuk.strftime('%H:%M') if jadwal.jam_masuk else '-',
                    'jam_keluar_display': jadwal.jam_keluar.strftime('%H:%M') if jadwal.jam_keluar else '-',
                    'days': []
                }
            jadwal_groups[group_name]['days'].append(jadwal.hari)
        
        try:
            assignment = PegawaiModeAssignment.objects.get(
                pegawai_id=pegawai_id,
                mode_id=mode.id
            )
            is_assigned = True
        except PegawaiModeAssignment.DoesNotExist:
            assignment = None
            is_assigned = False
        
        mode_type = "(GLOBAL)" if mode.is_global else f"({mode.cabang.nama})"
        
        modes_with_assignment.append({
            'mode': mode,
            'mode_type': mode_type,
            'jadwal_groups': jadwal_groups,
            'assignment': assignment,
            'is_assigned': is_assigned,
        })
    
    context = {
        'pegawai': pegawai,
        'modes_with_assignment': modes_with_assignment,
        'cabang_aktif': cabang_aktif,
    }
    
    return render(request, 'absensi_app/pengaturan/mode_jam_kerja/assign_mode_pegawai.html', context)


@login_required
@require_http_methods(["POST"])
def simpan_assign_mode_jam_kerja(request, pegawai_id):
    """Handle simpan assignment mode jam kerja dengan filter cabang"""
    
    try:
        pegawai = Pegawai.objects.get(id=pegawai_id)
    except Pegawai.DoesNotExist:
        return JsonResponse({'status': 'error', 'msg': 'Pegawai tidak ditemukan'}, status=404)
    
    # Validasi cabang pegawai
    cabang_aktif = get_active_cabang(request)
    if not cabang_aktif:
        return JsonResponse({'status': 'error', 'msg': 'Silakan pilih cabang terlebih dahulu'}, status=400)
    
    if str(pegawai.cabang_id) != str(cabang_aktif.id):
        return JsonResponse({'status': 'error', 'msg': 'Pegawai tidak ditemukan di cabang aktif'}, status=403)
    
    # Parse assignments
    try:
        assignments = json.loads(request.POST.get('assignments', '{}'))
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'msg': 'Format data tidak valid'}, status=400)
    
    if not assignments:
        return JsonResponse({'status': 'error', 'msg': 'Minimal 1 mode harus dipilih'}, status=400)
    
    # Process assignments
    from django.db import models as django_models
    created_count = 0
    updated_count = 0
    errors = []
    
    for mode_id_str, assignment_data in assignments.items():
        try:
            mode_id = int(mode_id_str)
            
            # ✅ Validasi: Mode harus GLOBAL atau sesuai cabang pegawai
            mode = MasterModeJamKerja.objects.get(
                id=mode_id,
                is_active=True
            )
            
            # ✅ CEK: Apakah mode bisa digunakan untuk pegawai ini?
            if mode.cabang_id is not None and mode.cabang_id != pegawai.cabang_id:
                errors.append(
                    f"Mode '{mode.nama}' adalah untuk cabang '{mode.cabang.nama}', "
                    f"tidak sesuai dengan pegawai di cabang '{pegawai.cabang.nama}'"
                )
                continue
            
            # Validasi jadwal
            group_id = assignment_data.get('group_id')
            jadwal = ModeJamKerjaJadwal.objects.get(id=group_id, mode_id=mode_id)
            
            # Build jadwal_per_hari
            jadwal_per_hari = {}
            jadwal_list = ModeJamKerjaJadwal.objects.filter(
                mode_id=mode_id,
                group_name=jadwal.group_name
            ).order_by('hari')
            
            for j in jadwal_list:
                jadwal_per_hari[str(j.hari)] = j.id
            
            # Save/Update assignment
            assignment, created = PegawaiModeAssignment.objects.update_or_create(
                pegawai_id=pegawai_id,
                mode_id=mode_id,
                defaults={
                    'jadwal_per_hari': jadwal_per_hari,
                    'is_active': True
                }
            )
            
            if created:
                created_count += 1
            else:
                updated_count += 1
        
        except (ValueError, MasterModeJamKerja.DoesNotExist, ModeJamKerjaJadwal.DoesNotExist) as e:
            errors.append(f"Mode {mode_id}: {str(e)}")
    
    # Build response
    msg = f"✅ Assignment disimpan!\n"
    msg += f"• Mode ditambah: {created_count}\n"
    msg += f"• Mode diperbarui: {updated_count}"
    
    if errors:
        msg += f"\n\n⚠️ Errors:\n" + "\n".join(errors)
    
    # Clear cache
    from .services import WorkModeService
    WorkModeService.clear_cache()
    
    return JsonResponse({
        'status': 'success',
        'msg': msg,
        'created': created_count,
        'updated': updated_count
    })


@login_required
def daftar_assign_mode_pegawai(request):
    """
    Halaman daftar semua pegawai dan mode jam kerja yang di-assign.
    
     Filter cabang diterapkan
    """
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    cabang_aktif = get_active_cabang(request)
    
    #  Filter pegawai berdasarkan cabang
    pegawai_list = Pegawai.objects.filter(is_active=True)
    
    if cabang_aktif:
        pegawai_list = pegawai_list.filter(cabang=cabang_aktif)
    
    pegawai_list = pegawai_list.select_related(
        'departemen', 'jabatan', 'cabang'
    ).order_by('nama_lengkap')
    
    all_modes = MasterModeJamKerja.objects.filter(
        is_active=True
    ).order_by('-is_default', '-priority', 'nama')
    
    from .models import PegawaiModeAssignment
    
    pegawai_data = []
    for pegawai in pegawai_list:
        assignments = PegawaiModeAssignment.objects.filter(
            pegawai=pegawai,
            is_active=True
        ).select_related('mode')
        
        mode_list = [a.mode.nama for a in assignments]
        
        pegawai_data.append({
            'pegawai': pegawai,
            'modes': mode_list,
            'mode_count': len(assignments),
            'has_modes': len(assignments) > 0
        })
    
    context = {
        'pegawai_data': pegawai_data,
        'all_modes': all_modes,
        'cabang_aktif': cabang_aktif,
    }
    
    return render(request, 'absensi_app/pegawai/daftar_assign_mode.html', context)


@login_required
def api_get_mode_assignment_form(request, pegawai_id):
    """Endpoint API: Mendapatkan data form mode assignment untuk ditampilkan di modal."""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        pegawai = get_object_or_404(Pegawai, pk=pegawai_id)
        
        from .models import MasterModeJamKerja, PegawaiModeAssignment
        
        all_modes = MasterModeJamKerja.objects.filter(
            is_active=True
        ).prefetch_related('jadwal_list', 'periode_list').order_by('-is_default', '-priority', 'nama')
        
        existing_assignments = PegawaiModeAssignment.objects.filter(
            pegawai=pegawai,
            is_active=True
        ).select_related('mode')
        
        modes_data = []
        
        for mode in all_modes:
            assignment = existing_assignments.filter(mode=mode).first()
            
            jadwal_groups = {}
            
            for jadwal in mode.jadwal_list.all():
                group_name = jadwal.group_name
                hari = jadwal.hari
                
                if group_name not in jadwal_groups:
                    jadwal_groups[group_name] = {}
                
                if hari not in jadwal_groups[group_name]:
                    jadwal_groups[group_name][hari] = []
                
                jadwal_groups[group_name][hari].append({
                    'id': jadwal.id,
                    'jam_masuk': jadwal.jam_masuk.strftime('%H:%M') if jadwal.jam_masuk else '-',
                    'jam_keluar': jadwal.jam_keluar.strftime('%H:%M') if jadwal.jam_keluar else '-',
                    'display': f"{group_name} ({jadwal.jam_masuk.strftime('%H:%M')}-{jadwal.jam_keluar.strftime('%H:%M')})" if jadwal.jam_masuk and jadwal.jam_keluar else group_name
                })
            
            active_periode = mode.periode_list.filter(
                is_active=True,
                tanggal_mulai__lte=date.today(),
                tanggal_selesai__gte=date.today()
            ).first()
            
            modes_data.append({
                'id': mode.id,
                'nama': mode.nama,
                'kode': mode.kode,
                'warna': mode.warna,
                'icon': mode.icon,
                'is_default': mode.is_default,
                'is_assigned': assignment is not None,
                'assignment_data': {
                    'jadwal_per_hari': assignment.jadwal_per_hari if assignment else {}
                } if assignment else {},
                'jadwal_groups': jadwal_groups,
                'periode_aktif': {
                    'nama': active_periode.nama,
                    'tanggal_mulai': active_periode.tanggal_mulai.strftime('%d %b %Y'),
                    'tanggal_selesai': active_periode.tanggal_selesai.strftime('%d %b %Y')
                } if active_periode else None,
                'hari_names': ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
            })
        
        return JsonResponse({
            "status": "success",
            "pegawai": {
                'id': pegawai.id,
                'userid': pegawai.userid,
                'nama': pegawai.nama_lengkap
            },
            "modes": modes_data
        })
        
    except Pegawai.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "msg": "Pegawai tidak ditemukan"
        }, status=404)
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "msg": str(e)
        }, status=500)

@login_required
def api_get_mode_assignments(request, pegawai_id):
    """✅ FIXED: API untuk mengambil assignment mode pegawai dengan data jadwal lengkap"""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        from .models import Pegawai, MasterModeJamKerja, PegawaiModeAssignment, ModeJamKerjaJadwal
        
        pegawai = get_object_or_404(Pegawai, pk=pegawai_id)
        
        all_modes = MasterModeJamKerja.objects.filter(
            is_active=True
        ).prefetch_related('jadwal_list').order_by('-is_default', '-priority', 'nama')
        
        existing_assignments = PegawaiModeAssignment.objects.filter(
            pegawai=pegawai,
            is_active=True
        ).select_related('mode')
        
        modes_data = []
        
        for mode in all_modes:
            assignment = existing_assignments.filter(mode=mode).first()
            
            # ✅ BUILD JADWAL GROUPS
            jadwal_groups = {}
            
            for jadwal in mode.jadwal_list.all().order_by('group_name', 'hari'):
                group_name = jadwal.group_name
                hari = jadwal.hari
                
                if group_name not in jadwal_groups:
                    jadwal_groups[group_name] = {
                        'name': group_name,
                        'days': {},
                        'first_jadwal_id': jadwal.id,  # ← PENTING
                        'jam_display': f"{jadwal.jam_masuk.strftime('%H:%M')}-{jadwal.jam_keluar.strftime('%H:%M')}" if jadwal.jam_masuk and jadwal.jam_keluar else '-'
                    }
                
                if hari not in jadwal_groups[group_name]['days']:
                    jadwal_groups[group_name]['days'][hari] = []
                
                jadwal_groups[group_name]['days'][hari].append({
                    'id': jadwal.id,
                    'jam_masuk': jadwal.jam_masuk.strftime('%H:%M') if jadwal.jam_masuk else '-',
                    'jam_keluar': jadwal.jam_keluar.strftime('%H:%M') if jadwal.jam_keluar else '-',
                })
            
            modes_data.append({
                'id': mode.id,
                'nama': mode.nama,
                'kode': mode.kode,
                'warna': mode.warna,
                'icon': mode.icon,
                'is_default': mode.is_default,
                'is_assigned': assignment is not None,
                'assignment_data': {
                    'jadwal_per_hari': assignment.jadwal_per_hari if assignment else {}
                } if assignment else {},
                'jadwal_groups': jadwal_groups,  # ← KIRIM JADWAL GROUPS
                'hari_names': ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
            })
        
        print(f"\n📊 DEBUG: Loaded {len(modes_data)} modes for pegawai {pegawai.nama_lengkap}")
        
        return JsonResponse({
            "status": "success",
            "pegawai": {
                'id': pegawai.id,
                'userid': pegawai.userid,
                'nama': pegawai.nama_lengkap
            },
            "modes": modes_data
        })
        
    except Pegawai.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "msg": "Pegawai tidak ditemukan"
        }, status=404)
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"❌ ERROR api_get_mode_assignments:\n{error_detail}")
        
        return JsonResponse({
            "status": "error",
            "msg": str(e)
        }, status=500)
        
@login_required
def api_save_mode_assignment_bulk(request):
    """
    ✅ FIXED: ADDITIVE MODE dengan jadwal_per_hari LENGKAP (semua hari)
    
    BEFORE: Hanya simpan 1 jadwal → detail pegawai hanya tampil 1 hari
    AFTER: Simpan SEMUA jadwal dalam group → detail pegawai tampil lengkap
    """
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({"status": "error", "msg": "Method tidak diizinkan"}, status=405)
    
    try:
        import json
        from django.db import transaction
        from .models import Pegawai, PegawaiModeAssignment, MasterModeJamKerja, ModeJamKerjaJadwal
        
        assignments_json = request.POST.get('assignments', '{}')
        
        print("\n" + "="*60)
        print("✅ BULK ASSIGN MODE JAM KERJA - FIXED VERSION")
        print("="*60)
        
        try:
            assignments_list = json.loads(assignments_json)
        except json.JSONDecodeError as e:
            return JsonResponse({
                "status": "error",
                "msg": f"Format JSON tidak valid: {str(e)}"
            }, status=400)
        
        print(f"📊 Total pegawai: {len(assignments_list)}")
        
        total_added = 0
        total_updated = 0
        total_failed = 0
        errors = []
        
        with transaction.atomic():
            for assignment_data in assignments_list:
                try:
                    pegawai_id = assignment_data.get('pegawai_id')
                    modes_data = assignment_data.get('modes', {})
                    
                    print(f"\n{'─'*60}")
                    print(f"👤 Pegawai ID: {pegawai_id}")
                    print(f"{'─'*60}")
                    
                    pegawai = Pegawai.objects.get(id=pegawai_id)
                    print(f"   Nama: {pegawai.nama_lengkap}")
                    
                    for mode_id_str, mode_assignment in modes_data.items():
                        try:
                            mode_id = int(mode_id_str)
                            mode = MasterModeJamKerja.objects.get(id=mode_id)
                            
                            print(f"\n   🔧 Mode: {mode.nama}")
                            
                            # ========================================
                            # ✅ CRITICAL FIX: Build jadwal_per_hari LENGKAP
                            # ========================================
                            
                            jadwal_per_hari_data = mode_assignment.get('jadwal_per_hari', {})
                            
                            print(f"   📥 Received jadwal_per_hari: {jadwal_per_hari_data}")
                            
                            # ✅ VALIDASI: Pastikan data lengkap
                            if not jadwal_per_hari_data or not any(v for v in jadwal_per_hari_data.values()):
                                print(f"   ⚠️  WARNING: jadwal_per_hari kosong!")
                                
                                # ✅ FALLBACK: Build dari group_id jika ada
                                group_id = mode_assignment.get('group_id')
                                if group_id:
                                    print(f"   🔄 Fallback: Building dari group_id={group_id}")
                                    
                                    # ✅ Ambil SEMUA jadwal dalam group ini
                                    jadwal_list = ModeJamKerjaJadwal.objects.filter(
                                        mode_id=mode_id,
                                        id=group_id
                                    ).first()
                                    
                                    if jadwal_list:
                                        # ✅ Ambil SEMUA jadwal dengan group_name yang sama
                                        all_jadwal_in_group = ModeJamKerjaJadwal.objects.filter(
                                            mode_id=mode_id,
                                            group_name=jadwal_list.group_name
                                        ).order_by('hari')
                                        
                                        jadwal_per_hari_data = {}
                                        for jdl in all_jadwal_in_group:
                                            jadwal_per_hari_data[str(jdl.hari)] = jdl.id
                                        
                                        print(f"   ✅ Built jadwal_per_hari: {jadwal_per_hari_data}")
                                    else:
                                        errors.append(f"{pegawai.nama_lengkap} - Mode {mode.nama}: Group ID {group_id} tidak ditemukan")
                                        total_failed += 1
                                        continue
                                else:
                                    errors.append(f"{pegawai.nama_lengkap} - Mode {mode.nama}: Tidak ada jadwal")
                                    total_failed += 1
                                    continue
                            
                            # ========================================
                            # ✅ VALIDASI FINAL: Pastikan semua jadwal valid
                            # ========================================
                            
                            valid_jadwal_ids = set(
                                ModeJamKerjaJadwal.objects.filter(
                                    mode_id=mode_id
                                ).values_list('id', flat=True)
                            )
                            
                            cleaned_jadwal = {}
                            for hari_str, jadwal_id in jadwal_per_hari_data.items():
                                if jadwal_id in valid_jadwal_ids:
                                    cleaned_jadwal[str(hari_str)] = jadwal_id
                                else:
                                    print(f"   ⚠️  Jadwal ID {jadwal_id} tidak valid, skip")
                            
                            if not cleaned_jadwal:
                                errors.append(f"{pegawai.nama_lengkap} - Mode {mode.nama}: Tidak ada jadwal valid")
                                total_failed += 1
                                continue
                            
                            print(f"   ✅ Final cleaned jadwal: {cleaned_jadwal}")
                            print(f"   📊 Total hari: {len(cleaned_jadwal)}")
                            
                            # ========================================
                            # ✅ SAVE/UPDATE ASSIGNMENT
                            # ========================================
                            
                            assignment, created = PegawaiModeAssignment.objects.update_or_create(
                                pegawai=pegawai,
                                mode=mode,
                                defaults={
                                    'jadwal_per_hari': cleaned_jadwal,
                                    'is_active': True
                                }
                            )
                            
                            if created:
                                total_added += 1
                                action = "ADDED"
                                icon = "➕"
                            else:
                                total_updated += 1
                                action = "UPDATED"
                                icon = "🔄"
                            
                            print(f"   ✅ {icon} {action}: {mode.nama} ({len(cleaned_jadwal)} hari)")
                            
                        except MasterModeJamKerja.DoesNotExist:
                            errors.append(f"{pegawai.nama_lengkap} - Mode {mode_id}: Mode tidak ditemukan")
                            total_failed += 1
                            print(f"   ❌ Mode {mode_id} not found")
                        except Exception as e:
                            errors.append(f"{pegawai.nama_lengkap} - Mode {mode_id}: {str(e)}")
                            total_failed += 1
                            print(f"   ❌ Exception: {str(e)}")
                    
                except Pegawai.DoesNotExist:
                    errors.append(f"Pegawai ID {pegawai_id}: Tidak ditemukan")
                    total_failed += 1
                    print(f"❌ Pegawai ID {pegawai_id} not found")
                except Exception as e:
                    errors.append(f"Pegawai ID {pegawai_id}: {str(e)}")
                    total_failed += 1
                    print(f"❌ Exception: {str(e)}")
        
        print(f"\n{'='*60}")
        print("📊 SUMMARY")
        print(f"{'='*60}")
        print(f"✅ Added: {total_added}")
        print(f"🔄 Updated: {total_updated}")
        print(f"❌ Failed: {total_failed}")
        print(f"{'='*60}\n")
        
        # ========================================
        # ✅ CLEAR CACHE SETELAH SUKSES
        # ========================================
        
        if total_added > 0 or total_updated > 0:
            from django.core.cache import cache
            from .services import WorkModeService
            
            print("🧹 Clearing cache...")
            
            # Clear global cache
            WorkModeService.clear_cache()
            
            # Clear pegawai-specific cache
            for assignment_data in assignments_list:
                pegawai_id = assignment_data.get('pegawai_id')
                if pegawai_id:
                    cache_key = f"pegawai_mode_detail_{pegawai_id}"
                    cache.delete(cache_key)
                    print(f"   ✅ Cleared cache for pegawai {pegawai_id}")
        
        # ========================================
        # ✅ BUILD RESPONSE MESSAGE
        # ========================================
        
        if total_added == 0 and total_updated == 0 and total_failed == 0:
            return JsonResponse({
                "status": "info",
                "msg": "ℹ️ Tidak ada perubahan",
                "added": 0,
                "updated": 0,
                "failed": 0
            })
        
        msg = "✅ Berhasil!\n\n"
        if total_added > 0:
            msg += f"• ➕ Mode ditambahkan: {total_added}\n"
        if total_updated > 0:
            msg += f"• 🔄 Mode diperbarui: {total_updated}\n"
        if total_failed > 0:
            msg += f"\n⚠️ {total_failed} assignment gagal"
            if errors:
                msg += "\n\nDetail Error:\n"
                for err in errors[:5]:
                    msg += f"• {err}\n"
                if len(errors) > 5:
                    msg += f"• ... dan {len(errors) - 5} error lainnya\n"
        
        msg += f"\n\n💡 Refresh halaman detail pegawai untuk melihat perubahan."
        
        return JsonResponse({
            "status": "success" if (total_added > 0 or total_updated > 0) else "warning",
            "msg": msg,
            "added": total_added,
            "updated": total_updated,
            "failed": total_failed
        })
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"\n❌ CRITICAL ERROR:\n{error_detail}")
        
        return JsonResponse({
            "status": "error",
            "msg": f"Server error: {str(e)}"
        }, status=500)    
    
# ==============================================================================
# IMPORT TAP DARI MESIN → SIMPAN KE TAP LOG
# ==============================================================================

@login_required
def sync_tap_to_log(request):
    """
     Sync data tap dari mesin ke TapLog dengan PROPER RESPONSE
    
    Response Format:
    {
        'status': 'success',
        'total_tap_saved': int,      
        'total_tap_duplicate': int,  
        'mesin_success': int,        
        'mesin_failed': int,         
        'message': str
    }
    """
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        from .models import MasterMesin, Pegawai, TapLog
        from datetime import datetime
        import concurrent.futures
        import threading
        
        # Parse date filter
        tanggal_mulai = request.GET.get('tanggal_mulai')
        tanggal_akhir = request.GET.get('tanggal_akhir')
        
        if tanggal_mulai:
            tanggal_mulai = datetime.strptime(tanggal_mulai, '%Y-%m-%d').date()
        
        if tanggal_akhir:
            tanggal_akhir = datetime.strptime(tanggal_akhir, '%Y-%m-%d').date()
        
        # Ambil mesin aktif
        mesin_list = MasterMesin.objects.filter(is_active=True)
        
        if not mesin_list.exists():
            return JsonResponse({
                "status": "error",
                "message": " Tidak ada mesin aktif!"
            }, status=404)
        
        # ========================================
        #  PARALLEL PROCESSING 
        # ========================================
        
        total_tap_saved = 0
        total_tap_duplicate = 0
        mesin_success = 0
        mesin_failed = 0
        mesin_results = []
        
        # Thread-safe counter
        counter_lock = threading.Lock()
        
        def process_mesin(mesin):
            """Process single mesin (thread-safe)"""
            nonlocal total_tap_saved, total_tap_duplicate, mesin_success, mesin_failed
            
            try:
                # Connect ke mesin
                conn = connect_to_fingerprint_machine(
                    mesin.ip_address,
                    mesin.port,
                    timeout=10  
                )
                conn.disable_device()
                
                # Ambil semua tap
                attendances = conn.get_attendance()
                
                conn.enable_device()
                conn.disconnect()
                
                if not attendances:
                    return {
                        'mesin': mesin.nama,
                        'status': 'no_data',
                        'tap_saved': 0,
                        'tap_duplicate': 0
                    }
                
                # Filter by date
                if tanggal_mulai or tanggal_akhir:
                    filtered = []
                    
                    for att in attendances:
                        att_date = att.timestamp.date()
                        
                        if tanggal_mulai and att_date < tanggal_mulai:
                            continue
                        if tanggal_akhir and att_date > tanggal_akhir:
                            continue
                        
                        filtered.append(att)
                    
                    attendances = filtered
                
                # Simpan ke TapLog
                saved_count = 0
                duplicate_count = 0
                
                for att in attendances:
                    try:
                        pegawai = Pegawai.objects.filter(
                            userid=str(att.user_id)
                        ).first()
                        
                        if not pegawai:
                            continue
                        
                        #  Check duplikat berdasarkan timestamp EXACT
                        exists = TapLog.objects.filter(
                            pegawai=pegawai,
                            tanggal=att.timestamp.date(),
                            waktu_tap=att.timestamp.time(),
                            punch_type=att.punch,
                            mesin=mesin
                        ).exists()
                        
                        if exists:
                            duplicate_count += 1
                            continue
                        
                        # Create new TapLog
                        TapLog.objects.create(
                            pegawai=pegawai,
                            tanggal=att.timestamp.date(),
                            waktu_tap=att.timestamp.time(),
                            punch_type=att.punch,
                            mesin=mesin,
                            is_processed=False
                        )
                        
                        saved_count += 1
                    
                    except Exception as e:
                        print(f"ERROR saving tap: {str(e)}")
                        continue
                
                # Update counter (thread-safe)
                with counter_lock:
                    total_tap_saved += saved_count
                    total_tap_duplicate += duplicate_count
                    mesin_success += 1
                
                return {
                    'mesin': mesin.nama,
                    'status': 'success',
                    'tap_saved': saved_count,
                    'tap_duplicate': duplicate_count
                }
                
            except Exception as e:
                with counter_lock:
                    mesin_failed += 1
                
                return {
                    'mesin': mesin.nama,
                    'status': 'error',
                    'error': str(e)
                }
        
        # ========================================
        #  EXECUTE PARALLEL (Max 4 threads)
        # ========================================
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
            futures = [executor.submit(process_mesin, mesin) for mesin in mesin_list]
            mesin_results = [future.result() for future in concurrent.futures.as_completed(futures)]
        
        # ========================================
        #  BUILD RESPONSE MESSAGE
        # ========================================
        
        if total_tap_saved == 0 and total_tap_duplicate == 0 and mesin_failed == len(mesin_list):
            # Semua mesin offline
            msg = f" Semua mesin offline!\n\n"
            msg += f"📡 Total mesin: {len(mesin_list)}\n"
            msg += f" Mesin gagal: {mesin_failed}\n\n"
            msg += "Detail:\n"
            for mr in mesin_results:
                if mr['status'] == 'error':
                    msg += f"• {mr['mesin']}: {mr['error']}\n"
            
            return JsonResponse({
                "status": "error",
                "total_tap_saved": 0,
                "total_tap_duplicate": 0,
                "mesin_success": 0,
                "mesin_failed": mesin_failed,
                "message": msg
            })
        
        elif total_tap_saved == 0 and total_tap_duplicate > 0:
            # Semua data sudah di-sync sebelumnya
            msg = f"ℹ️ Data sudah di-sync sebelumnya!\n\n"
            msg += f"📊 Hasil Sync:\n"
            msg += f"•  Tap baru tersimpan: 0\n"
            msg += f"• ⏭️ Tap duplikat (skip): {total_tap_duplicate}\n"
            msg += f"• 📡 Mesin berhasil: {mesin_success} / {len(mesin_list)}\n"
            
            if mesin_failed > 0:
                msg += f"•  Mesin gagal: {mesin_failed}\n"
            
            msg += f"\n💡 Next Step:\n"
            msg += f"Data tap sudah ada. Langsung klik tombol \"Proses ke Sesi\" untuk mengubah TapLog menjadi AbsensiSesi."
            
            return JsonResponse({
                "status": "success",
                "total_tap_saved": 0,
                "total_tap_duplicate": total_tap_duplicate,
                "mesin_success": mesin_success,
                "mesin_failed": mesin_failed,
                "message": msg
            })
        
        else:
            # Ada data baru
            msg = f" Sync Berhasil!\n\n"
            msg += f"📊 Hasil Sync:\n"
            msg += f"•  Tap baru tersimpan: {total_tap_saved}\n"
            msg += f"• ⏭️ Tap duplikat (skip): {total_tap_duplicate}\n"
            msg += f"• 📡 Mesin berhasil: {mesin_success} / {len(mesin_list)}\n"
            
            if mesin_failed > 0:
                msg += f"•  Mesin gagal: {mesin_failed}\n"
            
            msg += f"\nDetail per Mesin:\n"
            for mr in mesin_results:
                if mr['status'] == 'success':
                    msg += f"• {mr['mesin']}: {mr['tap_saved']} tap baru"
                    if mr['tap_duplicate'] > 0:
                        msg += f", {mr['tap_duplicate']} duplikat"
                    msg += "\n"
            
            msg += f"\n💡 Langkah Selanjutnya:\n"
            msg += f"Klik tombol \"Proses ke Sesi\" untuk mengubah TapLog menjadi AbsensiSesi."
            
            return JsonResponse({
                "status": "success",
                "total_tap_saved": total_tap_saved,
                "total_tap_duplicate": total_tap_duplicate,
                "mesin_success": mesin_success,
                "mesin_failed": mesin_failed,
                "message": msg
            })
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"ERROR sync_tap_to_log: {error_detail}")
        
        return JsonResponse({
            "status": "error",
            "message": f" Error: {str(e)}"
        }, status=500)


# ==============================================================================
# PROSES TAP LOG → BUAT SESI ABSENSI
# ==============================================================================
@login_required
def proses_tap_to_sesi(request):
    """
    Endpoint AJAX: Proses TapLog menjadi AbsensiSesi
    
     Proses SEMUA tap (bukan hanya hari ini)
    """
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        from .services import TapStackProcessor
        
        #  Ganti dari proses_semua_tap_hari_ini() → proses_semua_tap()
        result = TapStackProcessor.proses_semua_tap()
        
        if result['status'] == 'success':
            msg = f" Berhasil memproses tap menjadi sesi!\n\n"
            msg += f"Detail:\n"
            msg += f"• Total Pegawai: {result['total_pegawai']}\n"
            msg += f"• Total Sesi: {result['total_sesi']}\n"
            msg += f"• Total Tap: {result['total_tap']}\n\n"
            
            if result['detail']:
                msg += "Per Pegawai:\n"
                for item in result['detail'][:5]:
                    msg += f"• {item['pegawai_nama']}: {item['sesi_count']} sesi\n"
                
                if len(result['detail']) > 5:
                    msg += f"• ... dan {len(result['detail']) - 5} pegawai lainnya\n"
            
            return JsonResponse({
                "status": "success",
                "total_pegawai": result['total_pegawai'],
                "total_sesi": result['total_sesi'],
                "total_tap": result['total_tap'],
                "message": msg
            })
        else:
            return JsonResponse({
                "status": "error",
                "message": result.get('message', 'Tidak ada tap yang perlu diproses')
            })
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"ERROR proses_tap_to_sesi: {error_detail}")
        
        return JsonResponse({
            "status": "error",
            "message": f"Error: {str(e)}\n\nDetail:\n{error_detail}"
        }, status=500)

# ==============================================================================
# VIEW RIWAYAT SESI ABSENSI
# ==============================================================================

@login_required
def riwayat_sesi_absensi(request):
    """
    Halaman untuk melihat riwayat sesi absensi pegawai
    
     Menambahkan filter cabang yang konsisten
    """
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    from .models import AbsensiSesi, Pegawai
    from .forms import LaporanFilterForm
    from collections import defaultdict

    #  Ambil cabang aktif dari session
    cabang_aktif = get_active_cabang(request)
    
    form = LaporanFilterForm(request.GET)
    
    sesi_list = AbsensiSesi.objects.select_related(
        'pegawai', 
        'pegawai__departemen',
        'pegawai__cabang'  
    ).all() 
    
    #  Filter berdasarkan cabang aktif
    if cabang_aktif:
        sesi_list = sesi_list.filter(pegawai__cabang=cabang_aktif)
    
    if form.is_valid():
        tgl_mulai = form.cleaned_data.get('tanggal_mulai')
        tgl_akhir = form.cleaned_data.get('tanggal_akhir')
        
        if tgl_mulai:
            sesi_list = sesi_list.filter(tanggal_mulai__gte=tgl_mulai)
        if tgl_akhir:
            sesi_list = sesi_list.filter(tanggal_mulai__lte=tgl_akhir)
        
        search_employee = form.cleaned_data.get('search_employee')
        if search_employee:
            sesi_list = sesi_list.filter(
                Q(pegawai__nama_lengkap__icontains=search_employee) |
                Q(pegawai__userid__icontains=search_employee)
            )
    
    sesi_list = sesi_list.order_by(
        '-tanggal_mulai',
        'pegawai__userid',
        'tap_masuk_pertama'
    )
    
    # HITUNG NOMOR SESI PER HARI PER PEGAWAI
    sesi_count_per_day = defaultdict(int)
    for sesi in sesi_list:
        key = (sesi.pegawai.id, sesi.tanggal_mulai)
        sesi_count_per_day[key] += 1
    
    sesi_urutan = defaultdict(int)
    sesi_list_with_number = []
    
    for sesi in sesi_list:
        key = (sesi.pegawai.id, sesi.tanggal_mulai)
        sesi_urutan[key] += 1
        
        sesi.nomor_sesi = sesi_urutan[key]
        sesi.total_sesi_hari_ini = sesi_count_per_day[key]
        
        sesi_list_with_number.append(sesi)
    
    context = {
        'sesi_list': sesi_list_with_number,
        'form': form,
        'cabang_aktif': cabang_aktif,  
        'page_title': 'Riwayat Sesi Absensi',
    }
    
    return render(request, 'absensi_app/sesi/riwayat_sesi.html', context)

@login_required
def riwayat_sesi_per_pegawai(request, pk):
    """
    Halaman riwayat sesi absensi per pegawai
    
     Menambahkan filter cabang yang konsisten
    """
    if not request.user.is_staff:
        messages.error(request, "Akses ditolak.")
        return redirect('dashboard')
    
    from .models import Pegawai, AbsensiSesi
    from .forms import LaporanFilterForm
    from .services import TapStackProcessor
    
    pegawai = get_object_or_404(Pegawai, pk=pk)
    
    #  Validasi bahwa pegawai sesuai cabang aktif
    cabang_aktif = get_active_cabang(request)
    
    if cabang_aktif and pegawai.cabang != cabang_aktif:
        messages.error(
            request, 
            f"Pegawai {pegawai.nama_lengkap} tidak terdaftar di cabang {cabang_aktif.nama}"
        )
        return redirect('riwayat_sesi_absensi')
    
    form = LaporanFilterForm(request.GET)
    
    # Default: 1 bulan terakhir
    today = date.today()
    tanggal_mulai = today - timedelta(days=30)
    tanggal_akhir = today
    
    if form.is_valid():
        if form.cleaned_data.get('tanggal_mulai'):
            tanggal_mulai = form.cleaned_data['tanggal_mulai']
        if form.cleaned_data.get('tanggal_akhir'):
            tanggal_akhir = form.cleaned_data['tanggal_akhir']
    
    # Ambil summary sesi
    summary = TapStackProcessor.get_sesi_summary_untuk_pegawai(
        pegawai, 
        tanggal_mulai, 
        tanggal_akhir
    )
    
    context = {
        'pegawai': pegawai,
        'form': form,
        'sesi_list': summary['sesi_list'],
        'sesi_per_hari': summary['sesi_per_hari'],
        'total_sesi': summary['total_sesi'],
        'total_hari_kerja': summary['total_hari_kerja'],
        'tanggal_mulai': tanggal_mulai,
        'tanggal_akhir': tanggal_akhir,
        'cabang_aktif': cabang_aktif,  
    }
    
    return render(request, 'absensi_app/sesi/riwayat_sesi_per_pegawai.html', context)

# API - TAP LOG & SESI
@login_required
def api_get_tap_logs(request, pegawai_id):
    """
    API: Ambil semua tap logs untuk pegawai
    
    Returns:
        JSON: List tap logs
    """
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        from .models import Pegawai, TapLog
        
        pegawai = get_object_or_404(Pegawai, pk=pegawai_id)
        
        tanggal = request.GET.get('tanggal')
        if tanggal:
            tanggal = datetime.strptime(tanggal, '%Y-%m-%d').date()
        else:
            tanggal = date.today()
        
        tap_logs = TapLog.objects.filter(
            pegawai=pegawai,
            tanggal=tanggal
        ).order_by('waktu_tap')
        
        data = []
        for tap in tap_logs:
            data.append({
                'id': tap.id,
                'waktu_tap': tap.waktu_tap.strftime('%H:%M:%S'),
                'punch_type': tap.get_punch_type_display(),
                'punch_code': tap.punch_type,
                'is_processed': tap.is_processed,
                'mesin': tap.mesin.nama if tap.mesin else '-'
            })
        
        return JsonResponse({
            "status": "success",
            "pegawai": {
                'id': pegawai.id,
                'userid': pegawai.userid,
                'nama': pegawai.nama_lengkap
            },
            "tanggal": tanggal.strftime('%Y-%m-%d'),
            "tap_logs": data,
            "total": len(data)
        })
        
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@login_required
def api_get_sesi_detail(request, sesi_id):
    """
    API: Ambil detail sesi absensi beserta tap logs
    
    Returns:
        JSON: Detail sesi + list tap logs
    """
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        from .models import AbsensiSesi
        
        sesi = get_object_or_404(AbsensiSesi, pk=sesi_id)
        
        # Ambil tap logs via relation
        tap_relations = sesi.tap_relation.select_related('tap_log', 'tap_log__mesin').order_by('urutan_dalam_sesi')
        
        tap_logs_data = []
        for rel in tap_relations:
            tap = rel.tap_log
            tap_logs_data.append({
                'waktu_tap': tap.waktu_tap.strftime('%H:%M:%S'),
                'punch_type_display': tap.get_punch_type_display(),
                'punch_type_code': tap.punch_type,
                'mesin_nama': tap.mesin.nama if tap.mesin else '-',
                'urutan': rel.urutan_dalam_sesi
            })
        
        return JsonResponse({
            "status": "success",
            "data": {
                "sesi": {
                    'id': sesi.id,
                    'pegawai_id': sesi.pegawai.id,
                    'pegawai_nama': sesi.pegawai.nama_lengkap,
                    'pegawai_userid': sesi.pegawai.userid,
                    'tanggal_mulai': sesi.tanggal_mulai.strftime('%d %b %Y'),
                    'tanggal_selesai': sesi.tanggal_selesai.strftime('%d %b %Y'),
                    'tap_masuk_pertama': sesi.tap_masuk_pertama.strftime('%H:%M'),
                    'tap_masuk_terakhir': sesi.tap_masuk_terakhir.strftime('%H:%M') if sesi.tap_masuk_terakhir else '-',
                    'tap_pulang_pertama': sesi.tap_pulang_pertama.strftime('%H:%M') if sesi.tap_pulang_pertama else '-',
                    'tap_pulang_terakhir': sesi.tap_pulang_terakhir.strftime('%H:%M') if sesi.tap_pulang_terakhir else '-',
                    'jumlah_tap_masuk': sesi.jumlah_tap_masuk,
                    'jumlah_tap_pulang': sesi.jumlah_tap_pulang,
                    'status': sesi.status,
                    'is_cross_day': sesi.is_cross_day,  
                    'durasi': sesi.get_durasi_formatted()
                },
                "tap_logs": tap_logs_data
            }
        })
        
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)

# STATISTIK SESI
@login_required
def api_get_sesi_statistics(request):
    """
    API: Ambil statistik sesi hari ini
    
    Returns:
        JSON: {
            'total_sesi': int,
            'total_pegawai': int,
            'sesi_complete': int,
            'sesi_incomplete': int,
            ...
        }
    """
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        from .models import AbsensiSesi
        from django.db.models import Count
        
        today = date.today()
        
        sesi_today = AbsensiSesi.objects.filter(tanggal_mulai=today)
        
        stats = sesi_today.aggregate(
            total_sesi=Count('id'),
            sesi_complete=Count('id', filter=Q(status='Hadir')),
            sesi_incomplete=Count('id', filter=Q(status='Incomplete'))
        )
        
        # Hitung pegawai unik
        pegawai_ids = sesi_today.values_list('pegawai_id', flat=True).distinct()
        
        # Pegawai dengan multiple sesi
        multiple_sesi = sesi_today.values('pegawai').annotate(
            sesi_count=Count('id')
        ).filter(sesi_count__gt=1)
        
        return JsonResponse({
            "status": "success",
            "tanggal": today.strftime('%Y-%m-%d'),
            "statistics": {
                'total_sesi': stats['total_sesi'] or 0,
                'total_pegawai': len(pegawai_ids),
                'sesi_complete': stats['sesi_complete'] or 0,
                'sesi_incomplete': stats['sesi_incomplete'] or 0,
                'pegawai_multiple_sesi': multiple_sesi.count()
            }
        })
        
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)

@login_required
def hapus_sesi_absensi(request, pk):
    """
    View untuk menghapus sesi absensi beserta semua tap logs terkait
    """
    from .models import AbsensiSesi  
    
    sesi = get_object_or_404(AbsensiSesi, pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Simpan info untuk pesan
                pegawai_nama = sesi.pegawai.nama_lengkap
                tanggal_mulai = sesi.tanggal_mulai
                tanggal_selesai = sesi.tanggal_selesai
                
                # Hitung jumlah tap logs yang akan terhapus
                jumlah_tap_logs = sesi.tap_relation.count()
                
                # Hapus sesi (cascade akan menghapus semua TapLogSesiRelation terkait)
                sesi.delete()
                
                # Pesan sukses
                messages.success(
                    request,
                    f'Sesi untuk {pegawai_nama} '
                    f'({tanggal_mulai.strftime("%d %b %Y")} - {tanggal_selesai.strftime("%d %b %Y")}) '
                    f'berhasil dihapus beserta {jumlah_tap_logs} tap logs terkait.'
                )
                
                return redirect('riwayat_sesi_absensi')
                
        except Exception as e:
            messages.error(
                request,
                f'Gagal menghapus sesi absensi: {str(e)}'
            )
            return redirect('hapus_sesi_absensi', pk=pk)
    
    # GET request - tampilkan halaman konfirmasi
    context = {
        'sesi': sesi,
    }
    
    return render(request, 'absensi_app/sesi/hapus_sesi_absensi.html', context)

@login_required
def api_hapus_sesi(request, pk):
    """
    API endpoint untuk menghapus sesi absensi via AJAX
    """
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "message": "Akses ditolak"}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({"status": "error", "message": "Method tidak diizinkan"}, status=405)
    
    try:
        from .models import AbsensiSesi
        
        sesi = get_object_or_404(AbsensiSesi, pk=pk)
        
        with transaction.atomic():
            pegawai_nama = sesi.pegawai.nama_lengkap
            tanggal_mulai = sesi.tanggal_mulai
            tanggal_selesai = sesi.tanggal_selesai
            jumlah_tap_logs = sesi.tap_relation.count()
            
            sesi.delete()
            
            return JsonResponse({
                "status": "success",
                "message": f'Sesi untuk {pegawai_nama} ({tanggal_mulai.strftime("%d %b %Y")} - {tanggal_selesai.strftime("%d %b %Y")}) berhasil dihapus beserta {jumlah_tap_logs} tap logs terkait.'
            })
            
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": f'Gagal menghapus sesi: {str(e)}'
        }, status=500)

@login_required
def api_get_mode_pegawai_list(request, pk):
    """API: Ambil daftar pegawai yang menggunakan mode ini"""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        from .models import MasterModeJamKerja, Pegawai
        
        mode = get_object_or_404(MasterModeJamKerja, pk=pk)
        
        # Ambil pegawai yang menggunakan mode ini
        pegawai_list = Pegawai.objects.filter(
            mode_assignments__mode=mode,
            mode_assignments__is_active=True,
            is_active=True
        ).select_related('departemen', 'jabatan').distinct().order_by('userid')
        
        pegawai_data = []
        for p in pegawai_list:
            pegawai_data.append({
                'id': p.id,
                'userid': p.userid,
                'nama': p.nama_lengkap,
                'departemen': p.departemen.nama if p.departemen else None,
                'jabatan': p.jabatan.nama if p.jabatan else None
            })
        
        return JsonResponse({
            "status": "success",
            "pegawai": pegawai_data,
            "total": len(pegawai_data)
        })
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "msg": str(e)
        }, status=500)


@login_required
def api_get_mode_periode_list(request, pk):
    """API: Ambil daftar periode untuk mode ini"""
    if not request.user.is_staff:
        return JsonResponse({"status": "error", "msg": "Akses ditolak"}, status=403)
    
    try:
        from .models import MasterModeJamKerja, ModeJamKerjaPeriode
        
        mode = get_object_or_404(MasterModeJamKerja, pk=pk)
        
        periode_list = mode.periode_list.filter(
            is_active=True
        ).order_by('-tanggal_mulai')
        
        periode_data = []
        today = date.today()
        
        for p in periode_list:
            is_active_today = p.tanggal_mulai <= today <= p.tanggal_selesai
            
            periode_data.append({
                'id': p.id,
                'nama': p.nama,
                'tanggal_mulai': p.tanggal_mulai.strftime('%d %b %Y'),
                'tanggal_selesai': p.tanggal_selesai.strftime('%d %b %Y'),
                'is_active_today': is_active_today,
                'catatan': p.catatan
            })
        
        return JsonResponse({
            "status": "success",
            "periode": periode_data,
            "total": len(periode_data)
        })
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "msg": str(e)
        }, status=500)